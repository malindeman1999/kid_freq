from __future__ import annotations

import copy
import queue
import re
import shutil
import threading
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, List, Optional, Sequence

import matplotlib.pyplot as plt
import numpy as np
import tkinter as tk
from openpyxl import Workbook, load_workbook
from matplotlib import colors as mcolors
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.figure import Figure
from matplotlib.ticker import FuncFormatter
from scipy import stats
from tkinter import filedialog, messagebox, simpledialog, scrolledtext, ttk

try:
    import winsound
except Exception:  # pragma: no cover - non-Windows fallback
    winsound = None

from analysis_gui_support.analysis_io import (
    DATASETS_DIR,
    DEFAULT_DATASET_FILE,
    _dataset_dir,
    _dataset_pickle_path,
    _load_dataset,
    _load_vna_file,
    _load_vna_npy_mhz_db_deg,
    _try_load_vna_npy_pair,
    _read_app_state,
    _safe_name,
    _save_dataset,
    _write_app_state,
)
from analysis_gui_support.analysis_models import (
    Dataset,
    VNAScan,
    _current_user,
    _make_event,
    _read_polar_series,
)
from analysis_gui_support.gui_mixins.analysis_selector_plot_mixin import AnalysisSelectorPlotMixin
from analysis_gui_support.gui_mixins.baseline_filter_mixin import BaselineFilterMixin
from analysis_gui_support.gui_mixins.dsdf_convolution_mixin import DSDFConvolutionMixin
from analysis_gui_support.gui_mixins.gaussian_convolution_mixin import GaussianConvolutionMixin
from analysis_gui_support.gui_mixins.interpolation_smooth_mixin import InterpolationSmoothMixin
from analysis_gui_support.gui_mixins.normalization_mixin import NormalizationMixin
from analysis_gui_support.gui_mixins.phase_correction2_mixin import PhaseCorrection2Mixin
from analysis_gui_support.gui_mixins.resonance_selection_mixin import ResonanceSelectionMixin
from analysis_gui_support.gui_mixins.third_phase_correction_mixin import ThirdPhaseCorrectionMixin
from analysis_gui_support.gui_mixins.synthetic_generator_mixin import SyntheticGeneratorMixin
from analysis_gui_support.gui_mixins.unwrap_phase_mixin import UnwrapPhaseMixin

class DataAnalysisGUI(
    AnalysisSelectorPlotMixin,
    BaselineFilterMixin,
    DSDFConvolutionMixin,
    GaussianConvolutionMixin,
    InterpolationSmoothMixin,
    NormalizationMixin,
    PhaseCorrection2Mixin,
    ResonanceSelectionMixin,
    ThirdPhaseCorrectionMixin,
    SyntheticGeneratorMixin,
    UnwrapPhaseMixin,
):
    def _dataset_res_neighbor_initial_date(self) -> str:
        return str(getattr(self.dataset, "res_neighbor_initial_date", "") or "")

    def _sync_res_neighbor_initial_date(self, *, autosave: bool = False) -> None:
        if self.res_neighbor_dfrel_initial_date_var is None:
            return
        new_value = str(self.res_neighbor_dfrel_initial_date_var.get()).strip()
        if new_value == self._dataset_res_neighbor_initial_date():
            return
        self.dataset.res_neighbor_initial_date = new_value
        self._mark_dirty()
        self._refresh_status()
        if autosave:
            self._autosave_dataset()

    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("VNA Data Analysis")
        self.dataset_path = _read_app_state().resolve()
        self.dataset = _load_dataset(self.dataset_path)
        self._scrollable_plot_hosts: Dict[str, Dict[str, object]] = {}

        self.dataset_meta_var = tk.StringVar()
        self.dataset_label_var = tk.StringVar()
        self.scan_count_var = tk.StringVar()
        self.selection_var = tk.StringVar()
        self.saved_var = tk.StringVar()
        self.update_dates_mode_var = tk.StringVar(value="dir")
        self.synth_button: Optional[tk.Button] = None
        self.select_scans_button: Optional[tk.Button] = None
        self.unwrap_button: Optional[tk.Button] = None
        self.phase2_button: Optional[tk.Button] = None
        self.phase3_button: Optional[tk.Button] = None
        self.baseline_button: Optional[tk.Button] = None
        self.interp_button: Optional[tk.Button] = None
        self.norm_apply_large_button: Optional[tk.Button] = None
        self._dirty = False
        self.baseline_window: Optional[tk.Toplevel] = None
        self.baseline_canvas: Optional[FigureCanvasTkAgg] = None
        self.baseline_toolbar: Optional[NavigationToolbar2Tk] = None
        self.baseline_figure: Optional[Figure] = None
        self.width_slider: Optional[tk.Scale] = None
        self.step_slider: Optional[tk.Scale] = None
        self.low_slope_slider: Optional[tk.Scale] = None
        self.retain_slider: Optional[tk.Scale] = None
        self.center_slider: Optional[tk.Scale] = None
        self._baseline_after_id: Optional[str] = None
        self._baseline_preview_results: Dict[str, Dict[str, np.ndarray]] = {}
        self.baseline_status_var: Optional[tk.StringVar] = None
        self.baseline_progress: Optional[ttk.Progressbar] = None
        self.baseline_attach_button: Optional[tk.Button] = None
        self._baseline_worker_thread: Optional[threading.Thread] = None
        self._baseline_worker_queue: "queue.Queue[dict]" = queue.Queue()
        self._baseline_compute_running: bool = False
        self._baseline_recompute_pending: bool = False
        self._baseline_compute_context: Dict[str, object] = {}
        self._baseline_target_scan_keys_override: Optional[set[str]] = None
        self.interp_window: Optional[tk.Toplevel] = None
        self.interp_canvas: Optional[FigureCanvasTkAgg] = None
        self.interp_toolbar: Optional[NavigationToolbar2Tk] = None
        self.interp_figure: Optional[Figure] = None
        self.interp_status_var: Optional[tk.StringVar] = None
        self.interp_smooth_slider: Optional[tk.Scale] = None
        self.interp_attach_button: Optional[tk.Button] = None
        self.interp_preview: Dict[str, Dict[str, np.ndarray]] = {}
        self.norm_button: Optional[tk.Button] = None
        self.norm_window: Optional[tk.Toplevel] = None
        self.norm_canvas: Optional[FigureCanvasTkAgg] = None
        self.norm_toolbar: Optional[NavigationToolbar2Tk] = None
        self.norm_figure: Optional[Figure] = None
        self.norm_status_var: Optional[tk.StringVar] = None
        self.norm_attach_button: Optional[tk.Button] = None
        self.norm_preview: Dict[str, Dict[str, np.ndarray]] = {}
        self.gauss_button: Optional[tk.Button] = None
        self.gauss_window: Optional[tk.Toplevel] = None
        self.gauss_canvas: Optional[FigureCanvasTkAgg] = None
        self.gauss_toolbar: Optional[NavigationToolbar2Tk] = None
        self.gauss_figure: Optional[Figure] = None
        self.gauss_slider: Optional[tk.Scale] = None
        self.gauss_threshold_slider: Optional[tk.Scale] = None
        self.gauss_min_region_slider: Optional[tk.Scale] = None
        self.gauss_auto_y_var: Optional[tk.BooleanVar] = None
        self.gauss_status_var: Optional[tk.StringVar] = None
        self.gauss_attach_button: Optional[tk.Button] = None
        self.gauss_preview: Dict[str, Dict[str, np.ndarray]] = {}
        self.dsdf_button: Optional[tk.Button] = None
        self.dsdf_window: Optional[tk.Toplevel] = None
        self.dsdf_canvas: Optional[FigureCanvasTkAgg] = None
        self.dsdf_toolbar: Optional[NavigationToolbar2Tk] = None
        self.dsdf_figure: Optional[Figure] = None
        self.dsdf_fwhm_slider: Optional[tk.Scale] = None
        self.dsdf_threshold_slider: Optional[tk.Scale] = None
        self.dsdf_min_region_slider: Optional[tk.Scale] = None
        self.dsdf_auto_y_var: Optional[tk.BooleanVar] = None
        self.dsdf_show_phase_context_var: Optional[tk.BooleanVar] = None
        self.dsdf_use_corrected_context_var: Optional[tk.BooleanVar] = None
        self.dsdf_status_var: Optional[tk.StringVar] = None
        self.dsdf_attach_button: Optional[tk.Button] = None
        self.dsdf_preview: Dict[str, Dict[str, np.ndarray]] = {}
        self.synth_window: Optional[tk.Toplevel] = None
        self.synth_canvas: Optional[FigureCanvasTkAgg] = None
        self.synth_toolbar: Optional[NavigationToolbar2Tk] = None
        self.synth_figure: Optional[Figure] = None
        self.synth_source_var: Optional[tk.StringVar] = None
        self.synth_status_var: Optional[tk.StringVar] = None
        self.synth_auto_y_var: Optional[tk.BooleanVar] = None
        self.synth_generate_button: Optional[tk.Button] = None
        self.synth_num_res_slider: Optional[tk.Scale] = None
        self.synth_freq_offset_slider: Optional[tk.Scale] = None
        self.synth_num_files_slider: Optional[tk.Scale] = None
        self.synth_qc_slider: Optional[tk.Scale] = None
        self.synth_qi_slider: Optional[tk.Scale] = None
        self.synth_source_path = None
        self.synth_freq = None
        self.synth_preview_files: List[np.ndarray] = []
        self.synth_amp_ax = None
        self.synth_iq_ax = None
        self.unwrap_window: Optional[tk.Toplevel] = None
        self.unwrap_canvas: Optional[FigureCanvasTkAgg] = None
        self.unwrap_toolbar: Optional[NavigationToolbar2Tk] = None
        self.unwrap_figure: Optional[Figure] = None
        self.unwrap_threshold_slider: Optional[tk.Scale] = None
        self.unwrap_max_passes_slider: Optional[tk.Scale] = None
        self.unwrap_min_sep_slider: Optional[tk.Scale] = None
        self.unwrap_apply_exact_360_var: Optional[tk.BooleanVar] = None
        self.unwrap_correct_congruent_var: Optional[tk.BooleanVar] = None
        self.unwrap_p_random_var: Optional[tk.StringVar] = None
        self.unwrap_p_random_entry: Optional[tk.Entry] = None
        self.unwrap_auto_y_var: Optional[tk.BooleanVar] = None
        self.unwrap_mod360_var: Optional[tk.BooleanVar] = None
        self.unwrap_status_var: Optional[tk.StringVar] = None
        self.unwrap_progress: Optional[ttk.Progressbar] = None
        self.unwrap_update_button: Optional[tk.Button] = None
        self.unwrap_attach_button: Optional[tk.Button] = None
        self.unwrap_preview: Dict[str, Dict[str, np.ndarray]] = {}
        self.phase2_window: Optional[tk.Toplevel] = None
        self.phase2_canvas: Optional[FigureCanvasTkAgg] = None
        self.phase2_toolbar: Optional[NavigationToolbar2Tk] = None
        self.phase2_figure: Optional[Figure] = None
        self.phase2_auto_y_var: Optional[tk.BooleanVar] = None
        self.phase2_mod360_var: Optional[tk.BooleanVar] = None
        self.phase2_status_var: Optional[tk.StringVar] = None
        self.phase2_progress: Optional[ttk.Progressbar] = None
        self.phase2_update_button: Optional[tk.Button] = None
        self.phase3_window: Optional[tk.Toplevel] = None
        self.phase3_canvas: Optional[FigureCanvasTkAgg] = None
        self.phase3_toolbar: Optional[NavigationToolbar2Tk] = None
        self.phase3_figure: Optional[Figure] = None
        self.phase3_auto_y_var: Optional[tk.BooleanVar] = None
        self.phase3_mod360_var: Optional[tk.BooleanVar] = None
        self.phase3_status_var: Optional[tk.StringVar] = None
        self.res_button: Optional[tk.Button] = None
        self.res_window: Optional[tk.Toplevel] = None
        self.res_canvas: Optional[FigureCanvasTkAgg] = None
        self.res_toolbar: Optional[NavigationToolbar2Tk] = None
        self.res_figure: Optional[Figure] = None
        self.res_fit_button: Optional[tk.Button] = None
        self.res_status_var: Optional[tk.StringVar] = None
        self.res_fr_var: Optional[tk.StringVar] = None
        self.res_qi_var: Optional[tk.StringVar] = None
        self.res_qc_var: Optional[tk.StringVar] = None
        self.res_qc_phase_var: Optional[tk.StringVar] = None
        self.res_a_mag_var: Optional[tk.StringVar] = None
        self.res_a_phase_var: Optional[tk.StringVar] = None
        self.res_tau_var: Optional[tk.StringVar] = None
        self.res_fix_fr_var: Optional[tk.BooleanVar] = None
        self.res_fix_qi_var: Optional[tk.BooleanVar] = None
        self.res_fix_qc_var: Optional[tk.BooleanVar] = None
        self.res_fix_qc_phase_var: Optional[tk.BooleanVar] = None
        self.res_fix_a_mag_var: Optional[tk.BooleanVar] = None
        self.res_fix_a_phase_var: Optional[tk.BooleanVar] = None
        self.res_fix_tau_var: Optional[tk.BooleanVar] = None
        self.res_auto_y_var: Optional[tk.BooleanVar] = None
        self.res_display_mode_var: Optional[tk.StringVar] = None
        self.res_fit_mode_var: Optional[tk.StringVar] = None
        self.res_span_selector = None
        self.res_amp_ax = None
        self.res_iq_ax = None
        self.res_model_preview: Optional[dict] = None
        self.plot_scans_window: Optional[tk.Toplevel] = None
        self.plot_scans_canvas: Optional[FigureCanvasTkAgg] = None
        self.plot_scans_toolbar: Optional[NavigationToolbar2Tk] = None
        self.plot_scans_figure: Optional[Figure] = None
        self.plot_scans_show_amp_var: Optional[tk.BooleanVar] = None
        self.plot_scans_show_phase_var: Optional[tk.BooleanVar] = None
        self.plot_scans_group_var: Optional[tk.BooleanVar] = None
        self.plot_scans_data_mode_var: Optional[tk.StringVar] = None
        self.plot_scans_raw_radio: Optional[tk.Radiobutton] = None
        self.plot_scans_normalized_radio: Optional[tk.Radiobutton] = None
        self.plot_scans_show_gaussian_var: Optional[tk.BooleanVar] = None
        self.plot_scans_show_dsdf_var: Optional[tk.BooleanVar] = None
        self.plot_scans_show_2pi_var: Optional[tk.BooleanVar] = None
        self.plot_scans_show_vna_phase_var: Optional[tk.BooleanVar] = None
        self.plot_scans_show_other_phase_var: Optional[tk.BooleanVar] = None
        self.plot_scans_show_attached_res_var: Optional[tk.BooleanVar] = None
        self.plot_scans_auto_y_var: Optional[tk.BooleanVar] = None
        self.plot_scans_use_unwrapped_phase_var: Optional[tk.BooleanVar] = None
        self.plot_scans_status_var: Optional[tk.StringVar] = None
        self.scan_evolution_window: Optional[tk.Toplevel] = None
        self.scan_evolution_canvas: Optional[FigureCanvasTkAgg] = None
        self.scan_evolution_toolbar: Optional[NavigationToolbar2Tk] = None
        self.scan_evolution_figure: Optional[Figure] = None
        self.scan_evolution_status_var: Optional[tk.StringVar] = None
        self.scan_evolution_mod360_var: Optional[tk.BooleanVar] = None
        self.scan_evolution_show_gaussian_var: Optional[tk.BooleanVar] = None
        self.scan_evolution_show_dsdf_var: Optional[tk.BooleanVar] = None
        self.scan_evolution_show_phase_2pi_var: Optional[tk.BooleanVar] = None
        self.scan_evolution_show_phase_vna_var: Optional[tk.BooleanVar] = None
        self.scan_evolution_show_phase_other_var: Optional[tk.BooleanVar] = None
        self.scan_evolution_show_attached_res_var: Optional[tk.BooleanVar] = None
        self._scan_evolution_scan_key: Optional[str] = None
        self._scan_evolution_stage_rows: List[dict] = []
        self._scan_evolution_axes_rows: List[tuple[object, object, object]] = []
        self._scan_evolution_syncing_xlim: bool = False
        self.res_shift_corr_window: Optional[tk.Toplevel] = None
        self.res_shift_corr_canvas: Optional[FigureCanvasTkAgg] = None
        self.res_shift_corr_toolbar: Optional[NavigationToolbar2Tk] = None
        self.res_shift_corr_figure: Optional[Figure] = None
        self.res_shift_corr_status_var: Optional[tk.StringVar] = None
        self._res_shift_corr_axes: tuple[object, object] | None = None
        self.res_pair_dfdiff_hist_window: Optional[tk.Toplevel] = None
        self.res_pair_dfdiff_hist_canvas: Optional[FigureCanvasTkAgg] = None
        self.res_pair_dfdiff_hist_toolbar: Optional[NavigationToolbar2Tk] = None
        self.res_pair_dfdiff_hist_figure: Optional[Figure] = None
        self.res_pair_dfdiff_hist_status_var: Optional[tk.StringVar] = None
        self.res_pair_dfdiff_hist_sep_mhz_var: Optional[tk.DoubleVar] = None
        self.res_pair_dfdiff_hist_bin_mhz_var: Optional[tk.DoubleVar] = None
        self.res_pair_dfdiff_hist_capture_var: Optional[tk.DoubleVar] = None
        self.res_pair_dfdiff_hist_fit_mode_var: Optional[tk.StringVar] = None
        self.res_pair_dfdiff_hist_num_res_var: Optional[tk.IntVar] = None
        self.res_pair_dfdiff_hist_center_ghz_var: Optional[tk.DoubleVar] = None
        self.res_pair_dfdiff_hist_dfrel_var: Optional[tk.DoubleVar] = None
        self.res_pair_dfdiff_hist_freq_jitter_khz_var: Optional[tk.DoubleVar] = None
        self.res_pair_dfdiff_hist_sep_scale: Optional[tk.Scale] = None
        self.res_pair_dfdiff_hist_bin_scale: Optional[tk.Scale] = None
        self.res_pair_dfdiff_hist_capture_scale: Optional[tk.Scale] = None
        self.res_pair_dfdiff_hist_freq_jitter_scale: Optional[tk.Scale] = None
        self._res_pair_dfdiff_hist_ax = None
        self.res_neighbor_dfrel_window: Optional[tk.Toplevel] = None
        self.res_neighbor_dfrel_canvas: Optional[FigureCanvasTkAgg] = None
        self.res_neighbor_dfrel_toolbar: Optional[NavigationToolbar2Tk] = None
        self.res_neighbor_dfrel_figure: Optional[Figure] = None
        self.res_neighbor_dfrel_status_var: Optional[tk.StringVar] = None
        self.res_neighbor_dfrel_sep_rel_var: Optional[tk.DoubleVar] = None
        self.res_neighbor_dfrel_show_iqr_var: Optional[tk.BooleanVar] = None
        self.res_neighbor_dfrel_mode_var: Optional[tk.StringVar] = None
        self.res_neighbor_dfrel_initial_date_var: Optional[tk.StringVar] = None
        self.res_neighbor_dfrel_sep_scale: Optional[tk.Scale] = None
        self._res_neighbor_dfrel_ax = None
        self.res_neighbor_corr_window: Optional[tk.Toplevel] = None
        self.res_neighbor_corr_canvas: Optional[FigureCanvasTkAgg] = None
        self.res_neighbor_corr_toolbar: Optional[NavigationToolbar2Tk] = None
        self.res_neighbor_corr_figure: Optional[Figure] = None
        self.res_neighbor_corr_status_var: Optional[tk.StringVar] = None
        self.res_neighbor_corr_sep_rel_var: Optional[tk.DoubleVar] = None
        self.res_neighbor_corr_initial_date_var: Optional[tk.StringVar] = None
        self.res_neighbor_corr_show_curves_var: Optional[tk.BooleanVar] = None
        self.res_neighbor_corr_sep_scale: Optional[tk.Scale] = None
        self._res_neighbor_corr_axes: tuple[object, object] | None = None
        self.res_neighbor_scan_window: Optional[tk.Toplevel] = None
        self.res_neighbor_scan_canvas: Optional[FigureCanvasTkAgg] = None
        self.res_neighbor_scan_toolbar: Optional[NavigationToolbar2Tk] = None
        self.res_neighbor_scan_figure: Optional[Figure] = None
        self.res_neighbor_scan_status_var: Optional[tk.StringVar] = None
        self._res_neighbor_scan_source: str = "dfrel"
        self._res_neighbor_scan_ax = None
        self.attached_res_edit_window: Optional[tk.Toplevel] = None
        self.attached_res_edit_canvas: Optional[FigureCanvasTkAgg] = None
        self.attached_res_edit_toolbar: Optional[NavigationToolbar2Tk] = None
        self.attached_res_edit_figure: Optional[Figure] = None
        self.attached_res_edit_status_var: Optional[tk.StringVar] = None
        self.attached_res_edit_working_number_var: Optional[tk.StringVar] = None
        self.attached_res_edit_working_number_spinbox: Optional[tk.Spinbox] = None
        self.attached_res_edit_spacing_var: Optional[tk.DoubleVar] = None
        self.attached_res_edit_spacing_scale: Optional[tk.Scale] = None
        self.attached_res_edit_search_window_khz_var: Optional[tk.DoubleVar] = None
        self.attached_res_edit_search_window_scale: Optional[tk.Scale] = None
        self.attached_res_edit_truncate_var: Optional[tk.BooleanVar] = None
        self.attached_res_edit_truncate_threshold_var: Optional[tk.DoubleVar] = None
        self.attached_res_edit_truncate_threshold_scale: Optional[tk.Scale] = None
        self.attached_res_edit_add_button: Optional[tk.Button] = None
        self.attached_res_edit_renumber_button: Optional[tk.Button] = None
        self.attached_res_edit_undo_button: Optional[tk.Button] = None
        self.attached_res_edit_save_button: Optional[tk.Button] = None
        self.attached_res_edit_exit_button: Optional[tk.Button] = None
        self.attached_res_edit_ax = None
        self._attached_res_edit_points: List[dict] = []
        self._attached_res_edit_rows_cache: List[dict] = []
        self._attached_res_edit_offset_by_scan_key: Dict[str, float] = {}
        self._attached_res_edit_selected: Optional[tuple[str, str]] = None
        self._attached_res_edit_pending_add: bool = False
        self._attached_res_edit_default_xlim: Optional[tuple[float, float]] = None
        self._attached_res_edit_missing_normalized_warned: Optional[tuple[str, ...]] = None
        self._attached_res_edit_snapshot: Optional[dict] = None
        self._attached_res_edit_undo_stack: List[dict] = []
        self._attached_res_edit_changed: bool = False
        self._res_scan_key: Optional[str] = None
        self._res_selected_range: Optional[tuple[float, float]] = None
        self._res_manual_ylim: Optional[tuple[float, float]] = None
        self._last_resonance_scan_key: Optional[str] = None
        self._default_button_bg = "SystemButtonFace"
        self._default_button_activebg = "SystemButtonFace"
        self._build_layout()
        if self.synth_button is not None:
            self._default_button_bg = str(self.synth_button.cget("bg"))
            self._default_button_activebg = str(self.synth_button.cget("activebackground"))
        self._reload_transcript_ui()
        self._refresh_status()
        self._log("Application started.")

    def _ensure_scrollable_plot_host(self, key: str, window: tk.Misc) -> tuple[tk.Frame, tk.Frame]:
        host = self._scrollable_plot_hosts.get(key)
        if host is not None:
            return host["toolbar_frame"], host["inner_frame"]

        container = tk.Frame(window)
        container.pack(side="top", fill="both", expand=True)
        toolbar_frame = tk.Frame(container)
        toolbar_frame.pack(side="top", fill="x")
        viewport = tk.Frame(container)
        viewport.pack(side="top", fill="both", expand=True)
        scrollbar = ttk.Scrollbar(viewport, orient="vertical")
        scrollbar.pack(side="right", fill="y")
        scroll_canvas = tk.Canvas(viewport, highlightthickness=0, yscrollcommand=scrollbar.set)
        scroll_canvas.pack(side="left", fill="both", expand=True)
        inner_frame = tk.Frame(scroll_canvas)
        inner_window = scroll_canvas.create_window((0, 0), window=inner_frame, anchor="nw")
        scrollbar.configure(command=scroll_canvas.yview)

        def _sync_scrollregion(_event=None) -> None:
            scroll_canvas.configure(scrollregion=scroll_canvas.bbox("all"))

        def _sync_inner_width(event) -> None:
            scroll_canvas.itemconfigure(inner_window, width=event.width)

        def _on_mousewheel(event) -> str:
            delta = int(getattr(event, "delta", 0))
            if delta == 0:
                return "break"
            step = -int(delta / 120) if abs(delta) >= 120 else (-1 if delta > 0 else 1)
            scroll_canvas.yview_scroll(step, "units")
            return "break"

        inner_frame.bind("<Configure>", _sync_scrollregion)
        scroll_canvas.bind("<Configure>", _sync_inner_width)
        scroll_canvas.bind("<MouseWheel>", _on_mousewheel)
        inner_frame.bind("<MouseWheel>", _on_mousewheel)

        self._scrollable_plot_hosts[key] = {
            "container": container,
            "toolbar_frame": toolbar_frame,
            "scroll_canvas": scroll_canvas,
            "inner_frame": inner_frame,
        }
        return toolbar_frame, inner_frame

    def _destroy_scrollable_plot_host(self, key: str) -> None:
        host = self._scrollable_plot_hosts.pop(key, None)
        if host is None:
            return
        container = host.get("container")
        if isinstance(container, tk.Misc) and container.winfo_exists():
            container.destroy()

    def _set_scrollable_figure_size(
        self,
        key: str,
        figure: Optional[Figure],
        *,
        canvas_agg: Optional[FigureCanvasTkAgg] = None,
        width_in: float,
        row_count: int,
        row_height_in: float,
        min_height_in: float,
    ) -> None:
        if figure is None:
            return
        height_in = max(min_height_in, row_count * row_height_in)
        figure.set_size_inches(width_in, height_in, forward=True)
        if canvas_agg is not None:
            widget = canvas_agg.get_tk_widget()
            dpi = float(figure.get_dpi()) if figure.get_dpi() else 100.0
            widget.configure(
                width=max(1, int(round(width_in * dpi))),
                height=max(1, int(round(height_in * dpi))),
            )
        host = self._scrollable_plot_hosts.get(key)
        if host is None:
            return
        scroll_canvas = host.get("scroll_canvas")
        if isinstance(scroll_canvas, tk.Canvas) and scroll_canvas.winfo_exists():
            scroll_canvas.update_idletasks()
            scroll_canvas.configure(scrollregion=scroll_canvas.bbox("all"))

    @staticmethod
    def _scan_key(scan: VNAScan) -> str:
        return f"{scan.filename}|{scan.loaded_at}"

    def _prune_selected_scan_keys(self) -> None:
        valid_keys = {self._scan_key(scan) for scan in self.dataset.vna_scans}
        self.dataset.selected_scan_keys = [
            key for key in self.dataset.selected_scan_keys if key in valid_keys
        ]

    def _selected_scans(self) -> List[VNAScan]:
        selected_keys = set(self.dataset.selected_scan_keys)
        return [scan for scan in self.dataset.vna_scans if self._scan_key(scan) in selected_keys]

    def _build_layout(self) -> None:
        frame = tk.Frame(self.root, padx=10, pady=10)
        frame.pack(fill="both", expand=True)

        top = tk.Frame(frame)
        top.pack(side="top", fill="x")
        tk.Label(top, textvariable=self.dataset_meta_var, anchor="w", justify="left").pack(
            anchor="w"
        )
        tk.Label(top, text="Active dataset file:", anchor="w", justify="left").pack(
            anchor="w", pady=(6, 0)
        )
        tk.Label(
            top, textvariable=self.dataset_label_var, wraplength=900, justify="left", anchor="w"
        ).pack(anchor="w")
        tk.Label(top, textvariable=self.scan_count_var, anchor="w", justify="left").pack(
            anchor="w", pady=(8, 0)
        )
        tk.Label(top, textvariable=self.selection_var, anchor="w", justify="left").pack(
            anchor="w", pady=(2, 0)
        )
        tk.Label(top, textvariable=self.saved_var, anchor="w", justify="left").pack(
            anchor="w", pady=(2, 10)
        )

        bottom = tk.Frame(frame)
        bottom.pack(side="top", fill="both", expand=True)
        left = tk.Frame(bottom)
        left.pack(side="left", fill="y")
        right = tk.Frame(bottom, padx=12)
        right.pack(side="left", fill="both", expand=True)
        button_area = tk.Frame(left)
        button_area.pack(anchor="w")
        button_col1 = tk.Frame(button_area)
        button_col1.pack(side="left", anchor="n")
        button_col2 = tk.Frame(button_area, padx=8)
        button_col2.pack(side="left", anchor="n")

        button_width = 24
        left_button_specs: list[dict[str, object]] = []
        right_button_specs: list[dict[str, object]] = []

        left_button_specs.append({"text": "New Dataset", "command": self.start_new_dataset})
        left_button_specs.append({"text": "Rename Dataset", "command": self.rename_dataset_prefix})
        self.synth_button = tk.Button(
            button_col1, text="Generate Synthetic Data", width=button_width, command=self.open_synthetic_generator_window
        )
        left_button_specs.append({"button": self.synth_button})
        left_button_specs.append({"text": "Load Different Dataset", "command": self.load_different_dataset})
        left_button_specs.append({"text": "Load VNA Scan(s)", "command": self.load_vna_scan})
        left_button_specs.append({"text": "Remove VNA Scan(s)", "command": self.remove_vna_scans})
        left_button_specs.append({"text": "Load Resonators From Sheet", "command": self.open_resonance_sheet_loader})
        left_button_specs.append({"text": "Save Resonators To Sheet", "command": self.open_resonance_sheet_saver})
        self.select_scans_button = tk.Button(
            button_col1, text="Select Scans for Analysis", width=button_width, command=self.open_analysis_selector
        )
        left_button_specs.append({"button": self.select_scans_button})
        left_button_specs.append({"text": "Group Selected Scans", "command": self.group_selected_scans_for_plotting})
        left_button_specs.append({"text": "Plot Selected VNA Scans", "command": self.plot_selected_vna_scans})
        self.unwrap_button = tk.Button(
            button_col1, text="Phase Correction 1", width=button_width, command=self.open_unwrap_phase_window
        )
        left_button_specs.append({"button": self.unwrap_button})
        self.phase2_button = tk.Button(
            button_col1, text="Phase Correction 2", width=button_width, command=self.open_second_phase_correction_window
        )
        left_button_specs.append({"button": self.phase2_button})
        self.phase3_button = tk.Button(
            button_col1, text="Phase Correction 3", width=button_width, command=self.open_third_phase_correction_window
        )
        left_button_specs.append({"button": self.phase3_button})
        self.baseline_button = tk.Button(
            button_col1, text="Baseline Filtering", width=button_width, command=self.open_baseline_filter_window
        )
        left_button_specs.append({"button": self.baseline_button})
        self.interp_button = tk.Button(
            button_col1, text="Interp + Smooth", width=button_width, command=self.open_interp_smooth_window
        )
        left_button_specs.append({"button": self.interp_button})
        self.norm_button = tk.Button(
            button_col1, text="Normalize Baseline", width=button_width, command=self.open_normalization_window
        )
        left_button_specs.append({"button": self.norm_button})
        self.norm_apply_large_button = tk.Button(
            button_col1, text="Apply Large-Scan Baseline", width=button_width, command=self.apply_large_scan_baseline_to_selected
        )
        left_button_specs.append({"button": self.norm_apply_large_button})
        left_button_specs.append({"text": "Scan Evolution", "command": self.open_scan_evolution_window})

        self.gauss_button = tk.Button(
            button_col2, text="Gaussian Convolve |S21|", width=button_width, command=self.open_gaussian_convolution_window
        )
        right_button_specs.append({"button": self.gauss_button})
        self.dsdf_button = tk.Button(
            button_col2, text="Gaussian Convolve |dS21/df|", width=button_width, command=self.open_dsdf_convolution_window
        )
        right_button_specs.append({"button": self.dsdf_button})
        self.res_button = tk.Button(
            button_col2, text="Resonance Fitting", width=button_width, command=self.open_resonance_selection_window
        )
        right_button_specs.append({"button": self.res_button})
        right_button_specs.append({"text": "Mark Res. on Sel. Scans", "command": self.open_attached_resonance_editor})
        right_button_specs.append(
            {
                "text": "Reset Selected Scans",
                "command": self.clear_selected_scan_attachments,
            }
        )
        right_button_specs.append({"text": "Plot Resonator Markers", "command": self.open_attached_resonance_plotter})
        right_button_specs.append({"text": "Update Dates From Path", "command": self.open_update_dates_dialog})
        right_button_specs.append({"text": "Reorder Scans By Date", "command": self.reorder_vna_scans_by_date})
        right_button_specs.append({"text": "Pair df/f vs Time", "command": self.open_resonator_neighbor_dfrel_window})
        right_button_specs.append({"text": "Pair Self Corr. vs Time", "command": self.open_resonator_neighbor_corr_window})
        right_button_specs.append({"text": "Shift Correl. Between Freqs.", "command": self.open_resonator_shift_correlation_window})
        right_button_specs.append({"text": "Histogram df2-df1", "command": self.open_resonator_pair_dfdiff_hist_window})

        for parent, specs in ((button_col1, left_button_specs), (button_col2, right_button_specs)):
            for spec in specs:
                button = spec.get("button")
                if not isinstance(button, tk.Button):
                    button = tk.Button(
                        parent,
                        text=str(spec["text"]),
                        width=button_width,
                        command=spec["command"],
                    )
                button.pack(anchor="w", pady=spec.get("pady", 2))

        tk.Label(right, text="Transcript:", anchor="w", justify="left").pack(anchor="w", pady=(0, 2))
        self.log_text = scrolledtext.ScrolledText(right, width=110, height=10, state="disabled")
        self.log_text.pack(fill="both", expand=True)

    def _append_transcript_line(self, timestamp: str, message: str) -> None:
        self.log_text.configure(state="normal")
        self.log_text.insert("end", f"[{timestamp}] {message}\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _choose_one_selected_scan(self) -> Optional[VNAScan]:
        scans = self._selected_scans()
        if not scans:
            messagebox.showwarning("No selection", "Select scans for analysis first.")
            return None
        if len(scans) == 1:
            return scans[0]

        dialog = tk.Toplevel(self.root)
        dialog.title("Choose Scan")
        dialog.geometry("900x360")
        dialog.transient(self.root)
        dialog.grab_set()

        tk.Label(dialog, text="Choose one of the currently selected scans:").pack(anchor="w", padx=10, pady=(10, 4))
        listbox = tk.Listbox(dialog, width=130, height=12, selectmode=tk.SINGLE)
        listbox.pack(fill="both", expand=True, padx=10)
        for idx, scan in enumerate(scans):
            listbox.insert(
                idx,
                self._scan_dialog_label(
                    scan,
                    include_file_timestamp=True,
                    include_group=True,
                ),
            )
        listbox.selection_set(0)
        listbox.focus_set()

        chosen: dict[str, Optional[VNAScan]] = {"scan": None}

        def accept() -> None:
            selection = listbox.curselection()
            if not selection:
                messagebox.showwarning("No scan selected", "Choose one scan.", parent=dialog)
                return
            chosen["scan"] = scans[int(selection[0])]
            dialog.destroy()

        btns = tk.Frame(dialog)
        btns.pack(fill="x", padx=10, pady=10)
        tk.Button(btns, text="Cancel", width=12, command=dialog.destroy).pack(side="right")
        tk.Button(btns, text="Open", width=12, command=accept).pack(side="right", padx=(0, 8))
        listbox.bind("<Double-Button-1>", lambda _e: accept())
        dialog.wait_window()
        return chosen["scan"]

    def _scan_evolution_make_stage(
        self,
        *,
        name: str,
        freq_hz: np.ndarray,
        amp: np.ndarray,
        phase_deg: np.ndarray,
    ) -> Optional[dict]:
        freq = np.asarray(freq_hz, dtype=float)
        amp_arr = np.asarray(amp, dtype=float)
        phase_arr = np.asarray(phase_deg, dtype=float)
        if freq.ndim != 1 or amp_arr.ndim != 1 or phase_arr.ndim != 1:
            return None
        if freq.size == 0 or amp_arr.shape != freq.shape or phase_arr.shape != freq.shape:
            return None
        order = np.argsort(freq)
        freq = freq[order]
        amp_arr = amp_arr[order]
        phase_arr = phase_arr[order]
        real = amp_arr * np.cos(np.radians(phase_arr))
        imag = amp_arr * np.sin(np.radians(phase_arr))
        return {
            "name": name,
            "freq_hz": freq,
            "freq_ghz": freq / 1.0e9,
            "amp": amp_arr,
            "phase_deg": phase_arr,
            "real": real,
            "imag": imag,
        }

    def _scan_evolution_phase_display(self, stage: dict) -> np.ndarray:
        phase = np.asarray(stage["phase_deg"], dtype=float)
        if self.scan_evolution_mod360_var is not None and self.scan_evolution_mod360_var.get():
            return ((phase + 180.0) % 360.0) - 180.0
        return phase

    @staticmethod
    def _scan_evolution_nearest_values(
        query_freq_hz: Sequence[float],
        ref_freq_hz: Sequence[float],
        ref_values: Sequence[float],
    ) -> tuple[np.ndarray, np.ndarray]:
        query = np.asarray(query_freq_hz, dtype=float).ravel()
        ref_f = np.asarray(ref_freq_hz, dtype=float).ravel()
        ref_v = np.asarray(ref_values, dtype=float).ravel()
        if query.size == 0 or ref_f.size == 0 or ref_v.size == 0 or ref_f.size != ref_v.size:
            return np.asarray([], dtype=float), np.asarray([], dtype=float)
        idx = np.searchsorted(ref_f, query)
        idx = np.clip(idx, 0, max(ref_f.size - 1, 0))
        left = np.clip(idx - 1, 0, max(ref_f.size - 1, 0))
        use_left = np.abs(query - ref_f[left]) <= np.abs(query - ref_f[idx])
        chosen = np.where(use_left, left, idx)
        return ref_f[chosen], ref_v[chosen]

    def _scan_evolution_overlay_points(self, scan: VNAScan) -> dict[str, np.ndarray]:
        cand = scan.candidate_resonators
        gaussian = cand.get("gaussian_convolution", {})
        dsdf = cand.get("dsdf_gaussian_convolution", {})
        phase_points = cand.get("phase_class_points", {})
        return {
            "gaussian": np.asarray(gaussian.get("candidate_freq", np.array([])), dtype=float),
            "dsdf": np.asarray(dsdf.get("candidate_freq", np.array([])), dtype=float),
            "regular": np.asarray(phase_points.get("regular_freqs", np.array([])), dtype=float),
            "congruent": np.asarray(phase_points.get("irregular_congruent_freqs", np.array([])), dtype=float),
            "noncongruent": np.asarray(phase_points.get("irregular_noncongruent_freqs", np.array([])), dtype=float),
        }

    def _scan_evolution_attached_resonator_points(
        self,
        scan: VNAScan,
        *,
        phase_values: np.ndarray,
        amp_values: np.ndarray,
        real_values: np.ndarray,
        imag_values: np.ndarray,
    ) -> tuple[list[dict], list[dict], list[dict]]:
        attached = scan.candidate_resonators.get("sheet_resonances", {})
        assignments = attached.get("assignments") if isinstance(attached, dict) else {}
        if not isinstance(assignments, dict):
            return [], [], []
        ref_freq_hz = np.asarray(scan.freq, dtype=float)
        amp_ref = np.asarray(amp_values, dtype=float)
        phase_ref = np.asarray(phase_values, dtype=float)
        real_ref = np.asarray(real_values, dtype=float)
        imag_ref = np.asarray(imag_values, dtype=float)
        if (
            ref_freq_hz.size == 0
            or amp_ref.shape != ref_freq_hz.shape
            or phase_ref.shape != ref_freq_hz.shape
            or real_ref.shape != ref_freq_hz.shape
            or imag_ref.shape != ref_freq_hz.shape
        ):
            return [], [], []
        amp_points: list[dict] = []
        phase_points: list[dict] = []
        complex_points: list[dict] = []
        for resonator_number, record in assignments.items():
            if not isinstance(record, dict):
                continue
            try:
                target_hz = float(record.get("frequency_hz"))
            except Exception:
                continue
            x_hz_amp, y_amp = self._scan_evolution_nearest_values([target_hz], ref_freq_hz, amp_ref)
            x_hz_phase, y_phase = self._scan_evolution_nearest_values([target_hz], ref_freq_hz, phase_ref)
            x_hz_real, y_real = self._scan_evolution_nearest_values([target_hz], ref_freq_hz, real_ref)
            x_hz_imag, y_imag = self._scan_evolution_nearest_values([target_hz], ref_freq_hz, imag_ref)
            if x_hz_amp.size:
                amp_points.append({"x_hz": float(x_hz_amp[0]), "y": float(y_amp[0]), "label": str(resonator_number)})
            if x_hz_phase.size:
                phase_points.append({"x_hz": float(x_hz_phase[0]), "y": float(y_phase[0]), "label": str(resonator_number)})
            if x_hz_real.size and x_hz_imag.size:
                complex_points.append(
                    {
                        "x_hz": float(x_hz_real[0]),
                        "real": float(y_real[0]),
                        "imag": float(y_imag[0]),
                        "label": str(resonator_number),
                    }
                )
        return amp_points, phase_points, complex_points

    def _scan_evolution_add_overlays(
        self,
        ax,
        scan: VNAScan,
        *,
        values: np.ndarray,
        use_phase: bool,
        used_labels: set[str],
    ) -> None:
        ref_freq_hz = np.asarray(scan.freq, dtype=float)
        ref_y = np.asarray(values, dtype=float)
        marker_defs = [
            (
                self.scan_evolution_show_gaussian_var is not None and bool(self.scan_evolution_show_gaussian_var.get()),
                "gaussian",
                dict(linestyle="none", marker="o", markersize=8, markerfacecolor="none", markeredgewidth=1.5, color="green"),
                "Gaussian candidates",
            ),
            (
                self.scan_evolution_show_dsdf_var is not None and bool(self.scan_evolution_show_dsdf_var.get()),
                "dsdf",
                dict(linestyle="none", marker="D", markersize=6, color="red"),
                "dS21/df peaks",
            ),
            (
                self.scan_evolution_show_phase_2pi_var is not None and bool(self.scan_evolution_show_phase_2pi_var.get()),
                "regular",
                dict(linestyle="none", marker="o", markersize=4, color="black"),
                "2pi phase corrections",
            ),
            (
                self.scan_evolution_show_phase_vna_var is not None and bool(self.scan_evolution_show_phase_vna_var.get()),
                "congruent",
                dict(linestyle="none", marker="o", markersize=5, color="pink"),
                "VNA phase corrections",
            ),
            (
                self.scan_evolution_show_phase_other_var is not None and bool(self.scan_evolution_show_phase_other_var.get()),
                "noncongruent",
                dict(linestyle="none", marker="o", markersize=2.5, color="blue"),
                "Other phase discontinuities",
            ),
        ]
        points = self._scan_evolution_overlay_points(scan)
        for enabled, key, style, label in marker_defs:
            if not enabled:
                continue
            freq_pts = points[key]
            if freq_pts.size == 0:
                continue
            x_hz, y_pts = self._scan_evolution_nearest_values(freq_pts, ref_freq_hz, ref_y)
            if x_hz.size == 0:
                continue
            plot_label = label if label not in used_labels else None
            ax.plot(x_hz / 1.0e9, y_pts, label=plot_label, **style)
            if plot_label is not None:
                used_labels.add(label)

    def _scan_evolution_complex_overlay_points(
        self,
        scan: VNAScan,
        *,
        real_values: np.ndarray,
        imag_values: np.ndarray,
        freq_lo_ghz: float,
        freq_hi_ghz: float,
    ) -> list[tuple[np.ndarray, np.ndarray, dict, str]]:
        ref_freq_hz = np.asarray(scan.freq, dtype=float)
        real_ref = np.asarray(real_values, dtype=float)
        imag_ref = np.asarray(imag_values, dtype=float)
        if ref_freq_hz.size == 0 or real_ref.shape != ref_freq_hz.shape or imag_ref.shape != ref_freq_hz.shape:
            return []

        marker_defs = [
            (
                self.scan_evolution_show_gaussian_var is not None and bool(self.scan_evolution_show_gaussian_var.get()),
                "gaussian",
                dict(linestyle="none", marker="o", markersize=8, markerfacecolor="none", markeredgewidth=1.5, color="green"),
                "Gaussian candidates",
            ),
            (
                self.scan_evolution_show_dsdf_var is not None and bool(self.scan_evolution_show_dsdf_var.get()),
                "dsdf",
                dict(linestyle="none", marker="D", markersize=6, color="red"),
                "dS21/df peaks",
            ),
            (
                self.scan_evolution_show_phase_2pi_var is not None and bool(self.scan_evolution_show_phase_2pi_var.get()),
                "regular",
                dict(linestyle="none", marker="o", markersize=4, color="black"),
                "2pi phase corrections",
            ),
            (
                self.scan_evolution_show_phase_vna_var is not None and bool(self.scan_evolution_show_phase_vna_var.get()),
                "congruent",
                dict(linestyle="none", marker="o", markersize=5, color="pink"),
                "VNA phase corrections",
            ),
            (
                self.scan_evolution_show_phase_other_var is not None and bool(self.scan_evolution_show_phase_other_var.get()),
                "noncongruent",
                dict(linestyle="none", marker="o", markersize=2.5, color="blue"),
                "Other phase discontinuities",
            ),
        ]
        points = self._scan_evolution_overlay_points(scan)
        plotted: list[tuple[np.ndarray, np.ndarray, dict, str]] = []
        for enabled, key, style, label in marker_defs:
            if not enabled:
                continue
            freq_pts = np.asarray(points.get(key, np.array([])), dtype=float)
            if freq_pts.size == 0:
                continue
            freq_pts_ghz = freq_pts / 1.0e9
            visible_mask = np.isfinite(freq_pts_ghz) & (freq_pts_ghz >= freq_lo_ghz) & (freq_pts_ghz <= freq_hi_ghz)
            freq_visible = freq_pts[visible_mask]
            if freq_visible.size == 0:
                continue
            _x_hz, real_pts = self._scan_evolution_nearest_values(freq_visible, ref_freq_hz, real_ref)
            _x_hz_im, imag_pts = self._scan_evolution_nearest_values(freq_visible, ref_freq_hz, imag_ref)
            if real_pts.size == 0 or imag_pts.size == 0:
                continue
            plotted.append((real_pts, imag_pts, style, label))
        return plotted

    def _scan_evolution_toggle_phase_wrap(self) -> None:
        if self.scan_evolution_figure is None or self.scan_evolution_canvas is None:
            return
        self._render_scan_evolution_window()

    def _scan_evolution_stage_rows_for_scan(self, scan: VNAScan) -> list[dict]:
        stages: list[dict] = []
        raw_stage = self._scan_evolution_make_stage(
            name="Raw",
            freq_hz=scan.freq,
            amp=scan.amplitude(),
            phase_deg=scan.phase_deg_wrapped_raw(),
        )
        if raw_stage is not None:
            stages.append(raw_stage)

        if scan.has_dewrapped_phase():
            stage = self._scan_evolution_make_stage(
                name="Phase Corr. 1",
                freq_hz=scan.freq,
                amp=scan.amplitude(),
                phase_deg=scan.phase_deg_unwrapped(),
            )
            if stage is not None:
                stages.append(stage)

        phase2 = scan.candidate_resonators.get("phase_correction_2", {})
        amp2, phase2_deg = _read_polar_series(
            phase2,
            amplitude_key="corrected_amp",
            phase_key="corrected_phase_deg",
        )
        stage = self._scan_evolution_make_stage(
            name="Phase Corr. 2",
            freq_hz=scan.freq,
            amp=amp2,
            phase_deg=phase2_deg,
        )
        if stage is not None:
            stages.append(stage)

        phase3 = scan.candidate_resonators.get("phase_correction_3", {})
        amp3, phase3_deg = _read_polar_series(
            phase3,
            amplitude_key="corrected_amp",
            phase_key="corrected_phase_deg",
        )
        stage = self._scan_evolution_make_stage(
            name="Phase Corr. 3",
            freq_hz=scan.freq,
            amp=amp3,
            phase_deg=phase3_deg,
        )
        if stage is not None:
            stages.append(stage)

        bf = scan.baseline_filter if isinstance(scan.baseline_filter, dict) else {}
        norm = bf.get("normalized", {}) if isinstance(bf, dict) else {}
        norm_amp, norm_phase = _read_polar_series(
            norm,
            amplitude_key="norm_amp",
            phase_key="norm_phase_deg_unwrapped",
        )
        stage = self._scan_evolution_make_stage(
            name="Normalized",
            freq_hz=scan.freq,
            amp=norm_amp,
            phase_deg=norm_phase,
        )
        if stage is not None:
            stages.append(stage)

        return stages

    def open_scan_evolution_window(self) -> None:
        scan = self._choose_one_selected_scan()
        if scan is None:
            return

        if self.scan_evolution_window is not None and self.scan_evolution_window.winfo_exists():
            self.scan_evolution_window.lift()
            self._scan_evolution_scan_key = self._scan_key(scan)
            self._render_scan_evolution_window()
            return

        self.scan_evolution_window = tk.Toplevel(self.root)
        self.scan_evolution_window.title("Scan Evolution")
        self.scan_evolution_window.geometry("1500x980")
        self.scan_evolution_window.protocol("WM_DELETE_WINDOW", self._close_scan_evolution_window)
        self._scan_evolution_scan_key = self._scan_key(scan)

        controls = tk.Frame(self.scan_evolution_window, padx=8, pady=8)
        controls.pack(side="top", fill="x")
        self.scan_evolution_status_var = tk.StringVar(value="Showing selected scan evolution.")
        self.scan_evolution_mod360_var = tk.BooleanVar(value=True)
        self.scan_evolution_show_gaussian_var = tk.BooleanVar(value=False)
        self.scan_evolution_show_dsdf_var = tk.BooleanVar(value=False)
        self.scan_evolution_show_phase_2pi_var = tk.BooleanVar(value=False)
        self.scan_evolution_show_phase_vna_var = tk.BooleanVar(value=False)
        self.scan_evolution_show_phase_other_var = tk.BooleanVar(value=False)
        self.scan_evolution_show_attached_res_var = tk.BooleanVar(value=False)
        tk.Label(controls, textvariable=self.scan_evolution_status_var, anchor="w").pack(
            side="left", fill="x", expand=True
        )
        marker_controls = tk.Frame(self.scan_evolution_window, padx=8, pady=0)
        marker_controls.pack(side="top", fill="x")
        tk.Checkbutton(
            controls,
            text="Mod 360 phase",
            variable=self.scan_evolution_mod360_var,
            command=self._scan_evolution_toggle_phase_wrap,
        ).pack(side="right", padx=(0, 8))
        tk.Button(controls, text="Choose Scan", width=12, command=self._scan_evolution_choose_scan).pack(
            side="right"
        )
        tk.Button(controls, text="Reset View", width=12, command=self._scan_evolution_reset_view).pack(
            side="right", padx=(0, 8)
        )
        for text, var in (
            ("Gaussian", self.scan_evolution_show_gaussian_var),
            ("dS21/df", self.scan_evolution_show_dsdf_var),
            ("2pi", self.scan_evolution_show_phase_2pi_var),
            ("VNA phase", self.scan_evolution_show_phase_vna_var),
            ("Other phase", self.scan_evolution_show_phase_other_var),
            ("Resonators", self.scan_evolution_show_attached_res_var),
        ):
            tk.Checkbutton(
                marker_controls,
                text=text,
                variable=var,
                command=self._render_scan_evolution_window,
            ).pack(side="left", padx=(0, 8))

        self.scan_evolution_figure = Figure(figsize=(14, 9))
        self.scan_evolution_canvas = FigureCanvasTkAgg(self.scan_evolution_figure, master=self.scan_evolution_window)
        self.scan_evolution_toolbar = NavigationToolbar2Tk(self.scan_evolution_canvas, self.scan_evolution_window)
        self.scan_evolution_toolbar.update()
        self.scan_evolution_toolbar.pack(side="top", fill="x")
        self.scan_evolution_canvas.get_tk_widget().pack(fill="both", expand=True)

        self._render_scan_evolution_window()

    def _scan_evolution_choose_scan(self) -> None:
        scan = self._choose_one_selected_scan()
        if scan is None:
            return
        self._scan_evolution_scan_key = self._scan_key(scan)
        self._render_scan_evolution_window()

    def _close_scan_evolution_window(self) -> None:
        if self.scan_evolution_window is not None and self.scan_evolution_window.winfo_exists():
            self.scan_evolution_window.destroy()
        self.scan_evolution_window = None
        self.scan_evolution_canvas = None
        self.scan_evolution_toolbar = None
        self.scan_evolution_figure = None
        self.scan_evolution_status_var = None
        self.scan_evolution_mod360_var = None
        self.scan_evolution_show_gaussian_var = None
        self.scan_evolution_show_dsdf_var = None
        self.scan_evolution_show_phase_2pi_var = None
        self.scan_evolution_show_phase_vna_var = None
        self.scan_evolution_show_phase_other_var = None
        self.scan_evolution_show_attached_res_var = None
        self._scan_evolution_scan_key = None
        self._scan_evolution_stage_rows = []
        self._scan_evolution_axes_rows = []
        self._scan_evolution_syncing_xlim = False

    def _scan_evolution_current_scan(self) -> Optional[VNAScan]:
        if self._scan_evolution_scan_key is None:
            return None
        for scan in self.dataset.vna_scans:
            if self._scan_key(scan) == self._scan_evolution_scan_key:
                return scan
        return None

    def _scan_evolution_visible_xlim(self) -> Optional[tuple[float, float]]:
        if not self._scan_evolution_axes_rows:
            return None
        ax_amp = self._scan_evolution_axes_rows[0][0]
        try:
            x0, x1 = ax_amp.get_xlim()
        except Exception:
            return None
        return (float(x0), float(x1)) if float(x0) <= float(x1) else (float(x1), float(x0))

    def _scan_evolution_autoscale_amp_phase(self) -> None:
        xlim = self._scan_evolution_visible_xlim()
        if xlim is None:
            return
        lo, hi = xlim
        for stage, (ax_amp, ax_phase, _ax_complex) in zip(self._scan_evolution_stage_rows, self._scan_evolution_axes_rows):
            mask = np.isfinite(stage["freq_ghz"]) & (stage["freq_ghz"] >= lo) & (stage["freq_ghz"] <= hi)
            if np.any(mask):
                amp = np.asarray(stage["amp"], dtype=float)[mask]
                phase = self._scan_evolution_phase_display(stage)[mask]
                if amp.size:
                    amp_min = float(np.min(amp))
                    amp_max = float(np.max(amp))
                    amp_pad = 1.0 if amp_max <= amp_min else 0.05 * (amp_max - amp_min)
                    ax_amp.set_ylim(amp_min - amp_pad, amp_max + amp_pad)
                if phase.size:
                    if self.scan_evolution_mod360_var is not None and self.scan_evolution_mod360_var.get():
                        ax_phase.set_ylim(-180.0, 180.0)
                    else:
                        ph_min = float(np.min(phase))
                        ph_max = float(np.max(phase))
                        ph_pad = 1.0 if ph_max <= ph_min else 0.05 * (ph_max - ph_min)
                        ax_phase.set_ylim(ph_min - ph_pad, ph_max + ph_pad)

    def _scan_evolution_update_complex_axes(self) -> None:
        xlim = self._scan_evolution_visible_xlim()
        if xlim is None:
            return
        lo, hi = xlim
        scan = self._scan_evolution_current_scan()
        for stage, (_ax_amp, _ax_phase, ax_complex) in zip(self._scan_evolution_stage_rows, self._scan_evolution_axes_rows):
            ax_complex.clear()
            mask = np.isfinite(stage["freq_ghz"]) & (stage["freq_ghz"] >= lo) & (stage["freq_ghz"] <= hi)
            if np.any(mask):
                real = np.asarray(stage["real"], dtype=float)[mask]
                imag = np.asarray(stage["imag"], dtype=float)[mask]
                ax_complex.plot(real, imag, color="tab:green", linewidth=1.0)
                if scan is not None:
                    for real_pts, imag_pts, style, _label in self._scan_evolution_complex_overlay_points(
                        scan,
                        real_values=np.asarray(stage["real"], dtype=float),
                        imag_values=np.asarray(stage["imag"], dtype=float),
                        freq_lo_ghz=lo,
                        freq_hi_ghz=hi,
                    ):
                        ax_complex.plot(real_pts, imag_pts, **style)
                if self.scan_evolution_show_attached_res_var is not None and bool(self.scan_evolution_show_attached_res_var.get()):
                    if scan is not None:
                        _amp_points, _phase_points, complex_points = self._scan_evolution_attached_resonator_points(
                            scan,
                            phase_values=self._scan_evolution_phase_display(stage),
                            amp_values=np.asarray(stage["amp"], dtype=float),
                            real_values=np.asarray(stage["real"], dtype=float),
                            imag_values=np.asarray(stage["imag"], dtype=float),
                        )
                        complex_points = [
                            pt
                            for pt in complex_points
                            if lo <= float(pt["x_hz"]) / 1.0e9 <= hi
                        ]
                        if complex_points:
                            ax_complex.plot(
                                [float(pt["real"]) for pt in complex_points],
                                [float(pt["imag"]) for pt in complex_points],
                                linestyle="none",
                                marker="s",
                                markersize=5,
                                color="black",
                            )
                            for pt in complex_points:
                                ax_complex.annotate(
                                    str(pt["label"]),
                                    (float(pt["real"]), float(pt["imag"])),
                                    xytext=(4, 3),
                                    textcoords="offset points",
                                    fontsize=8,
                                    color="black",
                                )
                ax_complex.set_aspect("equal", adjustable="box")
                re_min = float(np.min(real))
                re_max = float(np.max(real))
                im_min = float(np.min(imag))
                im_max = float(np.max(imag))
                re_pad = 1.0 if re_max <= re_min else 0.05 * (re_max - re_min)
                im_pad = 1.0 if im_max <= im_min else 0.05 * (im_max - im_min)
                ax_complex.set_xlim(re_min - re_pad, re_max + re_pad)
                ax_complex.set_ylim(im_min - im_pad, im_max + im_pad)
            else:
                ax_complex.text(0.5, 0.5, "No data in range", ha="center", va="center", transform=ax_complex.transAxes)
            ax_complex.grid(True, alpha=0.3)
            ax_complex.set_xlabel("Real(S21)")
            ax_complex.set_ylabel("Imag(S21)")

    def _scan_evolution_on_xlim_changed(self, changed_ax) -> None:
        if self._scan_evolution_syncing_xlim:
            return
        try:
            xlim = changed_ax.get_xlim()
        except Exception:
            return
        self._scan_evolution_syncing_xlim = True
        try:
            for ax_amp, ax_phase, _ax_complex in self._scan_evolution_axes_rows:
                if ax_amp is not changed_ax:
                    ax_amp.set_xlim(xlim)
                if ax_phase is not changed_ax:
                    ax_phase.set_xlim(xlim)
            self._scan_evolution_autoscale_amp_phase()
            self._scan_evolution_update_complex_axes()
        finally:
            self._scan_evolution_syncing_xlim = False
        if self.scan_evolution_canvas is not None:
            self.scan_evolution_canvas.draw_idle()

    def _scan_evolution_reset_view(self) -> None:
        if not self._scan_evolution_axes_rows or not self._scan_evolution_stage_rows:
            return
        freq_min = min(float(stage["freq_ghz"][0]) for stage in self._scan_evolution_stage_rows)
        freq_max = max(float(stage["freq_ghz"][-1]) for stage in self._scan_evolution_stage_rows)
        x_pad = max((freq_max - freq_min) * 0.01, 1e-6)
        xlim = (freq_min - x_pad, freq_max + x_pad)
        self._scan_evolution_syncing_xlim = True
        try:
            for ax_amp, ax_phase, _ax_complex in self._scan_evolution_axes_rows:
                ax_amp.set_xlim(xlim)
                ax_phase.set_xlim(xlim)
        finally:
            self._scan_evolution_syncing_xlim = False
        self._scan_evolution_autoscale_amp_phase()
        self._scan_evolution_update_complex_axes()
        if self.scan_evolution_canvas is not None:
            self.scan_evolution_canvas.draw_idle()

    def _render_scan_evolution_window(self) -> None:
        if self.scan_evolution_figure is None or self.scan_evolution_canvas is None:
            return
        scan = self._scan_evolution_current_scan()
        if scan is None:
            return
        prior_limits: list[tuple[tuple[float, float], tuple[float, float], tuple[float, float]]] = []
        for ax_amp, ax_phase, ax_complex in self._scan_evolution_axes_rows:
            try:
                prior_limits.append(
                    (
                        tuple(ax_amp.get_xlim()),
                        tuple(ax_amp.get_ylim()),
                        tuple(ax_phase.get_ylim()),
                    )
                )
            except Exception:
                prior_limits = []
                break
        stages = self._scan_evolution_stage_rows_for_scan(scan)
        self._scan_evolution_stage_rows = stages
        self._scan_evolution_axes_rows = []
        self.scan_evolution_figure.clear()
        if not stages:
            ax = self.scan_evolution_figure.add_subplot(111)
            ax.text(0.5, 0.5, "No attached processing stages available for this scan.", ha="center", va="center")
            ax.axis("off")
            self.scan_evolution_canvas.draw_idle()
            return

        nrows = len(stages)
        axes = np.atleast_2d(self.scan_evolution_figure.subplots(nrows, 3, sharex=False, squeeze=False))
        used_overlay_labels: set[str] = set()
        used_res_labels: set[str] = set()
        for row_idx, stage in enumerate(stages):
            ax_amp = axes[row_idx, 0]
            ax_phase = axes[row_idx, 1]
            ax_complex = axes[row_idx, 2]
            phase_display = self._scan_evolution_phase_display(stage)
            ax_amp.plot(stage["freq_ghz"], stage["amp"], color="tab:blue", linewidth=1.0)
            ax_phase.plot(stage["freq_ghz"], phase_display, color="tab:orange", linewidth=1.0)
            self._scan_evolution_add_overlays(
                ax_amp,
                scan,
                values=np.asarray(stage["amp"], dtype=float),
                use_phase=False,
                used_labels=used_overlay_labels,
            )
            self._scan_evolution_add_overlays(
                ax_phase,
                scan,
                values=phase_display,
                use_phase=True,
                used_labels=used_overlay_labels,
            )
            if self.scan_evolution_show_attached_res_var is not None and bool(self.scan_evolution_show_attached_res_var.get()):
                amp_points, phase_points, _complex_points = self._scan_evolution_attached_resonator_points(
                    scan,
                    phase_values=phase_display,
                    amp_values=np.asarray(stage["amp"], dtype=float),
                    real_values=np.asarray(stage["real"], dtype=float),
                    imag_values=np.asarray(stage["imag"], dtype=float),
                )
                if amp_points:
                    ax_amp.plot(
                        [pt["x_hz"] / 1.0e9 for pt in amp_points],
                        [pt["y"] for pt in amp_points],
                        linestyle="none",
                        marker="s",
                        markersize=5,
                        color="black",
                        label=("Attached resonators" if "Attached resonators" not in used_res_labels else None),
                    )
                    used_res_labels.add("Attached resonators")
                if phase_points:
                    ax_phase.plot(
                        [pt["x_hz"] / 1.0e9 for pt in phase_points],
                        [pt["y"] for pt in phase_points],
                        linestyle="none",
                        marker="s",
                        markersize=5,
                        color="black",
                        label=("Attached resonators" if "Attached resonators" not in used_res_labels else None),
                    )
                    used_res_labels.add("Attached resonators")
            ax_amp.grid(True, alpha=0.3)
            ax_phase.grid(True, alpha=0.3)
            ax_amp.set_ylabel(f"{stage['name']}\nAmplitude")
            ax_phase.set_ylabel("Phase (deg)")
            if row_idx == 0 and (used_overlay_labels or used_res_labels):
                ax_phase.legend(loc="best", fontsize=8)
            if row_idx == 0:
                ax_amp.set_title("Amplitude", fontsize=11)
                ax_phase.set_title("Phase", fontsize=11)
                ax_complex.set_title("Complex Plane", fontsize=11)
            if row_idx == nrows - 1:
                ax_amp.set_xlabel("Frequency (GHz)")
                ax_phase.set_xlabel("Frequency (GHz)")
            self._scan_evolution_axes_rows.append((ax_amp, ax_phase, ax_complex))

        self.scan_evolution_figure.suptitle(
            f"Scan Evolution | {Path(scan.filename).name}",
            fontsize=12,
        )
        self.scan_evolution_figure.tight_layout()

        for ax_amp, ax_phase, _ax_complex in self._scan_evolution_axes_rows:
            ax_amp.callbacks.connect("xlim_changed", self._scan_evolution_on_xlim_changed)
            ax_phase.callbacks.connect("xlim_changed", self._scan_evolution_on_xlim_changed)

        if len(prior_limits) == len(self._scan_evolution_axes_rows):
            self._scan_evolution_syncing_xlim = True
            try:
                for (ax_amp, ax_phase, _ax_complex), (xlim, amp_ylim, phase_ylim) in zip(
                    self._scan_evolution_axes_rows,
                    prior_limits,
                ):
                    ax_amp.set_xlim(xlim)
                    ax_phase.set_xlim(xlim)
                    ax_amp.set_ylim(amp_ylim)
                    ax_phase.set_ylim(phase_ylim)
            finally:
                self._scan_evolution_syncing_xlim = False
            self._scan_evolution_update_complex_axes()
            self.scan_evolution_canvas.draw_idle()
        else:
            self._scan_evolution_reset_view()
        if self.scan_evolution_status_var is not None:
            self.scan_evolution_status_var.set(
                f"Showing {len(stages)} stage(s) for {Path(scan.filename).name}. Zoom amplitude or phase to update all panels."
            )

    def _reload_transcript_ui(self) -> None:
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        for entry in self.dataset.transcript:
            self.log_text.insert("end", f"[{entry['timestamp']}] {entry['message']}\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    def _log(self, message: str) -> None:
        timestamp = datetime.now().isoformat(timespec="seconds")
        self.dataset.transcript.append({"timestamp": timestamp, "message": message})
        self._append_transcript_line(timestamp, message)

    def _select_setting_option(
        self, title: str, prompt: str, options: List[str], default_index: int = 0
    ) -> Optional[int]:
        if not options:
            return None
        if len(options) == 1:
            return 0

        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        dialog.geometry("760x360")
        dialog.transient(self.root)
        dialog.grab_set()

        tk.Label(dialog, text=prompt, anchor="w", justify="left", wraplength=720).pack(
            fill="x", padx=10, pady=(10, 4)
        )
        listbox = tk.Listbox(dialog, width=110, height=12)
        listbox.pack(fill="both", expand=True, padx=10, pady=4)
        for item in options:
            listbox.insert(tk.END, item)
        if 0 <= default_index < len(options):
            listbox.selection_set(default_index)
            listbox.see(default_index)
        else:
            listbox.selection_set(0)

        selected_index: Dict[str, Optional[int]] = {"value": None}

        def choose() -> None:
            sel = listbox.curselection()
            if not sel:
                return
            selected_index["value"] = int(sel[0])
            dialog.destroy()

        def cancel() -> None:
            dialog.destroy()

        btns = tk.Frame(dialog)
        btns.pack(fill="x", padx=10, pady=(4, 10))
        tk.Button(btns, text="Use Selected", command=choose).pack(side="right")
        tk.Button(btns, text="Cancel", command=cancel).pack(side="right", padx=(0, 8))

        dialog.wait_window()
        return selected_index["value"]

    def _select_multiple_setting_options(
        self,
        title: str,
        prompt: str,
        options: List[str],
        default_indices: Optional[List[int]] = None,
    ) -> List[int]:
        if not options:
            return []

        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        dialog.geometry("760x420")
        dialog.transient(self.root)
        dialog.grab_set()

        tk.Label(dialog, text=prompt, anchor="w", justify="left", wraplength=720).pack(
            fill="x", padx=10, pady=(10, 4)
        )
        listbox = tk.Listbox(dialog, width=110, height=16, selectmode=tk.MULTIPLE)
        listbox.pack(fill="both", expand=True, padx=10, pady=4)
        for item in options:
            listbox.insert(tk.END, item)

        defaults = sorted({int(idx) for idx in (default_indices or []) if 0 <= int(idx) < len(options)})
        if defaults:
            for idx in defaults:
                listbox.selection_set(idx)
            listbox.see(defaults[0])

        selected_indices: Dict[str, List[int]] = {"value": []}

        def choose() -> None:
            selected_indices["value"] = [int(idx) for idx in listbox.curselection()]
            dialog.destroy()

        def cancel() -> None:
            dialog.destroy()

        btns = tk.Frame(dialog)
        btns.pack(fill="x", padx=10, pady=(4, 10))
        tk.Button(btns, text="Use Selected", command=choose).pack(side="right")
        tk.Button(btns, text="Cancel", command=cancel).pack(side="right", padx=(0, 8))

        dialog.wait_window()
        return selected_indices["value"]

    def _confirm_bulk_text_changes(self, title: str, prompt: str, lines: List[str]) -> bool:
        if not lines:
            return False
        dialog = tk.Toplevel(self.root)
        dialog.title(title)
        dialog.geometry("860x460")
        dialog.transient(self.root)
        dialog.grab_set()

        tk.Label(dialog, text=prompt, anchor="w", justify="left", wraplength=820).pack(
            fill="x", padx=10, pady=(10, 4)
        )
        listbox = tk.Listbox(dialog, width=125, height=18)
        listbox.pack(fill="both", expand=True, padx=10, pady=4)
        for line in lines:
            listbox.insert(tk.END, line)

        decision: Dict[str, bool] = {"approve": False}

        def approve() -> None:
            decision["approve"] = True
            dialog.destroy()

        def cancel() -> None:
            dialog.destroy()

        btns = tk.Frame(dialog)
        btns.pack(fill="x", padx=10, pady=(4, 10))
        tk.Button(btns, text="Apply All", command=approve).pack(side="right")
        tk.Button(btns, text="Cancel", command=cancel).pack(side="right", padx=(0, 8))

        dialog.wait_window()
        return bool(decision["approve"])

    @staticmethod
    def _scan_dialog_path_text(scan: VNAScan) -> str:
        source_dir = str(getattr(scan, "source_dir", "") or "").strip()
        try:
            path = Path(source_dir) if source_dir else Path(str(scan.filename)).resolve().parent
        except Exception:
            path = Path(source_dir) if source_dir else Path(str(scan.filename)).parent
        parts = [part for part in path.parts if part not in (path.anchor, "")]
        if not parts:
            return str(path)
        return str(Path(*parts[-2:])) if len(parts) >= 2 else parts[-1]

    def _scan_dialog_label(
        self,
        scan: VNAScan,
        *,
        index: Optional[int] = None,
        include_file_timestamp: bool = False,
        include_loaded_at: bool = False,
        include_group: bool = False,
    ) -> str:
        prefix = f"{int(index):03d} | " if index is not None else ""
        parts = [prefix + Path(str(scan.filename)).name, f"folder {self._scan_dialog_path_text(scan)}"]
        if include_file_timestamp:
            file_timestamp = str(getattr(scan, "file_timestamp", "")).strip() or "unknown"
            parts.append(f"file {file_timestamp}")
        if include_loaded_at:
            parts.append(f"loaded {scan.loaded_at}")
        if include_group:
            parts.append(f"group {int(scan.plot_group)}" if scan.plot_group is not None else "no group")
        return " | ".join(parts)

    @staticmethod
    def _scan_file_two_level_context(scan: VNAScan) -> str:
        try:
            path = Path(str(scan.filename)).resolve()
        except Exception:
            path = Path(str(scan.filename))
        parts = [part for part in path.parts if part not in (path.anchor, "")]
        if not parts:
            return str(path)
        if len(parts) >= 3:
            return str(Path(*parts[-3:]))
        return str(Path(*parts))

    @staticmethod
    def _scan_sort_stamp(scan: VNAScan) -> str:
        file_stamp = str(getattr(scan, "file_timestamp", "") or "").strip()
        if file_stamp:
            return file_stamp
        return str(getattr(scan, "loaded_at", "") or "").strip()

    @staticmethod
    def _scan_sort_date_label(scan: VNAScan) -> str:
        stamp = DataAnalysisGUI._scan_sort_stamp(scan)
        if not stamp:
            return "unknown date"
        return stamp.split("T", 1)[0]

    @staticmethod
    def _scan_sort_key(scan: VNAScan) -> tuple[int, object, str]:
        stamp = DataAnalysisGUI._scan_sort_stamp(scan)
        if scan.plot_group is None:
            group_key = (1, 0)
        else:
            group_key = (0, int(scan.plot_group))
        name_key = Path(str(scan.filename)).name.lower()
        if stamp:
            try:
                return (0, datetime.fromisoformat(stamp), group_key, name_key)
            except Exception:
                date_part = stamp.split("T", 1)[0]
                try:
                    return (0, datetime.fromisoformat(date_part), group_key, name_key)
                except Exception:
                    return (1, stamp.lower(), group_key, name_key)
        return (2, "", group_key, name_key)

    @staticmethod
    def _date_from_source_dir_name(source_dir: str) -> Optional[str]:
        source_text = str(source_dir).strip()
        if not source_text:
            return None
        candidates = [Path(source_text).name, source_text]
        for text in candidates:
            for match in re.finditer(r"(?<!\d)(\d{8})(?!\d)", text):
                token = match.group(1)
                try:
                    dt = datetime.strptime(token, "%Y%m%d")
                except Exception:
                    continue
                return dt.date().isoformat()
            for match in re.finditer(r"(?<!\d)(\d{6})(?=_|\s|$)", text):
                token = match.group(1)
                try:
                    dt = datetime.strptime(token, "%y%m%d")
                except Exception:
                    continue
                return dt.date().isoformat()
        return None

    @staticmethod
    def _date_from_filename_8digit(filename: str) -> Optional[str]:
        name_text = Path(str(filename)).name.strip()
        if not name_text:
            return None
        for match in re.finditer(r"(?<!\d)(\d{8})(?!\d)", name_text):
            token = match.group(1)
            try:
                dt = datetime.strptime(token, "%Y%m%d")
            except Exception:
                continue
            return dt.date().isoformat()
        return None

    @staticmethod
    def _replace_iso_date_fixed_1pm(new_date_iso: str) -> str:
        return f"{new_date_iso}T13:00:00"

    def _update_selected_vna_dates(self, mode: str) -> None:
        scans = self._selected_scans()
        if not scans:
            messagebox.showwarning(
                "No selection",
                "No scans selected for analysis.\nUse 'Select Scans for Analysis' first.",
            )
            return

        if mode == "filename":
            parse_date = lambda scan: self._date_from_filename_8digit(getattr(scan, "filename", ""))
            skip_text = (
                lambda scan: f"{self._scan_file_two_level_context(scan)}: "
                f"no YYYYMMDD date token found in filename {Path(str(getattr(scan, 'filename', ''))).name}"
            )
            confirm_title = "Update VNA Dates From Filename"
            confirm_message = (
                "The following selected VNA scans will have their file date updated to a date parsed from the filename "
                "(first YYYYMMDD token). This updated date will be used in plots."
            )
            no_updates_message = "No selected scans need a date update from filename."
            per_scan_event = "update_vna_file_timestamp_from_filename"
            dataset_event = "update_selected_vna_dates_from_filename"
            log_mode_text = "filename date"
            info_mode_text = "filename"
        else:
            parse_date = lambda scan: self._date_from_source_dir_name(getattr(scan, "source_dir", ""))
            skip_text = (
                lambda scan: f"{self._scan_file_two_level_context(scan)}: "
                f"no YYYYMMDD or YYMMDD_ / YYMMDD<space> date token found in {scan.source_dir}"
            )
            confirm_title = "Update VNA Dates From Source Directory"
            confirm_message = (
                "The following selected VNA scans will have their file date updated from the filesystem date to a date "
                "parsed from source_dir. This updated date will be used in plots."
            )
            no_updates_message = "No selected scans need a date update from source_dir."
            per_scan_event = "update_vna_file_timestamp_from_source_dir"
            dataset_event = "update_selected_vna_dates_from_source_dir"
            log_mode_text = "source_dir date"
            info_mode_text = "source_dir"

        proposed_updates: list[tuple[VNAScan, str, str, str]] = []
        skipped: list[str] = []
        for scan in scans:
            new_date_iso = parse_date(scan)
            if new_date_iso is None:
                skipped.append(skip_text(scan))
                continue
            old_timestamp = str(getattr(scan, "file_timestamp", "") or "").strip()
            new_timestamp = self._replace_iso_date_fixed_1pm(new_date_iso)
            if new_timestamp == old_timestamp:
                continue
            proposed_updates.append((scan, old_timestamp, new_timestamp, new_date_iso))

        if not proposed_updates:
            message = no_updates_message
            if skipped:
                message += "\n\nSkipped:\n" + "\n".join(skipped[:10])
            messagebox.showwarning("No updates", message)
            return

        preview_lines = [
            f"{self._scan_file_two_level_context(scan)}: {old if old else 'unknown'} -> {new}"
            for scan, old, new, _date_iso in proposed_updates
        ]
        if skipped:
            preview_lines.append("")
            preview_lines.append("Skipped:")
            preview_lines.extend(skipped[:10])
            if len(skipped) > 10:
                preview_lines.append(f"... and {len(skipped) - 10} more")

        approved = self._confirm_bulk_text_changes(
            confirm_title,
            confirm_message,
            preview_lines,
        )
        if not approved:
            return

        changed_count = 0
        for scan, old_timestamp, new_timestamp, new_date_iso in proposed_updates:
            scan.file_timestamp = new_timestamp
            scan.processing_history.append(
                _make_event(
                    per_scan_event,
                    {
                        "filename": scan.filename,
                        "source_dir": scan.source_dir,
                        "old_file_timestamp": old_timestamp,
                        "new_file_timestamp": new_timestamp,
                        "parsed_date": new_date_iso,
                    },
                )
            )
            changed_count += 1

        self.dataset.processing_history.append(
            _make_event(
                dataset_event,
                {
                    "selected_count": len(scans),
                    "updated_count": changed_count,
                    "skipped_count": len(skipped),
                },
            )
        )
        self._mark_dirty()
        self._refresh_status()
        self._autosave_dataset()
        self._log(
            f"Updated {changed_count} selected VNA file timestamp(s) from {log_mode_text}; skipped {len(skipped)}."
        )
        messagebox.showinfo(
            "Dates updated",
            f"Updated {changed_count} selected scan date(s) from {info_mode_text}."
            + (f"\nSkipped {len(skipped)} scan(s)." if skipped else ""),
        )

    def update_selected_vna_dates_from_source_dir(self) -> None:
        self._update_selected_vna_dates(mode="dir")

    def update_selected_vna_dates_from_path(self) -> None:
        mode = str(self.update_dates_mode_var.get() or "dir").strip().lower()
        if mode not in {"dir", "filename"}:
            mode = "dir"
        self._update_selected_vna_dates(mode=mode)

    def open_update_dates_dialog(self) -> None:
        dialog = tk.Toplevel(self.root)
        dialog.title("Update Dates From Path")
        dialog.geometry("420x190")
        dialog.transient(self.root)
        dialog.grab_set()

        tk.Label(
            dialog,
            text="Choose where to parse scan dates from:",
            anchor="w",
            justify="left",
        ).pack(anchor="w", padx=12, pady=(12, 8))

        mode_var = tk.StringVar(value=str(self.update_dates_mode_var.get() or "dir"))
        tk.Radiobutton(
            dialog,
            text="Directory (default behavior)",
            value="dir",
            variable=mode_var,
            anchor="w",
            justify="left",
        ).pack(anchor="w", padx=18, pady=(0, 4))
        tk.Radiobutton(
            dialog,
            text="Filename (first YYYYMMDD token)",
            value="filename",
            variable=mode_var,
            anchor="w",
            justify="left",
        ).pack(anchor="w", padx=18, pady=(0, 8))

        button_row = tk.Frame(dialog)
        button_row.pack(fill="x", padx=12, pady=(4, 12))

        def _run() -> None:
            chosen = str(mode_var.get() or "dir").strip().lower()
            if chosen not in {"dir", "filename"}:
                chosen = "dir"
            self.update_dates_mode_var.set(chosen)
            dialog.destroy()
            self._update_selected_vna_dates(chosen)

        tk.Button(button_row, text="Cancel", width=10, command=dialog.destroy).pack(side="right")
        tk.Button(button_row, text="Update", width=10, command=_run).pack(side="right", padx=(0, 8))

    def reorder_vna_scans_by_date(self) -> None:
        scans = list(self.dataset.vna_scans)
        if not scans:
            messagebox.showwarning("No data", "No VNA scans are loaded in this dataset.")
            return

        reordered = sorted(scans, key=self._scan_sort_key)
        if [id(scan) for scan in reordered] == [id(scan) for scan in scans]:
            messagebox.showinfo("Already ordered", "Scans are already ordered by date.")
            return

        preview_lines = []
        for idx, scan in enumerate(reordered):
            group_text = f"group {int(scan.plot_group)}" if scan.plot_group is not None else "no group"
            preview_lines.append(
                f"{idx:03d} | {self._scan_file_two_level_context(scan)} | {group_text} | {self._scan_sort_date_label(scan)}"
            )

        approved = self._confirm_bulk_text_changes(
            "Reorder VNA Scans By Date",
            "The scans below are in the proposed date-sorted order. Approve to reorder dataset scan order.",
            preview_lines,
        )
        if not approved:
            return

        selected_set = set(self.dataset.selected_scan_keys)
        self.dataset.vna_scans = reordered
        self.dataset.selected_scan_keys = [
            self._scan_key(scan) for scan in self.dataset.vna_scans if self._scan_key(scan) in selected_set
        ]
        self.dataset.processing_history.append(
            _make_event(
                "reorder_vna_scans_by_date",
                {
                    "scan_count": len(self.dataset.vna_scans),
                },
            )
        )
        self._mark_dirty()
        self._refresh_status()
        self._autosave_dataset()
        self._log(f"Reordered {len(self.dataset.vna_scans)} VNA scan(s) by date.")
        messagebox.showinfo(
            "Scans reordered",
            f"Reordered {len(self.dataset.vna_scans)} scan(s) by date. Dataset saved automatically.",
        )

    def _refresh_status(self) -> None:
        created = self.dataset.created_at if self.dataset.created_at else "Unassigned"
        name = self.dataset.dataset_name if self.dataset.dataset_name else "Unassigned"
        self.dataset_meta_var.set(f"Dataset: {name} | Created: {created}")
        self._prune_selected_scan_keys()
        self.dataset_label_var.set(str(self.dataset_path))
        self.scan_count_var.set(f"Loaded VNA scans in dataset: {len(self.dataset.vna_scans)}")
        selected_names = [Path(scan.filename).name for scan in self._selected_scans()]
        if len(selected_names) > 3:
            selected_text = ", ".join(selected_names[:3]) + f", ... (+{len(selected_names) - 3} more)"
        elif selected_names:
            selected_text = ", ".join(selected_names)
        else:
            selected_text = "None"
        self.selection_var.set(
            f"Selected scans for analysis ({len(selected_names)}): {selected_text}"
        )
        last_saved = self.dataset.last_saved_at if self.dataset.last_saved_at else "Never"
        size_text = "0.00 MB"
        try:
            if self.dataset_path.exists():
                size_mb = self.dataset_path.stat().st_size / (1024.0 * 1024.0)
                size_text = f"{size_mb:.2f} MB"
        except Exception:
            size_text = "Unknown"
        self.saved_var.set(f"Last saved: {last_saved} | Dataset file size: {size_text}")
        self._update_save_button_state()
        self._update_interp_button_state()
        self._update_norm_button_state()
        self._update_gauss_button_state()
        self._update_dsdf_button_state()
        self._update_unwrap_button_state()
        self._update_phase2_button_state()
        self._update_phase3_button_state()
        self._update_baseline_button_state()
        self._update_select_scans_button_state()
        self._update_res_button_state()

    def _has_data_to_save(self) -> bool:
        return len(self.dataset.vna_scans) > 0

    def _mark_dirty(self) -> None:
        self._dirty = True
        self._update_save_button_state()

    def _mark_clean(self) -> None:
        self._dirty = False
        self._update_save_button_state()

    def _update_save_button_state(self) -> None:
        return

    def _attach_save_and_close_baseline(self) -> None:
        self.attach_baseline_filter()
        self._close_baseline_window()

    def _attach_save_and_close_interp(self) -> None:
        self._interp_attach()
        self._interp_close()

    def _attach_save_and_close_norm(self) -> None:
        self._norm_attach()
        self._norm_close()

    def _attach_save_and_close_gauss(self) -> None:
        self._gauss_attach()
        self._gauss_close()

    def _attach_save_and_close_dsdf(self) -> None:
        self._dsdf_attach()
        self._dsdf_close()

    def _attach_save_and_close_unwrap(self) -> None:
        self._unwrap_attach()
        self._unwrap_close()

    def _attach_save_and_close_phase2(self) -> None:
        self._phase2_attach()
        self._phase2_close()

    def _attach_save_and_close_phase3(self) -> None:
        self._phase3_attach()
        self._phase3_close()

    def _path_identity_from_stem(self, path: Path) -> tuple[str, str]:
        match = re.match(r"^(?P<prefix>\d{8}_\d{6})_(?P<name>.+)$", path.stem)
        if not match:
            return "", ""
        return match.group("prefix"), match.group("name")

    def _backfill_dataset_identity_from_path(self) -> None:
        prefix, name = self._path_identity_from_stem(self.dataset_path.resolve())
        if not prefix or not name:
            return
        if not self.dataset.created_at:
            try:
                dt = datetime.strptime(prefix, "%Y%m%d_%H%M%S")
                self.dataset.created_at = dt.isoformat(timespec="seconds")
            except Exception:
                pass
        if not self.dataset.dataset_name:
            self.dataset.dataset_name = name

    def _reconcile_dataset_path_for_save(self) -> None:
        current_path = self.dataset_path.resolve()
        self._backfill_dataset_identity_from_path()
        if self.dataset.dataset_name and self.dataset.created_at:
            target_path = _dataset_pickle_path(self.dataset).resolve()
        else:
            target_path = current_path

        if target_path == current_path:
            self.dataset_path = target_path
            return

        if target_path.exists():
            raise FileExistsError(f"Target dataset file already exists:\n{target_path}")

        current_exists = current_path.exists()
        current_dir = current_path.parent.resolve()
        target_dir = target_path.parent.resolve()

        if current_exists and current_dir == target_dir:
            current_path.rename(target_path)
        elif current_exists and current_dir != target_dir:
            if current_dir == DATASETS_DIR.resolve():
                target_dir.mkdir(parents=True, exist_ok=True)
                shutil.move(str(current_path), str(target_path))
            else:
                if target_dir.exists():
                    raise FileExistsError(f"Target dataset folder already exists:\n{target_dir}")
                shutil.move(str(current_dir), str(target_dir))
                moved_path = target_dir / current_path.name
                if moved_path.exists() and moved_path != target_path:
                    moved_path.rename(target_path)
        else:
            target_dir.mkdir(parents=True, exist_ok=True)

        self.dataset_path = target_path

    def _persist_dataset(self) -> bool:
        try:
            if not self.dataset.created_at:
                self.dataset.created_at = datetime.now().isoformat(timespec="seconds")
            self._reconcile_dataset_path_for_save()

            self.dataset.processing_history.append(
                _make_event("save_dataset", {"dataset_path": str(self.dataset_path)})
            )
            _save_dataset(self.dataset, self.dataset_path)
            _write_app_state(self.dataset_path)
            self._mark_clean()
            self._refresh_status()
            return True
        except Exception as exc:
            self._mark_dirty()
            self._refresh_status()
            self._log(f"Save failed: {exc}")
            messagebox.showerror("Save failed", str(exc))
            return False

    def _autosave_dataset(self) -> bool:
        return self._persist_dataset()

    def start_new_dataset(self) -> None:
        if self.dataset.vna_scans or self._dirty:
            ok = messagebox.askyesno(
                "Start New Dataset",
                "Start a new empty dataset?\nUnsaved changes in the current dataset will be lost.",
            )
            if not ok:
                return

        proposed_name = simpledialog.askstring(
            "New Dataset Prefix",
            "Enter the prefix to use for the new dataset folder and pickle filename:",
            parent=self.root,
        )
        if proposed_name is None:
            return
        cleaned_name = _safe_name(proposed_name)
        if not cleaned_name:
            messagebox.showwarning("Invalid prefix", "Please enter a non-empty dataset prefix.")
            return

        for closer in (
            getattr(self, "_synth_close", None),
            getattr(self, "_close_baseline_window", None),
            getattr(self, "_interp_close", None),
            getattr(self, "_norm_close", None),
            getattr(self, "_gauss_close", None),
            getattr(self, "_dsdf_close", None),
            getattr(self, "_unwrap_close", None),
            getattr(self, "_phase2_close", None),
            getattr(self, "_phase3_close", None),
            getattr(self, "_res_close", None),
        ):
            if callable(closer):
                closer()

        created_at = datetime.now().isoformat(timespec="seconds")
        self.dataset = Dataset(
            source_file=str(DEFAULT_DATASET_FILE.resolve()),
            dataset_name=cleaned_name,
            created_at=created_at,
        )
        self.dataset_path = _dataset_pickle_path(self.dataset).resolve()
        self.dataset.source_file = str(self.dataset_path)
        _write_app_state(self.dataset_path)
        if self.res_neighbor_dfrel_initial_date_var is not None:
            self.res_neighbor_dfrel_initial_date_var.set(self._dataset_res_neighbor_initial_date())
        self._reload_transcript_ui()
        self._mark_clean()
        self._refresh_status()
        self._log(f"Started new empty dataset: {cleaned_name}")

    def rename_dataset_prefix(self) -> None:
        if not self.dataset.dataset_name or not self.dataset.created_at:
            messagebox.showwarning(
                "Rename unavailable",
                "This dataset does not have a saved prefix yet. Create or save the dataset first.",
            )
            return

        current_name = str(self.dataset.dataset_name).strip()
        proposed_name = simpledialog.askstring(
            "Rename Dataset Prefix",
            "Enter the new prefix to use for the dataset folder and pickle filename:",
            initialvalue=current_name,
            parent=self.root,
        )
        if proposed_name is None:
            return

        cleaned_name = _safe_name(proposed_name)
        if not cleaned_name:
            messagebox.showwarning("Invalid prefix", "Please enter a non-empty dataset prefix.")
            return
        if cleaned_name == current_name:
            return

        old_dataset_name = current_name
        old_dataset_path = self.dataset_path.resolve()
        old_dataset_source = str(self.dataset.source_file)
        old_dataset_dir = old_dataset_path.parent.resolve()
        old_history_len = len(self.dataset.processing_history)
        old_dir_exists = old_dataset_dir.exists()
        moved_dir = False
        legacy_pickle_path: Optional[Path] = None

        try:
            self.dataset.dataset_name = cleaned_name
            new_dataset_path = _dataset_pickle_path(self.dataset).resolve()
            new_dataset_dir = new_dataset_path.parent.resolve()

            if new_dataset_dir.exists():
                raise FileExistsError(f"Target dataset folder already exists:\n{new_dataset_dir}")

            if old_dir_exists and old_dataset_dir != new_dataset_dir:
                shutil.move(str(old_dataset_dir), str(new_dataset_dir))
                moved_dir = True

            if moved_dir:
                legacy_pickle_path = (new_dataset_dir / old_dataset_path.name).resolve()
            else:
                legacy_pickle_path = old_dataset_path

            self.dataset_path = new_dataset_path
            self.dataset.source_file = str(new_dataset_path)
            self.dataset.processing_history.append(
                _make_event(
                    "rename_dataset_prefix",
                    {
                        "old_dataset_name": old_dataset_name,
                        "new_dataset_name": cleaned_name,
                        "old_dataset_path": str(old_dataset_path),
                        "new_dataset_path": str(new_dataset_path),
                    },
                )
            )
            _save_dataset(self.dataset, self.dataset_path)
            if (
                legacy_pickle_path is not None
                and legacy_pickle_path.exists()
                and legacy_pickle_path != self.dataset_path
            ):
                legacy_pickle_path.unlink()
            _write_app_state(self.dataset_path)
            self._mark_clean()
            self._refresh_status()
            self._log(f"Renamed dataset prefix from {old_dataset_name} to {cleaned_name}.")
        except Exception as exc:
            if moved_dir and new_dataset_dir.exists() and not old_dataset_dir.exists():
                try:
                    shutil.move(str(new_dataset_dir), str(old_dataset_dir))
                except Exception:
                    pass
            del self.dataset.processing_history[old_history_len:]
            self.dataset.dataset_name = old_dataset_name
            self.dataset_path = old_dataset_path
            self.dataset.source_file = old_dataset_source
            self._mark_dirty()
            self._refresh_status()
            self._log(f"Rename failed: {exc}")
            messagebox.showerror("Rename failed", str(exc))

    def _selected_scans_have_attached_filter(self) -> bool:
        scans = self._baseline_pipeline_selected_scans()
        if not scans:
            return False
        for scan in scans:
            bf = scan.baseline_filter
            if not isinstance(bf, dict):
                return False
            amp, phase = _read_polar_series(
                bf,
                amplitude_key="filtered_amp",
                phase_key="filtered_phase_deg",
            )
            if amp.ndim != 1 or phase.ndim != 1 or amp.shape != phase.shape:
                return False
            if amp.size == 0:
                return False
        return True

    def _configure_action_button(
        self,
        button: Optional[tk.Button],
        *,
        available: bool,
        done_count: int = 0,
        total_count: int = 0,
    ) -> None:
        if button is None:
            return
        if not available:
            button.configure(state="disabled", bg="light grey", activebackground="light grey")
            return
        if total_count > 0 and done_count >= total_count:
            bg = "medium sea green"
        elif done_count > 0:
            bg = "light green"
        else:
            bg = self._default_button_bg
        button.configure(state="normal", bg=bg, activebackground=bg)

    def _selected_progress_counts(self, done_check) -> tuple[list[VNAScan], int]:
        scans = self._selected_scans()
        return scans, sum(1 for scan in scans if done_check(scan))

    def _baseline_target_scans(self) -> list[VNAScan]:
        scans = self._selected_scans()
        if self._baseline_target_scan_keys_override is not None:
            override = self._baseline_target_scan_keys_override
            return [scan for scan in (scans if scans else list(self.dataset.vna_scans)) if self._scan_key(scan) in override]
        return scans if scans else list(self.dataset.vna_scans)

    @staticmethod
    def _scan_marked_omitted_from_baseline_fit(scan: VNAScan) -> bool:
        bf = scan.baseline_filter
        return isinstance(bf, dict) and bool(bf.get("omit_from_baseline_fit", False))

    def _baseline_pipeline_selected_scans(self) -> list[VNAScan]:
        scans = self._selected_scans()
        if not scans:
            return []
        if self._baseline_target_scan_keys_override is not None:
            override = self._baseline_target_scan_keys_override
            return [scan for scan in scans if self._scan_key(scan) in override]
        return [scan for scan in scans if not self._scan_marked_omitted_from_baseline_fit(scan)]

    def _baseline_pipeline_omitted_selected_scans(self) -> list[VNAScan]:
        scans = self._selected_scans()
        if not scans:
            return []
        if self._baseline_target_scan_keys_override is not None:
            override = self._baseline_target_scan_keys_override
            return [scan for scan in scans if self._scan_key(scan) not in override]
        return [scan for scan in scans if self._scan_marked_omitted_from_baseline_fit(scan)]

    def _has_valid_phase2_output(self, scan: VNAScan) -> bool:
        phase2 = scan.candidate_resonators.get("phase_correction_2")
        if not isinstance(phase2, dict):
            return False
        amp, phase = _read_polar_series(
            phase2,
            amplitude_key="corrected_amp",
            phase_key="corrected_phase_deg",
        )
        return amp.shape == scan.freq.shape and phase.shape == scan.freq.shape

    def _has_valid_phase3_output(self, scan: VNAScan) -> bool:
        phase3 = scan.candidate_resonators.get("phase_correction_3")
        if not isinstance(phase3, dict):
            return False
        amp, phase = _read_polar_series(
            phase3,
            amplitude_key="corrected_amp",
            phase_key="corrected_phase_deg",
        )
        return amp.shape == scan.freq.shape and phase.shape == scan.freq.shape

    def _has_valid_baseline_filter_output(self, scan: VNAScan) -> bool:
        bf = scan.baseline_filter
        if not isinstance(bf, dict):
            return False
        amp, phase = _read_polar_series(
            bf,
            amplitude_key="filtered_amp",
            phase_key="filtered_phase_deg",
        )
        if amp.ndim != 1 or phase.ndim != 1 or amp.shape != phase.shape or amp.size == 0:
            return False
        keep = np.asarray(bf.get("retained_mask", np.array([])), dtype=bool)
        if keep.ndim != 1 or keep.size != scan.freq.size:
            return False
        return int(np.count_nonzero(keep)) == int(amp.size)

    def _has_valid_interp_output(self, scan: VNAScan) -> bool:
        bf = scan.baseline_filter
        if not isinstance(bf, dict):
            return False
        interp = bf.get("interp_smooth")
        if not isinstance(interp, dict):
            return False
        amp, phase = _read_polar_series(
            interp,
            amplitude_key="interp_amp",
            phase_key="interp_phase",
        )
        return amp.shape == scan.freq.shape and phase.shape == scan.freq.shape

    def _has_valid_normalized_output(self, scan: VNAScan) -> bool:
        bf = scan.baseline_filter
        if not isinstance(bf, dict):
            return False
        norm = bf.get("normalized")
        if not isinstance(norm, dict):
            return False
        amp, phase = _read_polar_series(
            norm,
            amplitude_key="norm_amp",
            phase_key="norm_phase_deg_unwrapped",
        )
        return amp.shape == scan.freq.shape and phase.shape == scan.freq.shape

    @staticmethod
    def _has_valid_candidate_attachment(scan: VNAScan, key: str) -> bool:
        payload = scan.candidate_resonators.get(key)
        return isinstance(payload, dict) and len(payload) > 0

    def _update_interp_button_state(self) -> None:
        scans = self._baseline_pipeline_selected_scans()
        done_count = sum(1 for scan in scans if self._has_valid_interp_output(scan))
        self._configure_action_button(
            self.interp_button,
            available=bool(scans) and all(self._has_valid_baseline_filter_output(scan) for scan in scans),
            done_count=done_count,
            total_count=len(scans),
        )

    def _selected_scans_have_attached_interp_data(self) -> bool:
        scans = self._baseline_pipeline_selected_scans()
        if not scans:
            return False
        for scan in scans:
            bf = scan.baseline_filter
            if not isinstance(bf, dict):
                return False
            interp = bf.get("interp_smooth")
            if not isinstance(interp, dict):
                return False
            amp, phase = _read_polar_series(
                interp,
                amplitude_key="interp_amp",
                phase_key="interp_phase",
            )
            if amp.ndim != 1 or phase.ndim != 1 or amp.shape != phase.shape:
                return False
            if amp.size == 0:
                return False
            if amp.size != scan.freq.size:
                return False
        return True

    def _update_norm_button_state(self) -> None:
        scans = self._baseline_pipeline_selected_scans()
        done_count = sum(1 for scan in scans if self._has_valid_normalized_output(scan))
        self._configure_action_button(
            self.norm_button,
            available=bool(scans) and all(self._has_valid_interp_output(scan) for scan in scans),
            done_count=done_count,
            total_count=len(scans),
        )
        self._configure_action_button(
            self.norm_apply_large_button,
            available=bool(scans)
            and all(self._has_valid_phase3_output(scan) for scan in scans)
            and any(self._has_valid_interp_output(scan) for scan in scans),
        )

    def _selected_scans_have_attached_normalized(self) -> bool:
        scans = self._selected_scans()
        if not scans:
            return False
        for scan in scans:
            bf = scan.baseline_filter
            if not isinstance(bf, dict):
                return False
            norm = bf.get("normalized")
            if not isinstance(norm, dict):
                return False
            amp, phase = _read_polar_series(
                norm,
                amplitude_key="norm_amp",
                phase_key="norm_phase_deg_unwrapped",
            )
            if amp.ndim != 1 or phase.ndim != 1 or amp.shape != phase.shape:
                return False
            if amp.size != scan.freq.size:
                return False
        return True

    def _update_res_button_state(self) -> None:
        scans, done_count = self._selected_progress_counts(
            lambda scan: self._has_valid_candidate_attachment(scan, "resonance_selection_view")
        )
        self._configure_action_button(
            self.res_button,
            available=bool(scans) and all(self._has_valid_normalized_output(scan) for scan in scans),
            done_count=done_count,
            total_count=len(scans),
        )

    def _update_gauss_button_state(self) -> None:
        scans, done_count = self._selected_progress_counts(
            lambda scan: self._has_valid_candidate_attachment(scan, "gaussian_convolution")
        )
        self._configure_action_button(
            self.gauss_button,
            available=bool(scans) and all(self._has_valid_normalized_output(scan) for scan in scans),
            done_count=done_count,
            total_count=len(scans),
        )

    def _update_dsdf_button_state(self) -> None:
        scans, done_count = self._selected_progress_counts(
            lambda scan: self._has_valid_candidate_attachment(scan, "dsdf_gaussian_convolution")
        )
        self._configure_action_button(
            self.dsdf_button,
            available=bool(scans) and all(self._has_valid_normalized_output(scan) for scan in scans),
            done_count=done_count,
            total_count=len(scans),
        )

    def _update_unwrap_button_state(self) -> None:
        scans, done_count = self._selected_progress_counts(lambda scan: scan.has_dewrapped_phase())
        self._configure_action_button(
            self.unwrap_button,
            available=bool(scans),
            done_count=done_count,
            total_count=len(scans),
        )

    def _update_phase2_button_state(self) -> None:
        scans, done_count = self._selected_progress_counts(self._has_valid_phase2_output)
        self._configure_action_button(
            self.phase2_button,
            available=bool(scans) and all(scan.has_dewrapped_phase() for scan in scans),
            done_count=done_count,
            total_count=len(scans),
        )

    def _update_phase3_button_state(self) -> None:
        scans, done_count = self._selected_progress_counts(self._has_valid_phase3_output)
        self._configure_action_button(
            self.phase3_button,
            available=bool(scans) and all(self._has_valid_phase2_output(scan) for scan in scans),
            done_count=done_count,
            total_count=len(scans),
        )

    def _update_baseline_button_state(self) -> None:
        scans = self._baseline_target_scans()
        done_count = sum(1 for scan in scans if self._has_valid_baseline_filter_output(scan))
        self._configure_action_button(
            self.baseline_button,
            available=bool(scans) and all(self._has_valid_phase3_output(scan) for scan in scans),
            done_count=done_count,
            total_count=len(scans),
        )

    def _update_select_scans_button_state(self) -> None:
        total_count = len(self.dataset.vna_scans)
        done_count = len(self._selected_scans())
        self._configure_action_button(
            self.select_scans_button,
            available=total_count > 0,
            done_count=done_count,
            total_count=total_count,
        )

    def save_dataset(self) -> None:
        if not self._has_data_to_save():
            self._log("Save skipped: no data to save.")
            self._update_save_button_state()
            return
        if not self.dataset.dataset_name:
            proposed_name = simpledialog.askstring(
                "Name dataset",
                "Enter a dataset name:",
                parent=self.root,
            )
            if proposed_name is None:
                return
            cleaned_name = _safe_name(proposed_name)
            if not cleaned_name:
                messagebox.showwarning("Invalid name", "Please enter a non-empty dataset name.")
                return
            self.dataset.dataset_name = cleaned_name
            if not self.dataset.created_at:
                self.dataset.created_at = datetime.now().isoformat(timespec="seconds")

        self._persist_dataset()

    def load_different_dataset(self) -> None:
        path_text = filedialog.askopenfilename(
            title="Select dataset file",
            initialdir=str(DATASETS_DIR.resolve()),
            filetypes=[("Pickle files", "*.pkl"), ("All files", "*.*")],
        )
        if not path_text:
            return

        new_path = Path(path_text)
        try:
            self.dataset = _load_dataset(new_path)
            self.dataset_path = new_path.resolve()
            _write_app_state(self.dataset_path)
            if self.res_neighbor_dfrel_initial_date_var is not None:
                self.res_neighbor_dfrel_initial_date_var.set(self._dataset_res_neighbor_initial_date())
            self._reload_transcript_ui()
            self._mark_clean()
            self._refresh_status()
            self._log(f"Dataset loaded: {self.dataset_path}")
            messagebox.showinfo("Dataset loaded", f"Loaded dataset:\n{self.dataset_path}")
        except Exception as exc:
            self._log(f"Load failed: {exc}")
            messagebox.showerror("Load failed", str(exc))

    def load_vna_scan(self) -> None:
        mode = self._choose_vna_load_mode()
        if mode is None:
            return

        path_texts = filedialog.askopenfilenames(
            title="Select VNA scan file(s)",
            filetypes=[
                ("Supported VNA files", "*.npy *.txt *.dat *.csv *.s2p"),
                ("NumPy files", "*.npy"),
                ("Text files", "*.txt *.dat *.csv"),
                ("Touchstone S2P files", "*.s2p"),
                ("All files", "*.*"),
            ],
        )
        if not path_texts:
            return

        added_count = 0
        failed: List[str] = []
        warnings: List[str] = []

        paths = [Path(path_text) for path_text in path_texts]

        def _add_scan(scan: VNAScan) -> None:
            self.dataset.vna_scans.append(scan)
            self.dataset.selected_scan_keys.append(self._scan_key(scan))
            self.dataset.processing_history.append(
                _make_event(
                    "add_vna_scan_to_dataset",
                    {"filename": scan.filename, "loaded_at": scan.loaded_at},
                )
            )

        if mode == "autodetect":
            pair_handled = False
            if len(paths) == 2:
                try:
                    pair_scan, pair_warning = _try_load_vna_npy_pair(paths[0], paths[1])
                except Exception as exc:
                    failed.append(f"{paths[0].name} + {paths[1].name}: {exc}")
                    pair_handled = True
                else:
                    if pair_scan is not None:
                        _add_scan(pair_scan)
                        self._mark_dirty()
                        added_count += 1
                        if pair_warning:
                            warnings.append(pair_warning)
                        pair_handled = True

            if not pair_handled:
                for path in paths:
                    try:
                        scan, warning = _load_vna_file(path)
                        _add_scan(scan)
                        self._mark_dirty()
                        added_count += 1
                        if warning:
                            warnings.append(warning)
                    except Exception as exc:
                        failed.append(f"{path.name}: {exc}")
        else:
            for path in paths:
                try:
                    if path.suffix.lower() != ".npy":
                        raise ValueError("Legacy MHz/dB/deg loader supports only .npy files.")
                    scan = _load_vna_npy_mhz_db_deg(path)
                    warning = None
                    _add_scan(scan)
                    self._mark_dirty()
                    added_count += 1
                    if warning:
                        warnings.append(warning)
                except Exception as exc:
                    failed.append(f"{path.name}: {exc}")

        self._refresh_status()
        self._log(f"VNA load result: added={added_count}, failed={len(failed)}")
        for warning in warnings:
            self._log(f"VNA load warning: {warning}")
        if added_count > 0:
            self._autosave_dataset()

        if added_count > 0 and not failed:
            if warnings:
                messagebox.showwarning(
                    "VNA scans loaded with assumptions",
                    f"Added {added_count} scan(s). Dataset saved automatically.\n\n"
                    + "\n".join(warnings[:8]),
                )
                return
            messagebox.showinfo(
                "VNA scans loaded",
                f"Added {added_count} scan(s). Dataset saved automatically.",
            )
            return

        if added_count > 0 and failed:
            detail_lines = [f"Added {added_count} scan(s), {len(failed)} failed."]
            if warnings:
                detail_lines.extend(["", "Assumption warnings:"])
                detail_lines.extend(warnings[:8])
            detail_lines.extend(["", "Failures:"])
            detail_lines.extend(failed[:8])
            messagebox.showwarning("VNA scan load partial", "\n".join(detail_lines))
            return

        messagebox.showerror(
            "VNA scan load failed",
            "No files were loaded.\n\n" + "\n".join(failed[:8]),
        )

    def _choose_vna_load_mode(self) -> Optional[str]:
        dialog = tk.Toplevel(self.root)
        dialog.title("Load VNA Scan(s)")
        dialog.geometry("880x300")
        dialog.transient(self.root)
        dialog.grab_set()

        mode_var = tk.StringVar(value="autodetect")
        tk.Label(
            dialog,
            text="Choose a load method before selecting file(s):",
            anchor="w",
            justify="left",
        ).pack(anchor="w", padx=12, pady=(12, 8))

        autodetect_text = (
            "Autodetect (default):\n"
            "  - .npy: (3,N)/(N,3) [freq, real, imag]\n"
            "  - .npy: (2,N)/(N,2) [freq, complex]\n"
            "  - two 1D .npy files: [freq] + [complex]\n"
            "  - text: 3-col [freq_Hz, real, imag]\n"
            "  - text: 2-col [freq_MHz, amp_dB] (phase assumed 0)\n"
            "  - .s2p: S21 (RI/MA/DB), freq units HZ/KHZ/MHZ/GHZ"
        )
        legacy_text = (
            "Legacy explicit .npy format:\n"
            "  - (3,N)/(N,3) [freq_MHz, amp_dB, phase_deg]\n"
            "  - use this for files like Be241202p1_hybrid_vna_PE20241213.npy"
        )

        tk.Radiobutton(
            dialog,
            text=autodetect_text,
            value="autodetect",
            variable=mode_var,
            anchor="w",
            justify="left",
            wraplength=840,
        ).pack(anchor="w", padx=12, pady=(0, 8))
        tk.Radiobutton(
            dialog,
            text=legacy_text,
            value="legacy_mhz_db_deg",
            variable=mode_var,
            anchor="w",
            justify="left",
            wraplength=840,
        ).pack(anchor="w", padx=12, pady=(0, 8))

        selection: dict[str, Optional[str]] = {"mode": None}

        def _choose_and_close() -> None:
            picked = str(mode_var.get() or "autodetect").strip().lower()
            if picked not in {"autodetect", "legacy_mhz_db_deg"}:
                picked = "autodetect"
            selection["mode"] = picked
            dialog.destroy()

        button_row = tk.Frame(dialog)
        button_row.pack(fill="x", padx=12, pady=(8, 12))
        tk.Button(button_row, text="Cancel", width=12, command=dialog.destroy).pack(side="right")
        tk.Button(button_row, text="Select File(s)...", width=14, command=_choose_and_close).pack(
            side="right", padx=(0, 8)
        )

        self.root.wait_window(dialog)
        return selection["mode"]

    def remove_vna_scans(self) -> None:
        if not self.dataset.vna_scans:
            messagebox.showwarning("No data", "No VNA scans are loaded in this dataset.")
            return

        selector = tk.Toplevel(self.root)
        selector.title("Remove VNA Scan(s)")
        selector.geometry("820x420")
        selector.transient(self.root)
        selector.grab_set()

        tk.Label(selector, text="Select scan(s) to remove from this dataset:").pack(
            anchor="w", padx=10, pady=(10, 4)
        )
        listbox = tk.Listbox(selector, width=130, height=16, selectmode=tk.MULTIPLE)
        listbox.pack(fill="both", expand=True, padx=10, pady=4)

        for idx, scan in enumerate(self.dataset.vna_scans):
            label = self._scan_dialog_label(scan, index=idx, include_loaded_at=True)
            listbox.insert(tk.END, label)

        def do_remove() -> None:
            indices = sorted(listbox.curselection())
            if not indices:
                selector.destroy()
                return
            names = [Path(self.dataset.vna_scans[i].filename).name for i in indices]
            ok = messagebox.askyesno(
                "Confirm Remove",
                f"Remove {len(indices)} scan(s) from dataset?\n\n"
                + "\n".join(names[:8])
                + ("\n..." if len(names) > 8 else ""),
                parent=selector,
            )
            if not ok:
                return

            removed = []
            for i in reversed(indices):
                scan = self.dataset.vna_scans.pop(i)
                removed.append(scan)

            self.dataset.selected_scan_keys = [
                self._scan_key(scan)
                for scan in self.dataset.vna_scans
                if self._scan_key(scan) in set(self.dataset.selected_scan_keys)
            ]
            self.dataset.processing_history.append(
                _make_event(
                    "remove_vna_scans_from_dataset",
                    {
                        "removed_count": len(removed),
                        "removed_files": [s.filename for s in removed],
                    },
                )
            )
            self._mark_dirty()
            self._refresh_status()
            self._log(f"Removed {len(removed)} VNA scan(s) from dataset.")
            self._autosave_dataset()
            selector.destroy()

        btns = tk.Frame(selector)
        btns.pack(fill="x", padx=10, pady=(4, 10))
        tk.Button(btns, text="Cancel", command=selector.destroy).pack(side="right")
        tk.Button(btns, text="Remove Selected", command=do_remove).pack(side="right", padx=(0, 8))

    def group_selected_scans_for_plotting(self) -> None:
        scans = self._selected_scans()
        if not scans:
            messagebox.showwarning(
                "No selection",
                "No scans selected for analysis.\nUse 'Select Scans for Analysis' first.",
            )
            return

        selector = tk.Toplevel(self.root)
        selector.title("Group Selected Scans")
        selector.geometry("900x460")
        selector.transient(self.root)
        selector.grab_set()

        existing_groups = sorted(
            {int(scan.plot_group) for scan in self.dataset.vna_scans if scan.plot_group is not None}
        )
        prompt = (
            "Choose the subset of currently selected analysis scans to join into one plot group.\n"
            "Each scan can belong to only one group. Enter a positive integer to assign, or leave blank to clear."
        )
        if existing_groups:
            prompt += "\nExisting groups: " + ", ".join(str(g) for g in existing_groups)
        tk.Label(selector, text=prompt, anchor="w", justify="left", wraplength=860).pack(
            anchor="w", padx=10, pady=(10, 6)
        )

        group_frame = tk.Frame(selector)
        group_frame.pack(fill="x", padx=10, pady=(0, 6))
        tk.Label(group_frame, text="Group number:").pack(side="left")
        group_entry = tk.Entry(group_frame, width=12)
        group_entry.pack(side="left", padx=(6, 12))
        tk.Label(group_frame, text="Blank clears group for the chosen subset.").pack(side="left")

        listbox = tk.Listbox(selector, width=160, height=16, selectmode=tk.EXTENDED)
        listbox.pack(fill="both", expand=True, padx=10, pady=4)
        for idx, scan in enumerate(scans):
            label = self._scan_dialog_label(
                scan,
                index=idx,
                include_file_timestamp=True,
                include_loaded_at=True,
                include_group=True,
            )
            listbox.insert(tk.END, label)
            listbox.selection_set(idx)

        def do_apply() -> None:
            indices = sorted(listbox.curselection())
            if not indices:
                messagebox.showwarning("No subset selected", "Select at least one scan in the grouping window.", parent=selector)
                return

            text = group_entry.get().strip()
            new_group: Optional[int] = None
            if text:
                try:
                    new_group = int(text)
                except ValueError:
                    messagebox.showwarning("Invalid group", "Enter a positive integer group number.", parent=selector)
                    return
                if new_group <= 0:
                    messagebox.showwarning("Invalid group", "Enter a positive integer group number.", parent=selector)
                    return

            chosen_scans = [scans[i] for i in indices]
            changed = 0
            for scan in chosen_scans:
                if scan.plot_group != new_group:
                    scan.plot_group = new_group
                    changed += 1

            if changed == 0:
                self._log("Grouping unchanged for selected scan subset.")
                selector.destroy()
                return

            self.dataset.processing_history.append(
                _make_event(
                    "group_selected_scans_for_plotting",
                    {
                        "selected_count": len(chosen_scans),
                        "plot_group": new_group,
                        "filenames": [scan.filename for scan in chosen_scans],
                    },
                )
            )
            self._mark_dirty()
            self._refresh_status()
            self._autosave_dataset()
            if new_group is None:
                self._log(f"Cleared plot groups for {changed} selected scan(s).")
                messagebox.showinfo("Plot groups updated", f"Cleared plot groups for {changed} selected scan(s).", parent=selector)
            else:
                self._log(f"Assigned plot group {new_group} to {changed} selected scan(s).")
                messagebox.showinfo(
                    "Plot groups updated",
                    f"Assigned plot group {new_group} to {changed} selected scan(s).",
                    parent=selector,
                )
            selector.destroy()

        btns = tk.Frame(selector)
        btns.pack(fill="x", padx=10, pady=(4, 10))
        tk.Button(btns, text="Select All", command=lambda: listbox.select_set(0, tk.END)).pack(side="left")
        tk.Button(btns, text="Clear Selection", command=lambda: listbox.selection_clear(0, tk.END)).pack(
            side="left", padx=(6, 0)
        )
        tk.Button(btns, text="Cancel", command=selector.destroy).pack(side="right")
        tk.Button(btns, text="Apply Group", command=do_apply).pack(side="right", padx=(0, 8))

    def clear_selected_scan_attachments(self) -> None:
        scans = self._selected_scans()
        if not scans:
            messagebox.showwarning(
                "No selection",
                "No scans selected for analysis.\nUse 'Select Scans for Analysis' first.",
            )
            return

        selector = tk.Toplevel(self.root)
        selector.title("Clear Attachments")
        selector.geometry("920x480")
        selector.transient(self.root)
        selector.grab_set()

        prompt = (
            "Choose the subset of selected analysis scans to reset.\n"
            "This returns each chosen scan to its original loaded state, while keeping plot group membership."
        )
        tk.Label(selector, text=prompt, anchor="w", justify="left", wraplength=880).pack(
            anchor="w", padx=10, pady=(10, 6)
        )

        listbox = tk.Listbox(selector, width=145, height=18, selectmode=tk.MULTIPLE)
        listbox.pack(fill="both", expand=True, padx=10, pady=4)
        for idx, scan in enumerate(scans):
            label = self._scan_dialog_label(
                scan,
                index=idx,
                include_loaded_at=True,
                include_group=True,
            )
            listbox.insert(tk.END, label)

        def do_clear() -> None:
            indices = sorted(listbox.curselection())
            if not indices:
                messagebox.showwarning(
                    "No subset selected",
                    "Select at least one scan in the clear-attachments window.",
                    parent=selector,
                )
                return

            chosen_scans = [scans[i] for i in indices]
            names = [Path(scan.filename).name for scan in chosen_scans]
            ok = messagebox.askyesno(
                "Confirm Clear",
                f"Clear attachments for {len(chosen_scans)} selected scan(s)?\n\n"
                + "\n".join(names[:8])
                + ("\n..." if len(names) > 8 else ""),
                parent=selector,
            )
            if not ok:
                return

            cleared = 0
            for scan in chosen_scans:
                load_events = [
                    event
                    for event in scan.processing_history
                    if str(event.get("action", "")).startswith("load_vna_")
                ]
                scan.s21_phase_deg_unwrapped = None
                scan.baseline_filter = {}
                scan.candidate_resonators = {}
                scan.processing_history = load_events[:1]
                cleared += 1

            self.dataset.processing_history.append(
                _make_event(
                    "clear_selected_scan_attachments",
                    {
                        "selected_count": len(chosen_scans),
                        "filenames": [scan.filename for scan in chosen_scans],
                    },
                )
            )
            self._mark_dirty()
            self._refresh_status()
            self._autosave_dataset()
            self._log(f"Cleared attachments for {cleared} selected scan(s).")
            messagebox.showinfo(
                "Attachments cleared",
                f"Cleared attachments for {cleared} selected scan(s).\nPlot groups were preserved.",
                parent=selector,
            )
            selector.destroy()

        btns = tk.Frame(selector)
        btns.pack(fill="x", padx=10, pady=(4, 10))
        tk.Button(btns, text="Select All", command=lambda: listbox.select_set(0, tk.END)).pack(side="left")
        tk.Button(btns, text="Clear Selection", command=lambda: listbox.selection_clear(0, tk.END)).pack(
            side="left", padx=(6, 0)
        )
        tk.Button(btns, text="Cancel", command=selector.destroy).pack(side="right")
        tk.Button(btns, text="Clear Attachments", command=do_clear).pack(side="right", padx=(0, 8))

    def open_resonance_sheet_loader(self) -> None:
        path_text = filedialog.askopenfilename(
            parent=self.root,
            title="Select resonance spreadsheet",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
        )
        if not path_text:
            return
        sheet_path = Path(path_text)
        if not sheet_path.exists():
            messagebox.showwarning("Missing file", "Select a valid spreadsheet file first.", parent=self.root)
            return
        try:
            loaded_count = self._load_resonances_from_sheet(sheet_path)
        except Exception as exc:
            self._log(f"Load resonators from sheet failed: {exc}")
            messagebox.showerror("Load failed", str(exc), parent=self.root)
            return
        messagebox.showinfo(
            "Resonators loaded",
            f"Loaded {loaded_count} resonator assignment(s) from:\n{sheet_path}",
            parent=self.root,
        )

    def open_resonance_sheet_saver(self) -> None:
        path_text = filedialog.asksaveasfilename(
            parent=self.root,
            title="Save resonance spreadsheet",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not path_text:
            return
        sheet_path = Path(path_text)
        if str(sheet_path).strip() == "":
            messagebox.showwarning("Missing file", "Select an output spreadsheet path first.", parent=self.root)
            return
        try:
            saved_count = self._save_resonances_to_sheet(sheet_path)
        except Exception as exc:
            self._log(f"Save resonators to sheet failed: {exc}")
            messagebox.showerror("Save failed", str(exc), parent=self.root)
            return
        messagebox.showinfo(
            "Resonators saved",
            f"Saved {saved_count} resonator assignment(s) to:\n{sheet_path}",
            parent=self.root,
        )

    def open_attached_resonance_plotter(self) -> None:
        dialog = tk.Toplevel(self.root)
        dialog.title("Plot Resonator Markers")
        dialog.geometry("760x150")
        dialog.transient(self.root)
        dialog.grab_set()
        status_var = tk.StringVar(value="Ready to generate PDF.")

        generate_button = None

        def run() -> None:
            if generate_button is not None:
                generate_button.configure(state="disabled")
            status_var.set("Preparing pages...")
            dialog.update_idletasks()

            def on_progress(completed: int, total: int) -> None:
                status_var.set(f"Generating page {completed} of {total}...")
                dialog.update_idletasks()

            try:
                saved = self._plot_attached_resonances(progress_callback=on_progress)
            except Exception as exc:
                if generate_button is not None:
                    generate_button.configure(state="normal")
                status_var.set("PDF generation failed.")
                self._log(f"Plot resonator markers failed: {exc}")
                messagebox.showerror("Plot failed", str(exc), parent=dialog)
                return
            if not saved:
                if generate_button is not None:
                    generate_button.configure(state="normal")
                status_var.set("No PDF was generated.")
                messagebox.showwarning("No plots saved", "No plot files were generated.", parent=dialog)
                return
            out_dir = saved[0].parent
            status_var.set("PDF generation complete.")
            self._log(f"Plotted resonator markers into {out_dir}")
            messagebox.showinfo(
                "Plots saved",
                f"Saved {len(saved)} plot file(s) to:\n{out_dir}",
                parent=dialog,
            )
            dialog.destroy()

        top = tk.Frame(dialog, padx=10, pady=10)
        top.pack(fill="both", expand=True)
        tk.Label(
            top,
            text="Generate a single PDF using the same grouped resonator-marker view as the editor: one full-span page plus overlapping 100 MHz zoom pages.",
            anchor="w",
            justify="left",
            wraplength=720,
        ).pack(anchor="w", pady=(0, 8))
        tk.Label(top, textvariable=status_var, anchor="w", justify="left").pack(anchor="w")

        btns = tk.Frame(top)
        btns.pack(fill="x", pady=(12, 0))
        tk.Button(btns, text="Cancel", width=12, command=dialog.destroy).pack(side="right")
        generate_button = tk.Button(btns, text="Generate PDF", width=14, command=run)
        generate_button.pack(side="right", padx=(0, 8))

    @staticmethod
    def _resonator_shift_test_sort_stamp(scan: VNAScan) -> str:
        text = str(getattr(scan, "file_timestamp", "") or "").strip()
        if text:
            return text
        return str(getattr(scan, "loaded_at", "") or "").strip()

    @staticmethod
    def _resonator_shift_test_date_label(scan: VNAScan) -> str:
        stamp = DataAnalysisGUI._resonator_shift_test_sort_stamp(scan)
        if stamp:
            return stamp.split("T", 1)[0]
        return "unknown date"

    @staticmethod
    def _resonator_shift_parse_timestamp(text: object) -> Optional[datetime]:
        stamp = str(text or "").strip()
        if not stamp:
            return None
        try:
            return datetime.fromisoformat(stamp)
        except Exception:
            pass
        for fmt in ("%Y-%m-%d", "%Y%m%d_%H%M%S", "%Y%m%d"):
            try:
                return datetime.strptime(stamp, fmt)
            except Exception:
                continue
        return None

    def _resonator_shift_test_units(self) -> list[dict]:
        scans = self._selected_scans()
        if not scans:
            raise ValueError("No scans are selected for analysis.")

        units_by_key: dict[tuple[str, object], dict] = {}
        ordered_keys: list[tuple[str, object]] = []
        for scan_index, scan in enumerate(scans):
            payload = scan.candidate_resonators.get("sheet_resonances")
            if not isinstance(payload, dict):
                continue
            assignments = payload.get("assignments")
            if not isinstance(assignments, dict):
                continue
            if scan.plot_group is not None:
                unit_key: tuple[str, object] = ("group", int(scan.plot_group))
                detail_label = f"Group {int(scan.plot_group)}"
            else:
                unit_key = ("scan", self._scan_key(scan))
                detail_label = Path(scan.filename).name
            unit = units_by_key.get(unit_key)
            if unit is None:
                sort_stamp = self._resonator_shift_test_sort_stamp(scan)
                unit = {
                    "key": unit_key,
                    "sort_stamp": sort_stamp,
                    "timestamp_dt": self._resonator_shift_parse_timestamp(sort_stamp),
                    "order_index": scan_index,
                    "date_label": self._resonator_shift_test_date_label(scan),
                    "detail_label": detail_label,
                    "resonator_lists": {},
                }
                units_by_key[unit_key] = unit
                ordered_keys.append(unit_key)
            else:
                sort_stamp = self._resonator_shift_test_sort_stamp(scan)
                if sort_stamp and (not unit["sort_stamp"] or sort_stamp < unit["sort_stamp"]):
                    unit["sort_stamp"] = sort_stamp
                    unit["timestamp_dt"] = self._resonator_shift_parse_timestamp(sort_stamp)
                    unit["date_label"] = self._resonator_shift_test_date_label(scan)

            freq = np.asarray(scan.freq, dtype=float)
            if freq.size == 0:
                continue
            freq_min = float(np.min(freq))
            freq_max = float(np.max(freq))
            resonator_lists = unit["resonator_lists"]
            for resonator_number, record in assignments.items():
                if not isinstance(record, dict):
                    continue
                try:
                    target_hz = float(record.get("frequency_hz"))
                except Exception:
                    continue
                if not np.isfinite(target_hz) or not (freq_min <= target_hz <= freq_max):
                    continue
                resonator_label = str(resonator_number).strip()
                if not resonator_label:
                    continue
                resonator_lists.setdefault(resonator_label, []).append(target_hz)

        units: list[dict] = []
        for unit_key in ordered_keys:
            unit = units_by_key[unit_key]
            resonators: dict[str, float] = {}
            for resonator_label, values in unit["resonator_lists"].items():
                arr = np.asarray(values, dtype=float)
                arr = arr[np.isfinite(arr)]
                if arr.size:
                    resonators[resonator_label] = float(np.mean(arr))
            if not resonators:
                continue
            units.append(
                {
                    "key": unit["key"],
                    "sort_stamp": str(unit["sort_stamp"]),
                    "timestamp_dt": unit.get("timestamp_dt"),
                    "order_index": int(unit["order_index"]),
                    "date_label": str(unit["date_label"]),
                    "detail_label": str(unit["detail_label"]),
                    "label": f"{unit['date_label']} | {unit['detail_label']}",
                    "resonators": resonators,
                }
            )

        units.sort(key=lambda item: (str(item["sort_stamp"]), int(item["order_index"])))
        return units

    @staticmethod
    def _resonator_neighbor_parse_initial_date(text: str) -> Optional[datetime]:
        date_text = str(text).strip()
        if not date_text:
            return None
        try:
            return datetime.strptime(date_text, "%Y-%m-%d")
        except ValueError as exc:
            raise ValueError("Initial date must use YYYY-MM-DD.") from exc

    def _resonator_neighbor_dfrel_data(
        self,
        max_neighbor_sep_rel: float,
        initial_date_text: str = "",
    ) -> dict:
        tests = self._resonator_shift_test_units()
        if len(tests) < 2:
            raise ValueError("At least two selected test dates with marked resonators are required.")

        dated_tests = [test for test in tests if test.get("timestamp_dt") is not None]
        if len(dated_tests) < 2:
            raise ValueError("At least two selected test dates need valid file timestamps.")

        initial_time = self._resonator_neighbor_parse_initial_date(initial_date_text)
        base_time = initial_time if initial_time is not None else min(test["timestamp_dt"] for test in dated_tests)
        for test in tests:
            timestamp_dt = test.get("timestamp_dt")
            test["elapsed_days"] = (
                float((timestamp_dt - base_time).total_seconds()) / 86400.0
                if isinstance(timestamp_dt, datetime)
                else np.nan
            )

        values_by_resonator: dict[str, list[float]] = {}
        for test in tests:
            for resonator_label, freq_hz in test["resonators"].items():
                values_by_resonator.setdefault(resonator_label, []).append(float(freq_hz))

        mean_freq_by_resonator: dict[str, float] = {}
        for resonator_label, values in values_by_resonator.items():
            arr = np.asarray(values, dtype=float)
            arr = arr[np.isfinite(arr)]
            if arr.size:
                mean_freq_by_resonator[resonator_label] = float(np.mean(arr))

        ordered_labels = sorted(
            mean_freq_by_resonator,
            key=lambda label: (mean_freq_by_resonator[label], self._resonator_sort_key(label)),
        )
        if len(ordered_labels) < 2:
            raise ValueError("Need at least two marked resonators to build neighboring pairs.")

        threshold_rel = max(0.0, float(max_neighbor_sep_rel))
        pair_series: list[dict] = []
        for idx in range(len(ordered_labels) - 1):
            low_label = ordered_labels[idx]
            high_label = ordered_labels[idx + 1]
            low_mean_hz = float(mean_freq_by_resonator[low_label])
            high_mean_hz = float(mean_freq_by_resonator[high_label])
            pair_mean_hz = 0.5 * (low_mean_hz + high_mean_hz)
            pair_sep_hz = high_mean_hz - low_mean_hz
            pair_sep_rel = pair_sep_hz / pair_mean_hz if pair_mean_hz != 0.0 else np.nan
            if not np.isfinite(pair_sep_rel) or pair_sep_rel < 0.0 or pair_sep_rel > threshold_rel:
                continue
            points: list[dict] = []
            for test in tests:
                elapsed_days = float(test.get("elapsed_days", np.nan))
                if not np.isfinite(elapsed_days):
                    continue
                test_resonators = test["resonators"]
                if low_label not in test_resonators or high_label not in test_resonators:
                    continue
                low_freq_hz = float(test_resonators[low_label])
                high_freq_hz = float(test_resonators[high_label])
                df_hz = high_freq_hz - low_freq_hz
                if not np.isfinite(df_hz) or pair_mean_hz == 0.0:
                    continue
                points.append(
                    {
                        "elapsed_days": elapsed_days,
                        "df_over_f": float(df_hz / pair_mean_hz),
                        "df_hz": df_hz,
                        "low_freq_hz": low_freq_hz,
                        "high_freq_hz": high_freq_hz,
                        "test_label": test["label"],
                    }
                )
            if len(points) < 2:
                continue
            points.sort(key=lambda item: float(item["elapsed_days"]))
            pair_series.append(
                {
                    "low_label": low_label,
                    "high_label": high_label,
                    "label": f"{low_label}-{high_label}",
                    "mean_freq_hz": pair_mean_hz,
                    "mean_sep_hz": pair_sep_hz,
                    "mean_sep_rel": float(pair_sep_rel),
                    "points": points,
                }
            )

        if not pair_series:
            raise ValueError(
                "No adjacent resonator pairs met the relative df/f threshold with at least two dated measurements."
            )

        pair_series.sort(
            key=lambda item: (float(item["mean_freq_hz"]), self._resonator_sort_key(str(item["label"])))
        )
        mean_pair_freqs_hz = np.asarray([float(item["mean_freq_hz"]) for item in pair_series], dtype=float)
        return {
            "tests": tests,
            "pair_series": pair_series,
            "mean_pair_freqs_hz": mean_pair_freqs_hz,
            "threshold_rel": float(max_neighbor_sep_rel),
            "elapsed_time_origin": base_time,
            "elapsed_time_origin_source": "initial_date" if initial_time is not None else "first_dataset",
        }

    @staticmethod
    def _resonator_neighbor_pair_colors(data: dict) -> tuple[mcolors.Normalize, object]:
        mean_pair_freqs_hz = np.asarray(data["mean_pair_freqs_hz"], dtype=float)
        vmin = float(np.min(mean_pair_freqs_hz))
        vmax = float(np.max(mean_pair_freqs_hz))
        if not np.isfinite(vmin) or not np.isfinite(vmax):
            raise ValueError("Could not determine pair mean frequencies for coloring.")
        if vmax <= vmin:
            vmax = vmin + 1.0
        return mcolors.Normalize(vmin=vmin, vmax=vmax), plt.cm.get_cmap("rainbow_r")

    def _resonator_neighbor_scan_overlay_state(self, threshold_rel: float, initial_date_text: str = "") -> dict:
        data = self._resonator_neighbor_dfrel_data(threshold_rel, initial_date_text=initial_date_text)
        norm, cmap = self._resonator_neighbor_pair_colors(data)
        pair_series = data["pair_series"]

        pair_color_by_label: dict[str, object] = {}
        resonator_color_lists: dict[str, list[np.ndarray]] = {}
        for pair in pair_series:
            color = cmap(norm(float(pair["mean_freq_hz"])))
            pair["color"] = color
            pair_color_by_label[str(pair["label"])] = color
            for resonator_label in (str(pair["low_label"]), str(pair["high_label"])):
                resonator_color_lists.setdefault(resonator_label, []).append(np.asarray(color[:3], dtype=float))

        resonator_colors: dict[str, tuple[float, float, float, float]] = {}
        for resonator_label, colors in resonator_color_lists.items():
            if not colors:
                continue
            mean_rgb = np.mean(np.vstack(colors), axis=0)
            resonator_colors[resonator_label] = (
                float(np.clip(mean_rgb[0], 0.0, 1.0)),
                float(np.clip(mean_rgb[1], 0.0, 1.0)),
                float(np.clip(mean_rgb[2], 0.0, 1.0)),
                1.0,
            )

        return {
            "data": data,
            "norm": norm,
            "cmap": cmap,
            "pair_color_by_label": pair_color_by_label,
            "resonator_colors": resonator_colors,
        }

    @staticmethod
    def _resonator_neighbor_summary_by_time(pair_series: list[dict]) -> list[dict]:
        values_by_time: dict[float, list[float]] = {}
        for pair in pair_series:
            for point in pair["points"]:
                elapsed_days = float(point["elapsed_days"])
                df_over_f = float(point.get("plot_df_over_f", point["df_over_f"]))
                if not np.isfinite(elapsed_days) or not np.isfinite(df_over_f):
                    continue
                values_by_time.setdefault(elapsed_days, []).append(df_over_f)

        summary: list[dict] = []
        for elapsed_days in sorted(values_by_time):
            arr = np.asarray(values_by_time[elapsed_days], dtype=float)
            arr = arr[np.isfinite(arr)]
            if arr.size == 0:
                continue
            mean_value = float(np.mean(arr))
            std_value = float(np.std(arr))
            q1, median, q3 = np.percentile(arr, [25.0, 50.0, 75.0])
            iqr = float(q3 - q1)
            lower_bound = float(q1 - 1.5 * iqr)
            upper_bound = float(q3 + 1.5 * iqr)
            whisker_low = float(np.min(arr[arr >= lower_bound])) if np.any(arr >= lower_bound) else float(np.min(arr))
            whisker_high = float(np.max(arr[arr <= upper_bound])) if np.any(arr <= upper_bound) else float(np.max(arr))
            outliers = arr[(arr < whisker_low) | (arr > whisker_high)]
            summary.append(
                {
                    "elapsed_days": float(elapsed_days),
                    "count": int(arr.size),
                    "mean": mean_value,
                    "std": std_value,
                    "lower": float(mean_value - std_value),
                    "upper": float(mean_value + std_value),
                    "q1": float(q1),
                    "median": float(median),
                    "q3": float(q3),
                    "whisker_low": whisker_low,
                    "whisker_high": whisker_high,
                    "outliers": np.asarray(outliers, dtype=float),
                }
            )
        return summary

    @staticmethod
    def _resonator_neighbor_plot_series(pair_series: list[dict], mode: str) -> list[dict]:
        mode_key = str(mode).strip().lower()
        plotted: list[dict] = []
        for pair in pair_series:
            raw_points = pair.get("points", [])
            if not raw_points:
                continue
            baseline = float(raw_points[0]["df_over_f"])
            points: list[dict] = []
            for point in raw_points:
                y_value = float(point["df_over_f"])
                if mode_key == "change":
                    y_value -= baseline
                updated = dict(point)
                updated["plot_df_over_f"] = float(y_value)
                points.append(updated)
            updated_pair = dict(pair)
            updated_pair["points"] = points
            plotted.append(updated_pair)
        return plotted

    @staticmethod
    def _resonator_neighbor_drift_rate_summary(pair_series: list[dict]) -> list[dict]:
        values_by_time: dict[float, list[float]] = {}
        for pair in pair_series:
            points = pair.get("points", [])
            if len(points) < 2:
                continue
            for prev_point, point in zip(points[:-1], points[1:]):
                elapsed_prev = float(prev_point["elapsed_days"])
                elapsed_days = float(point["elapsed_days"])
                y_prev = float(prev_point.get("plot_df_over_f", prev_point["df_over_f"]))
                y_value = float(point.get("plot_df_over_f", point["df_over_f"]))
                delta_days = elapsed_days - elapsed_prev
                if (
                    not np.isfinite(elapsed_prev)
                    or not np.isfinite(elapsed_days)
                    or not np.isfinite(y_prev)
                    or not np.isfinite(y_value)
                    or delta_days <= 0.0
                ):
                    continue
                values_by_time.setdefault(elapsed_days, []).append(float((y_value - y_prev) / delta_days))

        summary: list[dict] = []
        for elapsed_days in sorted(values_by_time):
            arr = np.asarray(values_by_time[elapsed_days], dtype=float)
            arr = arr[np.isfinite(arr)]
            if arr.size == 0:
                continue
            mean_value = float(np.mean(arr))
            std_value = float(np.std(arr))
            summary.append(
                {
                    "elapsed_days": float(elapsed_days),
                    "count": int(arr.size),
                    "mean": mean_value,
                    "std": std_value,
                    "lower": float(mean_value - std_value),
                    "upper": float(mean_value + std_value),
                }
            )
        return summary

    @staticmethod
    def _resonator_neighbor_drift_rate_series(pair_series: list[dict]) -> list[dict]:
        drift_series: list[dict] = []
        for pair in pair_series:
            points = pair.get("points", [])
            if len(points) < 2:
                continue
            drift_points: list[dict] = []
            for prev_point, point in zip(points[:-1], points[1:]):
                elapsed_prev = float(prev_point["elapsed_days"])
                elapsed_days = float(point["elapsed_days"])
                y_prev = float(prev_point.get("plot_df_over_f", prev_point["df_over_f"]))
                y_value = float(point.get("plot_df_over_f", point["df_over_f"]))
                delta_days = elapsed_days - elapsed_prev
                if (
                    not np.isfinite(elapsed_prev)
                    or not np.isfinite(elapsed_days)
                    or not np.isfinite(y_prev)
                    or not np.isfinite(y_value)
                    or delta_days <= 0.0
                ):
                    continue
                drift_points.append(
                    {
                        "elapsed_days": float(elapsed_days),
                        "drift_rate": float((y_value - y_prev) / delta_days),
                    }
                )
            if not drift_points:
                continue
            updated_pair = dict(pair)
            updated_pair["drift_points"] = drift_points
            drift_series.append(updated_pair)
        return drift_series

    @staticmethod
    def _resonator_neighbor_interval_drift_data(data: dict) -> list[dict]:
        tests = list(data.get("tests", []))
        pair_series = list(data.get("pair_series", []))
        if len(tests) < 2 or not pair_series:
            return []

        intervals: list[dict] = []
        for prev_test, test in zip(tests[:-1], tests[1:]):
            elapsed_prev = float(prev_test.get("elapsed_days", np.nan))
            elapsed_days = float(test.get("elapsed_days", np.nan))
            delta_days = elapsed_days - elapsed_prev
            if (
                not np.isfinite(elapsed_prev)
                or not np.isfinite(elapsed_days)
                or delta_days <= 0.0
            ):
                continue
            intervals.append(
                {
                    "start_label": str(prev_test.get("label", "")),
                    "end_label": str(test.get("label", "")),
                    "start_elapsed_days": elapsed_prev,
                    "end_elapsed_days": elapsed_days,
                    "mid_elapsed_days": float(0.5 * (elapsed_prev + elapsed_days)),
                    "delta_days": float(delta_days),
                    "pair_drift_by_label": {},
                }
            )

        if not intervals:
            return []

        for pair in pair_series:
            pair_label = str(pair.get("label", ""))
            point_by_test_label = {
                str(point.get("test_label", "")): point
                for point in pair.get("points", [])
                if str(point.get("test_label", "")).strip()
            }
            for interval in intervals:
                prev_point = point_by_test_label.get(interval["start_label"])
                point = point_by_test_label.get(interval["end_label"])
                if prev_point is None or point is None:
                    continue
                y_prev = float(prev_point["df_over_f"])
                y_value = float(point["df_over_f"])
                delta_days = float(interval["delta_days"])
                if (
                    not np.isfinite(y_prev)
                    or not np.isfinite(y_value)
                    or delta_days <= 0.0
                ):
                    continue
                interval["pair_drift_by_label"][pair_label] = float((y_value - y_prev) / delta_days)
        return intervals

    @staticmethod
    def _resonator_neighbor_pair_correlation(x_values: np.ndarray, y_values: np.ndarray) -> float:
        x = np.asarray(x_values, dtype=float)
        y = np.asarray(y_values, dtype=float)
        mask = np.isfinite(x) & np.isfinite(y)
        x = x[mask]
        y = y[mask]
        if x.size == 0:
            return np.nan
        if x.size == 1:
            return 1.0 if np.allclose(x, y, rtol=0.0, atol=0.0) else np.nan
        x_centered = x - float(np.mean(x))
        y_centered = y - float(np.mean(y))
        denom = float(np.linalg.norm(x_centered) * np.linalg.norm(y_centered))
        if denom <= 0.0:
            return 1.0 if np.allclose(x, y, rtol=0.0, atol=0.0) else np.nan
        return float(np.dot(x_centered, y_centered) / denom)

    def _resonator_neighbor_self_correlation_summary(self, data: dict) -> dict:
        intervals = self._resonator_neighbor_interval_drift_data(data)
        if not intervals:
            raise ValueError("Need at least one consecutive dated interval to analyze pair drift correlation.")

        reference_idx = next(
            (
                idx
                for idx, interval in enumerate(intervals)
                if len(interval["pair_drift_by_label"]) >= 2
            ),
            None,
        )
        if reference_idx is None:
            raise ValueError("Need at least two neighboring pairs on one interval to compute correlation.")

        reference_interval = intervals[reference_idx]
        reference_drift = dict(reference_interval["pair_drift_by_label"])
        pair_mean_freq_by_label = {
            str(pair.get("label", "")): float(pair.get("mean_freq_hz", np.nan))
            for pair in data.get("pair_series", [])
        }
        points: list[dict] = []
        for interval_idx, interval in enumerate(intervals):
            current_drift = dict(interval["pair_drift_by_label"])
            common_labels = sorted(
                set(reference_drift).intersection(current_drift),
                key=self._resonator_sort_key,
            )
            corr_value = np.nan
            q1_value = np.nan
            q3_value = np.nan
            contribution_by_label: dict[str, float] = {}
            if common_labels:
                x = np.asarray([float(reference_drift[label]) for label in common_labels], dtype=float)
                y = np.asarray([float(current_drift[label]) for label in common_labels], dtype=float)
                corr_value = self._resonator_neighbor_pair_correlation(x, y)
                if x.size >= 2:
                    x_centered = x - float(np.mean(x))
                    y_centered = y - float(np.mean(y))
                    x_std = float(np.std(x))
                    y_std = float(np.std(y))
                    if x_std > 0.0 and y_std > 0.0:
                        contributions = (x_centered / x_std) * (y_centered / y_std)
                        for label, value in zip(common_labels, contributions):
                            contribution_by_label[str(label)] = float(value)
                        q1_value, q3_value = np.percentile(contributions, [25.0, 75.0])
                if interval_idx == reference_idx and np.isfinite(corr_value):
                    corr_value = 1.0

            drift_arr = np.asarray(list(current_drift.values()), dtype=float)
            drift_arr = drift_arr[np.isfinite(drift_arr)]
            abs_arr = np.abs(drift_arr)
            mean_abs = float(np.mean(abs_arr)) if abs_arr.size else np.nan
            std_abs = float(np.std(abs_arr)) if abs_arr.size else np.nan
            rms = float(np.sqrt(np.mean(np.square(drift_arr)))) if drift_arr.size else np.nan

            points.append(
                {
                    "interval_idx": int(interval_idx),
                    "pair_count": int(len(current_drift)),
                    "common_pair_count": int(len(common_labels)),
                    "start_elapsed_days": float(interval["start_elapsed_days"]),
                    "mid_elapsed_days": float(interval["mid_elapsed_days"]),
                    "end_elapsed_days": float(interval["end_elapsed_days"]),
                    "delta_days": float(interval["delta_days"]),
                    "corr_to_initial": float(corr_value) if np.isfinite(corr_value) else np.nan,
                    "q1_contribution": float(q1_value) if np.isfinite(q1_value) else np.nan,
                    "q3_contribution": float(q3_value) if np.isfinite(q3_value) else np.nan,
                    "contribution_by_label": contribution_by_label,
                    "mean_abs_drift_rate": mean_abs,
                    "std_abs_drift_rate": std_abs,
                    "rms_drift_rate": rms,
                }
            )

        return {
            "intervals": intervals,
            "points": points,
            "reference_interval_idx": int(reference_idx),
            "reference_interval": reference_interval,
            "pair_mean_freq_by_label": pair_mean_freq_by_label,
        }

    def _resonator_neighbor_scan_control_values(self, source: str | None = None) -> tuple[float, str]:
        source_key = str(source or self._res_neighbor_scan_source or "dfrel").strip().lower()
        if source_key == "corr":
            threshold_rel = (
                float(self.res_neighbor_corr_sep_rel_var.get())
                if self.res_neighbor_corr_sep_rel_var is not None
                else 0.004
            )
            initial_date_text = (
                str(self.res_neighbor_corr_initial_date_var.get())
                if self.res_neighbor_corr_initial_date_var is not None
                else self._dataset_res_neighbor_initial_date()
            )
            return threshold_rel, initial_date_text

        threshold_rel = (
            float(self.res_neighbor_dfrel_sep_rel_var.get())
            if self.res_neighbor_dfrel_sep_rel_var is not None
            else 0.004
        )
        initial_date_text = (
            str(self.res_neighbor_dfrel_initial_date_var.get())
            if self.res_neighbor_dfrel_initial_date_var is not None
            else self._dataset_res_neighbor_initial_date()
        )
        return threshold_rel, initial_date_text

    def open_resonator_neighbor_corr_window(self) -> None:
        if self.res_neighbor_corr_window is not None and self.res_neighbor_corr_window.winfo_exists():
            self.res_neighbor_corr_window.lift()
            self._render_resonator_neighbor_corr_window()
            return

        self.res_neighbor_corr_window = tk.Toplevel(self.root)
        self.res_neighbor_corr_window.title("Neighbor Pair Self Correlation vs Time")
        self.res_neighbor_corr_window.geometry("1380x900")
        self.res_neighbor_corr_window.protocol("WM_DELETE_WINDOW", self._close_resonator_neighbor_corr_window)

        controls = tk.Frame(self.res_neighbor_corr_window, padx=8, pady=8)
        controls.pack(side="top", fill="x")
        control_row = tk.Frame(controls)
        control_row.pack(side="top", fill="x", anchor="w")
        self.res_neighbor_corr_status_var = tk.StringVar(
            value="Showing correlation of each pair-drift interval with the initial interval."
        )
        self.res_neighbor_corr_sep_rel_var = tk.DoubleVar(value=0.004)
        self.res_neighbor_corr_initial_date_var = tk.StringVar(value=self._dataset_res_neighbor_initial_date())
        self.res_neighbor_corr_show_curves_var = tk.BooleanVar(value=True)

        tk.Label(control_row, text="Initial Date").pack(side="left", padx=(0, 4))
        initial_date_entry = tk.Entry(control_row, width=12, textvariable=self.res_neighbor_corr_initial_date_var)
        initial_date_entry.pack(side="left", padx=(0, 4))
        initial_date_entry.bind("<Return>", lambda _event: self._render_resonator_neighbor_corr_window())
        initial_date_entry.bind("<FocusOut>", lambda _event: self._render_resonator_neighbor_corr_window())
        tk.Label(control_row, text="YYYY-MM-DD").pack(side="left", padx=(0, 12))
        self.res_neighbor_corr_sep_scale = tk.Scale(
            control_row,
            from_=0.0,
            to=0.04,
            resolution=0.0001,
            orient="horizontal",
            length=260,
            digits=5,
            label="Max pair mean separation (df/f)",
            variable=self.res_neighbor_corr_sep_rel_var,
            command=lambda _value: self._render_resonator_neighbor_corr_window(),
        )
        self.res_neighbor_corr_sep_scale.pack(side="left", padx=(0, 12))
        tk.Checkbutton(
            control_row,
            text="Colored Curves",
            variable=self.res_neighbor_corr_show_curves_var,
            command=self._render_resonator_neighbor_corr_window,
        ).pack(side="left", padx=(0, 12))
        tk.Button(
            control_row,
            text="Show On Scans",
            width=13,
            command=lambda: self.open_resonator_neighbor_scan_window(source="corr"),
        ).pack(side="left", padx=(0, 8))
        tk.Button(control_row, text="Refresh", width=10, command=self._render_resonator_neighbor_corr_window).pack(
            side="left",
            padx=(0, 8),
        )
        tk.Label(controls, textvariable=self.res_neighbor_corr_status_var, anchor="w", justify="left").pack(
            side="top",
            fill="x",
            expand=True,
            pady=(6, 0),
        )

        self.res_neighbor_corr_figure = Figure(figsize=(12.5, 7.8))
        self.res_neighbor_corr_canvas = FigureCanvasTkAgg(
            self.res_neighbor_corr_figure,
            master=self.res_neighbor_corr_window,
        )
        self.res_neighbor_corr_toolbar = NavigationToolbar2Tk(
            self.res_neighbor_corr_canvas,
            self.res_neighbor_corr_window,
        )
        self.res_neighbor_corr_toolbar.update()
        self.res_neighbor_corr_toolbar.pack(side="top", fill="x")
        self.res_neighbor_corr_canvas.get_tk_widget().pack(fill="both", expand=True)

        self._render_resonator_neighbor_corr_window()

    def _close_resonator_neighbor_corr_window(self) -> None:
        if self.res_neighbor_corr_window is not None and self.res_neighbor_corr_window.winfo_exists():
            self.res_neighbor_corr_window.destroy()
        self.res_neighbor_corr_window = None
        self.res_neighbor_corr_canvas = None
        self.res_neighbor_corr_toolbar = None
        self.res_neighbor_corr_figure = None
        self.res_neighbor_corr_status_var = None
        self.res_neighbor_corr_sep_rel_var = None
        self.res_neighbor_corr_initial_date_var = None
        self.res_neighbor_corr_show_curves_var = None
        self.res_neighbor_corr_sep_scale = None
        self._res_neighbor_corr_axes = None

    def _render_resonator_neighbor_corr_window(self) -> None:
        if self.res_neighbor_corr_figure is None or self.res_neighbor_corr_canvas is None:
            return

        self.res_neighbor_corr_figure.clear()
        ax_corr = self.res_neighbor_corr_figure.add_subplot(211)
        ax_mag = self.res_neighbor_corr_figure.add_subplot(212, sharex=ax_corr)
        self._res_neighbor_corr_axes = (ax_corr, ax_mag)

        threshold_rel = (
            float(self.res_neighbor_corr_sep_rel_var.get())
            if self.res_neighbor_corr_sep_rel_var is not None
            else 0.004
        )
        initial_date_text = (
            str(self.res_neighbor_corr_initial_date_var.get())
            if self.res_neighbor_corr_initial_date_var is not None
            else ""
        )

        try:
            data = self._resonator_neighbor_dfrel_data(threshold_rel, initial_date_text=initial_date_text)
            summary = self._resonator_neighbor_self_correlation_summary(data)
        except Exception as exc:
            ax_corr.text(0.5, 0.5, str(exc), ha="center", va="center", transform=ax_corr.transAxes)
            ax_corr.set_axis_off()
            ax_mag.set_axis_off()
            if self.res_neighbor_corr_status_var is not None:
                self.res_neighbor_corr_status_var.set(str(exc))
            self.res_neighbor_corr_canvas.draw_idle()
            return

        points = summary["points"]
        x_end = np.asarray([float(item["end_elapsed_days"]) for item in points], dtype=float)
        y_corr = np.asarray([float(item["corr_to_initial"]) for item in points], dtype=float)
        q1_corr = np.asarray([float(item["q1_contribution"]) for item in points], dtype=float)
        q3_corr = np.asarray([float(item["q3_contribution"]) for item in points], dtype=float)
        pair_counts = np.asarray([int(item["pair_count"]) for item in points], dtype=int)
        common_counts = np.asarray([int(item["common_pair_count"]) for item in points], dtype=int)
        y_mag = np.asarray([float(item["mean_abs_drift_rate"]) for item in points], dtype=float)
        y_mag_std = np.asarray([float(item["std_abs_drift_rate"]) for item in points], dtype=float)
        norm, cmap = self._resonator_neighbor_pair_colors(data)
        pair_mean_freq_by_label = dict(summary.get("pair_mean_freq_by_label", {}))
        show_curves = (
            bool(self.res_neighbor_corr_show_curves_var.get())
            if self.res_neighbor_corr_show_curves_var is not None
            else True
        )

        contribution_series: dict[str, list[tuple[float, float]]] = {}
        for point in points:
            x_value = float(point["end_elapsed_days"])
            if not np.isfinite(x_value):
                continue
            contribution_by_label = point.get("contribution_by_label", {})
            if not isinstance(contribution_by_label, dict):
                continue
            for pair_label, value in contribution_by_label.items():
                y_value = float(value)
                if not np.isfinite(y_value):
                    continue
                contribution_series.setdefault(str(pair_label), []).append((x_value, y_value))

        q_mask = np.isfinite(x_end) & np.isfinite(q1_corr) & np.isfinite(q3_corr)
        if np.any(q_mask):
            ax_corr.fill_between(
                x_end[q_mask],
                q1_corr[q_mask],
                q3_corr[q_mask],
                color="0.2",
                alpha=0.22,
                linewidth=0.0,
                zorder=1,
                label="Middle 50%",
            )

        if show_curves:
            for pair_label, series in sorted(
                contribution_series.items(),
                key=lambda item: (
                    pair_mean_freq_by_label.get(item[0], np.nan),
                    self._resonator_sort_key(item[0]),
                ),
            ):
                pair_freq_hz = float(pair_mean_freq_by_label.get(pair_label, np.nan))
                color = cmap(norm(pair_freq_hz)) if np.isfinite(pair_freq_hz) else "0.55"
                xy = np.asarray(series, dtype=float)
                if xy.ndim != 2 or xy.shape[0] == 0:
                    continue
                order = np.argsort(xy[:, 0])
                ax_corr.plot(
                    xy[order, 0],
                    xy[order, 1],
                    color=color,
                    linewidth=1.2,
                    alpha=0.9,
                    marker="o",
                    markersize=3.5,
                    zorder=2,
                )

        corr_mask = np.isfinite(x_end) & np.isfinite(y_corr)
        if np.any(corr_mask):
            ax_corr.plot(
                x_end[corr_mask],
                y_corr[corr_mask],
                color="black",
                linewidth=4.0,
                marker="o",
                markersize=5.5,
                zorder=4,
                label="Mean correlation",
            )
        missing_mask = np.isfinite(x_end) & ~np.isfinite(y_corr)
        if np.any(missing_mask):
            ax_corr.plot(
                x_end[missing_mask],
                np.zeros(np.count_nonzero(missing_mask), dtype=float),
                linestyle="none",
                marker="x",
                markersize=6.0,
                color="tab:red",
                alpha=0.8,
                zorder=4,
            )

        ref_idx = int(summary["reference_interval_idx"])
        if 0 <= ref_idx < x_end.size and np.isfinite(x_end[ref_idx]):
            ax_corr.axvline(x_end[ref_idx], color="0.45", linestyle=":", linewidth=1.0, alpha=0.9)

        for x_value, y_value, common_count in zip(x_end, y_corr, common_counts):
            if not np.isfinite(x_value) or not np.isfinite(y_value):
                continue
            ax_corr.annotate(
                f"n={int(common_count)}",
                (float(x_value), float(y_value)),
                xytext=(0, 8),
                textcoords="offset points",
                ha="center",
                fontsize=8,
                alpha=0.85,
            )

        mag_mask = np.isfinite(x_end) & np.isfinite(y_mag)
        if np.any(mag_mask):
            lower = np.maximum(0.0, y_mag[mag_mask] - np.nan_to_num(y_mag_std[mag_mask], nan=0.0))
            upper = y_mag[mag_mask] + np.nan_to_num(y_mag_std[mag_mask], nan=0.0)
            ax_mag.fill_between(
                x_end[mag_mask],
                lower,
                upper,
                color="0.2",
                alpha=0.22,
                linewidth=0.0,
                zorder=1,
                label="Mean |drift| +/- 1 std",
            )
            ax_mag.plot(
                x_end[mag_mask],
                y_mag[mag_mask],
                color="black",
                linewidth=1.8,
                marker="o",
                markersize=5.0,
                zorder=3,
                label="Mean |drift|",
            )

        for x_value, y_value, pair_count in zip(x_end, y_mag, pair_counts):
            if not np.isfinite(x_value) or not np.isfinite(y_value):
                continue
            ax_mag.annotate(
                f"n={int(pair_count)}",
                (float(x_value), float(y_value)),
                xytext=(0, 8),
                textcoords="offset points",
                ha="center",
                fontsize=8,
                alpha=0.85,
            )

        ax_corr.axhline(0.0, color="0.5", linewidth=0.8, linestyle="--")
        ax_corr.set_ylim(-2.0, 2.0)
        ax_corr.set_ylabel("Correlation to Initial Interval")
        ax_corr.set_title("Neighbor Pair Drift Self Correlation")
        ax_corr.grid(True, alpha=0.3)
        if contribution_series:
            sm = plt.cm.ScalarMappable(norm=norm, cmap=cmap)
            sm.set_array([])
            colorbar = self.res_neighbor_corr_figure.colorbar(sm, ax=ax_corr, pad=0.02)
            colorbar.set_label("Pair Mean Frequency (GHz)")
            tick_vals = colorbar.get_ticks()
            tick_vals = [tick for tick in tick_vals if float(norm.vmin) <= float(tick) <= float(norm.vmax)]
            if tick_vals:
                colorbar.set_ticks(tick_vals)
            colorbar.ax.yaxis.set_major_formatter(FuncFormatter(lambda value, _pos: f"{value / 1.0e9:.3f}"))
            vmin = float(norm.vmin)
            vmax = float(norm.vmax)
            freq_span_hz = max(vmax - vmin, 1.0)
            sorted_pair_freqs_hz = np.sort(
                np.asarray(
                    [
                        float(freq_hz)
                        for freq_hz in pair_mean_freq_by_label.values()
                        if np.isfinite(float(freq_hz))
                    ],
                    dtype=float,
                )
            )
            if sorted_pair_freqs_hz.size >= 2:
                neighbor_steps_hz = np.diff(sorted_pair_freqs_hz)
                neighbor_steps_hz = neighbor_steps_hz[np.isfinite(neighbor_steps_hz) & (neighbor_steps_hz > 0.0)]
                band_halfwidth_hz = (
                    0.18 * float(np.min(neighbor_steps_hz))
                    if neighbor_steps_hz.size
                    else 0.012 * freq_span_hz
                )
            else:
                band_halfwidth_hz = 0.018 * freq_span_hz
            band_halfwidth_hz = min(max(band_halfwidth_hz, 0.004 * freq_span_hz), 0.045 * freq_span_hz)

            band_edges: list[tuple[float, float]] = []
            cursor_hz = vmin
            for pair_freq_hz in sorted_pair_freqs_hz:
                lo_hz = max(vmin, float(pair_freq_hz) - band_halfwidth_hz)
                hi_hz = min(vmax, float(pair_freq_hz) + band_halfwidth_hz)
                if lo_hz > cursor_hz:
                    band_edges.append((cursor_hz, lo_hz))
                cursor_hz = max(cursor_hz, hi_hz)
            if cursor_hz < vmax:
                band_edges.append((cursor_hz, vmax))

            for lo_hz, hi_hz in band_edges:
                if hi_hz <= lo_hz:
                    continue
                mid_hz = 0.5 * (lo_hz + hi_hz)
                base_color = np.asarray(cmap(norm(float(mid_hz)))[:3], dtype=float)
                dark_color = tuple(np.clip(0.55 * base_color, 0.0, 1.0))
                colorbar.ax.axhspan(
                    lo_hz,
                    hi_hz,
                    xmin=0.0,
                    xmax=1.0,
                    facecolor=dark_color,
                    edgecolor="none",
                    alpha=0.95,
                )

        ax_mag.set_xlabel("Interval End Time (days)")
        ax_mag.set_ylabel("Mean |df/f per day|")
        ax_mag.set_title("Neighbor Pair Drift Magnitude vs Time")
        ax_mag.grid(True, alpha=0.3)
        if np.any(mag_mask):
            ax_mag.legend(loc="best", fontsize=8)

        self.res_neighbor_corr_figure.tight_layout()
        if self.res_neighbor_corr_status_var is not None:
            origin_dt = data.get("elapsed_time_origin")
            origin_text = origin_dt.strftime("%Y-%m-%d") if isinstance(origin_dt, datetime) else "unknown"
            ref_interval = summary["reference_interval"]
            self.res_neighbor_corr_status_var.set(
                f"Elapsed-time origin: {origin_text}. Showing {len(points)} consecutive interval(s) across "
                f"{len(data['pair_series'])} neighboring pair(s). Top: "
                f"{'colored traces show per-pair contribution, ' if show_curves else ''}"
                f"grey band = middle 50%, black line = mean correlation with the reference interval ending at "
                f"{float(ref_interval['end_elapsed_days']):.3f} day(s). Bottom: mean absolute pair drift rate with +/- 1 std; "
                f"threshold {threshold_rel:.4f} df/f."
            )
        self.res_neighbor_corr_canvas.draw_idle()

    def open_resonator_neighbor_dfrel_window(self) -> None:
        if self.res_neighbor_dfrel_window is not None and self.res_neighbor_dfrel_window.winfo_exists():
            self.res_neighbor_dfrel_window.lift()
            self._render_resonator_neighbor_dfrel_window()
            return

        self.res_neighbor_dfrel_window = tk.Toplevel(self.root)
        self.res_neighbor_dfrel_window.title("Neighbor Pair df/f vs Time")
        self.res_neighbor_dfrel_window.geometry("1380x860")
        self.res_neighbor_dfrel_window.protocol("WM_DELETE_WINDOW", self._close_resonator_neighbor_dfrel_window)

        controls = tk.Frame(self.res_neighbor_dfrel_window, padx=8, pady=8)
        controls.pack(side="top", fill="x")
        control_row = tk.Frame(controls)
        control_row.pack(side="top", fill="x", anchor="w")
        self.res_neighbor_dfrel_status_var = tk.StringVar(
            value="Showing neighboring resonator-pair separation df/f versus elapsed time."
        )
        self.res_neighbor_dfrel_sep_rel_var = tk.DoubleVar(value=0.004)
        self.res_neighbor_dfrel_show_iqr_var = tk.BooleanVar(value=True)
        self.res_neighbor_dfrel_mode_var = tk.StringVar(value="drift")
        self.res_neighbor_dfrel_initial_date_var = tk.StringVar(value=self._dataset_res_neighbor_initial_date())

        tk.Label(control_row, text="Initial Date").pack(side="left", padx=(0, 4))
        initial_date_entry = tk.Entry(control_row, width=12, textvariable=self.res_neighbor_dfrel_initial_date_var)
        initial_date_entry.pack(side="left", padx=(0, 4))
        initial_date_entry.bind(
            "<Return>",
            lambda _event: (self._sync_res_neighbor_initial_date(autosave=True), self._render_resonator_neighbor_dfrel_window()),
        )
        initial_date_entry.bind(
            "<FocusOut>",
            lambda _event: (self._sync_res_neighbor_initial_date(autosave=True), self._render_resonator_neighbor_dfrel_window()),
        )
        tk.Label(control_row, text="YYYY-MM-DD").pack(side="left", padx=(0, 12))
        self.res_neighbor_dfrel_sep_scale = tk.Scale(
            control_row,
            from_=0.0,
            to=0.04,
            resolution=0.0001,
            orient="horizontal",
            length=260,
            digits=5,
            label="Max pair mean separation (df/f)",
            variable=self.res_neighbor_dfrel_sep_rel_var,
            command=lambda _value: self._render_resonator_neighbor_dfrel_window(),
        )
        self.res_neighbor_dfrel_sep_scale.pack(side="left", padx=(0, 12))
        tk.Radiobutton(
            control_row,
            text="Mean +/- Std Spacing",
            value="drift",
            variable=self.res_neighbor_dfrel_mode_var,
            command=self._render_resonator_neighbor_dfrel_window,
        ).pack(side="left", padx=(0, 8))
        tk.Radiobutton(
            control_row,
            text="Spacing Displacement",
            value="change",
            variable=self.res_neighbor_dfrel_mode_var,
            command=self._render_resonator_neighbor_dfrel_window,
        ).pack(side="left", padx=(0, 8))
        tk.Radiobutton(
            control_row,
            text="Spacing",
            value="spacing",
            variable=self.res_neighbor_dfrel_mode_var,
            command=self._render_resonator_neighbor_dfrel_window,
        ).pack(side="left", padx=(0, 8))
        tk.Checkbutton(
            control_row,
            text="Summary Band",
            variable=self.res_neighbor_dfrel_show_iqr_var,
            command=self._render_resonator_neighbor_dfrel_window,
        ).pack(side="left", padx=(0, 12))
        tk.Button(control_row, text="Show On Scans", width=13, command=self.open_resonator_neighbor_scan_window).pack(
            side="left",
            padx=(0, 8),
        )
        tk.Button(control_row, text="Refresh", width=10, command=self._render_resonator_neighbor_dfrel_window).pack(
            side="left",
            padx=(0, 8),
        )
        tk.Label(controls, textvariable=self.res_neighbor_dfrel_status_var, anchor="w", justify="left").pack(
            side="top",
            fill="x",
            expand=True,
            pady=(6, 0),
        )

        self.res_neighbor_dfrel_figure = Figure(figsize=(12.5, 7))
        self.res_neighbor_dfrel_canvas = FigureCanvasTkAgg(
            self.res_neighbor_dfrel_figure,
            master=self.res_neighbor_dfrel_window,
        )
        self.res_neighbor_dfrel_toolbar = NavigationToolbar2Tk(
            self.res_neighbor_dfrel_canvas,
            self.res_neighbor_dfrel_window,
        )
        self.res_neighbor_dfrel_toolbar.update()
        self.res_neighbor_dfrel_toolbar.pack(side="top", fill="x")
        self.res_neighbor_dfrel_canvas.get_tk_widget().pack(fill="both", expand=True)

        self._render_resonator_neighbor_dfrel_window()

    def _close_resonator_neighbor_dfrel_window(self) -> None:
        if self.res_neighbor_dfrel_window is not None and self.res_neighbor_dfrel_window.winfo_exists():
            self.res_neighbor_dfrel_window.destroy()
        self.res_neighbor_dfrel_window = None
        self.res_neighbor_dfrel_canvas = None
        self.res_neighbor_dfrel_toolbar = None
        self.res_neighbor_dfrel_figure = None
        self.res_neighbor_dfrel_status_var = None
        self.res_neighbor_dfrel_sep_rel_var = None
        self.res_neighbor_dfrel_show_iqr_var = None
        self.res_neighbor_dfrel_mode_var = None
        self.res_neighbor_dfrel_initial_date_var = None
        self.res_neighbor_dfrel_sep_scale = None
        self._res_neighbor_dfrel_ax = None

    def _render_resonator_neighbor_dfrel_window(self) -> None:
        if self.res_neighbor_dfrel_figure is None or self.res_neighbor_dfrel_canvas is None:
            return

        self.res_neighbor_dfrel_figure.clear()
        ax = self.res_neighbor_dfrel_figure.add_subplot(111)
        self._res_neighbor_dfrel_ax = ax

        threshold_rel = (
            float(self.res_neighbor_dfrel_sep_rel_var.get())
            if self.res_neighbor_dfrel_sep_rel_var is not None
            else 0.004
        )
        initial_date_text = (
            str(self.res_neighbor_dfrel_initial_date_var.get())
            if self.res_neighbor_dfrel_initial_date_var is not None
            else ""
        )
        try:
            overlay_state = self._resonator_neighbor_scan_overlay_state(
                threshold_rel,
                initial_date_text=initial_date_text,
            )
        except Exception as exc:
            ax.text(0.5, 0.5, str(exc), ha="center", va="center", transform=ax.transAxes)
            ax.set_axis_off()
            if self.res_neighbor_dfrel_status_var is not None:
                self.res_neighbor_dfrel_status_var.set(str(exc))
            self.res_neighbor_dfrel_canvas.draw_idle()
            return

        data = overlay_state["data"]
        mode = (
            str(self.res_neighbor_dfrel_mode_var.get())
            if self.res_neighbor_dfrel_mode_var is not None
            else "change"
        ).strip().lower()
        pair_series = self._resonator_neighbor_plot_series(data["pair_series"], mode)
        mean_pair_freqs_hz = np.asarray(data["mean_pair_freqs_hz"], dtype=float)
        norm = overlay_state["norm"]
        cmap = overlay_state["cmap"]
        vmin = float(norm.vmin)
        vmax = float(norm.vmax)
        show_iqr = bool(self.res_neighbor_dfrel_show_iqr_var.get()) if self.res_neighbor_dfrel_show_iqr_var is not None else True

        summary = []
        drift_single_interval_xlim: Optional[tuple[float, float]] = None
        if mode == "drift":
            summary = self._resonator_neighbor_drift_rate_summary(pair_series)
            drift_series = self._resonator_neighbor_drift_rate_series(pair_series)
            if summary:
                x_summary = np.asarray([float(item["elapsed_days"]) for item in summary], dtype=float)
                lower_summary = np.asarray([float(item["lower"]) for item in summary], dtype=float)
                upper_summary = np.asarray([float(item["upper"]) for item in summary], dtype=float)
                if show_iqr:
                    valid_mask = np.isfinite(x_summary) & np.isfinite(lower_summary) & np.isfinite(upper_summary)
                    valid_count = int(np.count_nonzero(valid_mask))
                    if valid_count >= 2:
                        ax.fill_between(
                            x_summary[valid_mask],
                            lower_summary[valid_mask],
                            upper_summary[valid_mask],
                            color="0.15",
                            alpha=0.28,
                            zorder=3.2,
                            linewidth=0.0,
                            label="Mean +/- 1 std",
                        )
                        ax.plot(
                            x_summary[valid_mask],
                            lower_summary[valid_mask],
                            color="black",
                            linewidth=0.9,
                            alpha=0.9,
                            zorder=3.35,
                        )
                        ax.plot(
                            x_summary[valid_mask],
                            upper_summary[valid_mask],
                            color="black",
                            linewidth=0.9,
                            alpha=0.9,
                            zorder=3.35,
                        )
                    elif valid_count == 1:
                        idx = int(np.flatnonzero(valid_mask)[0])
                        x0 = float(x_summary[idx])
                        mean0 = float(summary[idx]["mean"])
                        lower0 = float(lower_summary[idx])
                        upper0 = float(upper_summary[idx])
                        x_left = x0 - 0.5
                        x_right = x0 + 0.5
                        ax.fill_between(
                            np.asarray([x_left, x_right], dtype=float),
                            np.asarray([lower0, lower0], dtype=float),
                            np.asarray([upper0, upper0], dtype=float),
                            color="0.15",
                            alpha=0.28,
                            zorder=3.2,
                            linewidth=0.0,
                            label="Mean +/- 1 std",
                        )
                        ax.plot(
                            np.asarray([x_left, x_right, x_right, x_left, x_left], dtype=float),
                            np.asarray([lower0, lower0, upper0, upper0, lower0], dtype=float),
                            color="black",
                            linewidth=1.0,
                            alpha=0.95,
                            zorder=3.35,
                        )
                        ax.plot(
                            np.asarray([x_left, x_right], dtype=float),
                            np.asarray([mean0, mean0], dtype=float),
                            color="black",
                            linewidth=1.6,
                            alpha=0.95,
                            zorder=4.3,
                        )
                        drift_single_interval_xlim = (x0 - 5.0, x0 + 5.0)
            for pair in drift_series:
                color = pair["color"]
                x = np.asarray([float(pt["elapsed_days"]) for pt in pair["drift_points"]], dtype=float)
                y = np.asarray([float(pt["drift_rate"]) for pt in pair["drift_points"]], dtype=float)
                ax.plot(
                    x,
                    y,
                    color=color,
                    linewidth=1.5,
                    alpha=0.9,
                    marker="o",
                    markersize=4.0,
                    zorder=4.0,
                )
        else:
            summary = self._resonator_neighbor_summary_by_time(pair_series)
            if show_iqr and summary:
                x_summary = np.asarray([float(item["elapsed_days"]) for item in summary], dtype=float)
                lower = np.asarray([float(item["lower"]) for item in summary], dtype=float)
                upper = np.asarray([float(item["upper"]) for item in summary], dtype=float)
                q1 = np.asarray([float(item["q1"]) for item in summary], dtype=float)
                median = np.asarray([float(item["median"]) for item in summary], dtype=float)
                q3 = np.asarray([float(item["q3"]) for item in summary], dtype=float)
                ax.fill_between(
                    x_summary,
                    lower,
                    upper,
                    color="0.85",
                    alpha=0.6,
                    zorder=1.4,
                    linewidth=0.0,
                    label="Mean +/- 1 std",
                )
                ax.fill_between(
                    x_summary,
                    q1,
                    q3,
                    color="0.15",
                    alpha=0.42,
                    zorder=3.2,
                    linewidth=0.0,
                    label="Middle 50%",
                )
                ax.plot(
                    x_summary,
                    q1,
                    color="black",
                    linewidth=0.9,
                    alpha=0.95,
                    zorder=3.35,
                )
                ax.plot(
                    x_summary,
                    q3,
                    color="black",
                    linewidth=0.9,
                    alpha=0.95,
                    zorder=3.35,
                )
                ax.plot(
                    x_summary,
                    median,
                    color="black",
                    linewidth=6.0,
                    alpha=1.0,
                    zorder=4.2,
                    label="Median",
                )

            for pair in pair_series:
                color = pair["color"]
                x = np.asarray([float(pt["elapsed_days"]) for pt in pair["points"]], dtype=float)
                y = np.asarray([float(pt.get("plot_df_over_f", pt["df_over_f"])) for pt in pair["points"]], dtype=float)
                ax.plot(
                    x,
                    y,
                    color=color,
                    linewidth=1.5,
                    alpha=0.9,
                    marker="o",
                    markersize=4.0,
                )

        ax.axhline(0.0, color="0.5", linewidth=0.8, linestyle="--")
        ax.grid(True, alpha=0.3)
        ax.set_xlabel("Elapsed Time (days)")
        if mode == "drift":
            ax.set_ylabel("Neighbor Pair Gap Drift Rate (df/f per day)")
            ax.set_title("Neighbor Pair Gap Drift Rate vs Time")
        elif mode == "change":
            ax.set_ylabel("Neighbor Pair Separation Displacement df/f")
            ax.set_title("Neighboring Resonator-Pair Separation Displacement vs Time")
        else:
            ax.set_ylabel("Neighbor Pair Separation df/f")
            ax.set_title("Neighboring Resonator-Pair Relative Separation vs Time")
        if mode == "drift" and drift_single_interval_xlim is not None:
            ax.set_xlim(*drift_single_interval_xlim)

        if show_iqr and summary:
            ax.legend(loc="best", fontsize=8)

        sm = plt.cm.ScalarMappable(norm=norm, cmap=cmap)
        sm.set_array([])
        colorbar = self.res_neighbor_dfrel_figure.colorbar(sm, ax=ax, pad=0.02)
        colorbar.set_label("Pair Mean Frequency (GHz)")
        tick_vals = colorbar.get_ticks()
        tick_vals = [tick for tick in tick_vals if vmin <= float(tick) <= vmax]
        if tick_vals:
            colorbar.set_ticks(tick_vals)
        colorbar.ax.yaxis.set_major_formatter(FuncFormatter(lambda value, _pos: f"{value / 1.0e9:.3f}"))

        freq_span_hz = max(vmax - vmin, 1.0)
        sorted_pair_freqs_hz = np.sort(mean_pair_freqs_hz)
        if sorted_pair_freqs_hz.size >= 2:
            neighbor_steps_hz = np.diff(sorted_pair_freqs_hz)
            neighbor_steps_hz = neighbor_steps_hz[np.isfinite(neighbor_steps_hz) & (neighbor_steps_hz > 0.0)]
            band_halfwidth_hz = (
                0.18 * float(np.min(neighbor_steps_hz))
                if neighbor_steps_hz.size
                else 0.012 * freq_span_hz
            )
        else:
            band_halfwidth_hz = 0.018 * freq_span_hz
        band_halfwidth_hz = min(max(band_halfwidth_hz, 0.004 * freq_span_hz), 0.045 * freq_span_hz)

        band_edges: list[tuple[float, float]] = []
        cursor_hz = vmin
        for pair_freq_hz in sorted_pair_freqs_hz:
            lo_hz = max(vmin, float(pair_freq_hz) - band_halfwidth_hz)
            hi_hz = min(vmax, float(pair_freq_hz) + band_halfwidth_hz)
            if lo_hz > cursor_hz:
                band_edges.append((cursor_hz, lo_hz))
            cursor_hz = max(cursor_hz, hi_hz)
        if cursor_hz < vmax:
            band_edges.append((cursor_hz, vmax))

        for lo_hz, hi_hz in band_edges:
            if hi_hz <= lo_hz:
                continue
            mid_hz = 0.5 * (lo_hz + hi_hz)
            base_color = np.asarray(cmap(norm(float(mid_hz)))[:3], dtype=float)
            dark_color = tuple(np.clip(0.55 * base_color, 0.0, 1.0))
            colorbar.ax.axhspan(lo_hz, hi_hz, xmin=0.0, xmax=1.0, facecolor=dark_color, edgecolor="none", alpha=0.95)

        self.res_neighbor_dfrel_figure.tight_layout()
        if self.res_neighbor_dfrel_status_var is not None:
            origin_dt = data.get("elapsed_time_origin")
            origin_text = (
                origin_dt.strftime("%Y-%m-%d")
                if isinstance(origin_dt, datetime)
                else "unknown"
            )
            origin_prefix = f"Elapsed-time origin: {origin_text}. "
            if mode == "drift":
                summary_text = " Summary overlay: grey band = mean +/- 1 std; colored traces = individual pair drift rates." if show_iqr else ""
                self.res_neighbor_dfrel_status_var.set(
                    f"{origin_prefix}Showing mean/std of pair-gap drift rate across {len(summary)} elapsed-time point(s) from {len(pair_series)} neighboring pair(s) and {len(data['tests'])} selected test unit(s); threshold {threshold_rel:.4f} df/f.{summary_text}"
                )
            else:
                summary_text = " Summary overlay: grey band = middle 50%, black line = median." if show_iqr else ""
                mode_text = "spacing displacement from initial" if mode == "change" else "spacing"
                self.res_neighbor_dfrel_status_var.set(
                    f"{origin_prefix}Showing {len(pair_series)} neighboring pair curve(s) across {len(data['tests'])} selected test unit(s) in {mode_text} mode; threshold {threshold_rel:.4f} df/f.{summary_text}"
                )
        self.res_neighbor_dfrel_canvas.draw_idle()
        if self.res_neighbor_scan_window is not None and self.res_neighbor_scan_window.winfo_exists():
            self._render_resonator_neighbor_scan_window()

    def _draw_resonator_neighbor_scan_overlay(
        self,
        ax,
        *,
        rows: list[dict],
        overlay_state: dict,
        xlim_ghz: Optional[tuple[float, float]] = None,
        spacing: float = 1.5,
        truncate_threshold: float = 1.5,
        title: str = "",
    ) -> dict:
        data = overlay_state["data"]
        pair_series = data["pair_series"]
        resonator_colors = overlay_state["resonator_colors"]
        neutral_color = (0.45, 0.45, 0.45, 1.0)
        offset_by_scan_key, tick_info = self._attached_resonance_editor_offset_map(rows, spacing)

        freq_min = min(float(row["freq"][0]) for row in rows) / 1.0e9
        freq_max = max(float(row["freq"][-1]) for row in rows) / 1.0e9
        x_pad = max((freq_max - freq_min) * 0.01, 1e-6)
        resonator_tracks: dict[str, list[tuple[float, float]]] = {}
        points_by_scan_key: dict[str, dict[str, dict]] = {}
        total_markers = 0

        for row in rows:
            scan_key = str(row["scan_key"])
            offset = float(offset_by_scan_key[scan_key])
            freq_ghz = np.asarray(row["freq"], dtype=float) / 1.0e9
            amp_display = np.minimum(np.asarray(row["amp"], dtype=float), truncate_threshold)
            y = amp_display + offset
            ax.plot(freq_ghz, y, linewidth=1.0, color="tab:blue", alpha=0.8, zorder=1)

            row_points: dict[str, dict] = {}
            for resonator in row["resonators"]:
                resonator_label = str(resonator["resonator_number"])
                target_hz = float(resonator["target_hz"])
                target_ghz = target_hz / 1.0e9
                if xlim_ghz is not None and not (xlim_ghz[0] <= target_ghz <= xlim_ghz[1]):
                    continue
                y_pt = self._interpolate_y(row["freq"], amp_display, target_hz) + offset
                color = resonator_colors.get(resonator_label, neutral_color)
                row_points[resonator_label] = {"x_ghz": target_ghz, "y": y_pt, "color": color}
                resonator_tracks.setdefault(resonator_label, []).append((target_ghz, y_pt))
                ax.plot(
                    [target_ghz],
                    [y_pt],
                    linestyle="none",
                    marker="o",
                    markersize=6,
                    markerfacecolor="none",
                    markeredgecolor=color,
                    markeredgewidth=1.5,
                    zorder=4,
                )
                ax.text(
                    target_ghz,
                    y_pt - 0.18,
                    resonator_label,
                    ha="center",
                    va="top",
                    fontsize=8,
                    color=color,
                    zorder=5,
                )
                total_markers += 1
            points_by_scan_key[scan_key] = row_points

        for resonator_label, points in resonator_tracks.items():
            if len(points) < 2:
                continue
            points = sorted(points, key=lambda item: item[1], reverse=True)
            ax.plot(
                [pt[0] for pt in points],
                [pt[1] for pt in points],
                linestyle=":",
                linewidth=2.0,
                color=resonator_colors.get(resonator_label, neutral_color),
                alpha=0.95,
                zorder=2,
            )

        connector_count = 0
        visible_pair_labels: set[str] = set()
        for row in rows:
            row_points = points_by_scan_key.get(str(row["scan_key"]), {})
            for pair in pair_series:
                low_label = str(pair["low_label"])
                high_label = str(pair["high_label"])
                if low_label not in row_points or high_label not in row_points:
                    continue
                low_pt = row_points[low_label]
                high_pt = row_points[high_label]
                ax.plot(
                    [float(low_pt["x_ghz"]), float(high_pt["x_ghz"])],
                    [float(low_pt["y"]), float(high_pt["y"])],
                    linestyle=":",
                    linewidth=2.4,
                    color=pair["color"],
                    alpha=0.9,
                    zorder=3,
                )
                connector_count += 1
                visible_pair_labels.add(str(pair["label"]))

        y_low = min(
            float(np.min(np.minimum(np.asarray(row["amp"], dtype=float), truncate_threshold)))
            + float(offset_by_scan_key[str(row["scan_key"])])
            for row in rows
        )
        y_high = max(
            float(np.max(np.minimum(np.asarray(row["amp"], dtype=float), truncate_threshold)))
            + float(offset_by_scan_key[str(row["scan_key"])])
            for row in rows
        )
        ax.set_xlabel("Frequency (GHz)")
        ax.set_ylabel("Normalized |S21| + vertical offset")
        ax.grid(True, alpha=0.3)
        ax.set_yticks([item[0] for item in tick_info])
        ax.set_yticklabels([item[1] for item in tick_info], fontsize=8)
        ax.set_ylim(y_low - 0.2, y_high + 0.2)
        if xlim_ghz is None:
            ax.set_xlim(freq_min - 0.5 * x_pad, freq_max + 2.0 * x_pad)
        else:
            ax.set_xlim(xlim_ghz)
        if title:
            ax.set_title(title)
        return {
            "marker_count": total_markers,
            "connector_count": connector_count,
            "visible_pair_labels": visible_pair_labels,
        }

    def open_resonator_neighbor_scan_window(self, source: str = "dfrel") -> None:
        self._res_neighbor_scan_source = str(source or "dfrel").strip().lower()
        if self.res_neighbor_scan_window is not None and self.res_neighbor_scan_window.winfo_exists():
            self.res_neighbor_scan_window.lift()
            self._render_resonator_neighbor_scan_window()
            return

        self.res_neighbor_scan_window = tk.Toplevel(self.root)
        self.res_neighbor_scan_window.title("Neighbor Pair Resonators On Scans")
        self.res_neighbor_scan_window.geometry("1380x900")
        self.res_neighbor_scan_window.protocol("WM_DELETE_WINDOW", self._close_resonator_neighbor_scan_window)

        controls = tk.Frame(self.res_neighbor_scan_window, padx=8, pady=8)
        controls.pack(side="top", fill="x")
        self.res_neighbor_scan_status_var = tk.StringVar(
            value="Showing marked resonators on selected VNA scans with active neighboring-pair coloring."
        )
        tk.Label(controls, textvariable=self.res_neighbor_scan_status_var, anchor="w", justify="left").pack(
            side="left",
            fill="x",
            expand=True,
        )
        tk.Button(controls, text="Save PDF", width=10, command=self._save_resonator_neighbor_scan_pdf).pack(
            side="right",
            padx=(0, 8),
        )
        tk.Button(controls, text="Refresh", width=10, command=self._render_resonator_neighbor_scan_window).pack(
            side="right"
        )

        self.res_neighbor_scan_figure = Figure(figsize=(12.5, 7.5))
        self.res_neighbor_scan_canvas = FigureCanvasTkAgg(
            self.res_neighbor_scan_figure,
            master=self.res_neighbor_scan_window,
        )
        self.res_neighbor_scan_toolbar = NavigationToolbar2Tk(
            self.res_neighbor_scan_canvas,
            self.res_neighbor_scan_window,
        )
        self.res_neighbor_scan_toolbar.update()
        self.res_neighbor_scan_toolbar.pack(side="top", fill="x")
        self.res_neighbor_scan_canvas.get_tk_widget().pack(fill="both", expand=True)

        self._render_resonator_neighbor_scan_window()

    def _close_resonator_neighbor_scan_window(self) -> None:
        if self.res_neighbor_scan_window is not None and self.res_neighbor_scan_window.winfo_exists():
            self.res_neighbor_scan_window.destroy()
        self.res_neighbor_scan_window = None
        self.res_neighbor_scan_canvas = None
        self.res_neighbor_scan_toolbar = None
        self.res_neighbor_scan_figure = None
        self.res_neighbor_scan_status_var = None
        self._res_neighbor_scan_source = "dfrel"
        self._res_neighbor_scan_ax = None

    def _render_resonator_neighbor_scan_window(self) -> None:
        if self.res_neighbor_scan_figure is None or self.res_neighbor_scan_canvas is None:
            return

        self.res_neighbor_scan_figure.clear()
        ax = self.res_neighbor_scan_figure.add_subplot(111)
        self._res_neighbor_scan_ax = ax

        threshold_rel, initial_date_text = self._resonator_neighbor_scan_control_values()
        rows, warnings = self._selected_scans_for_attached_resonance_editor()
        if not rows:
            message = "No selected scans with normalized data."
            if warnings:
                message += " Missing normalized data for: " + ", ".join(warnings[:6])
            ax.text(0.5, 0.5, message, ha="center", va="center", transform=ax.transAxes)
            ax.set_axis_off()
            if self.res_neighbor_scan_status_var is not None:
                self.res_neighbor_scan_status_var.set(message)
            self.res_neighbor_scan_canvas.draw_idle()
            return

        try:
            overlay_state = self._resonator_neighbor_scan_overlay_state(
                threshold_rel,
                initial_date_text=initial_date_text,
            )
        except Exception as exc:
            ax.text(0.5, 0.5, str(exc), ha="center", va="center", transform=ax.transAxes)
            ax.set_axis_off()
            if self.res_neighbor_scan_status_var is not None:
                self.res_neighbor_scan_status_var.set(str(exc))
            self.res_neighbor_scan_canvas.draw_idle()
            return

        data = overlay_state["data"]
        draw_stats = self._draw_resonator_neighbor_scan_overlay(
            ax,
            rows=rows,
            overlay_state=overlay_state,
            title="Marked Resonators On Selected Scans With Neighbor-Pair Coloring",
        )

        self.res_neighbor_scan_figure.tight_layout()
        if self.res_neighbor_scan_status_var is not None:
            status = (
                f"Showing {int(draw_stats['marker_count'])} marker(s), {len(data['pair_series'])} active neighboring pair(s), and "
                f"{int(draw_stats['connector_count'])} same-scan pair connector(s); threshold {threshold_rel:.4f} df/f."
            )
            if warnings:
                status += " Missing normalized data for: " + ", ".join(warnings[:6])
                if len(warnings) > 6:
                    status += f", ... (+{len(warnings) - 6} more)"
            self.res_neighbor_scan_status_var.set(status)
        self.res_neighbor_scan_canvas.draw_idle()

    def _save_resonator_neighbor_scan_pdf(self) -> None:
        threshold_rel, initial_date_text = self._resonator_neighbor_scan_control_values()
        rows, warnings = self._selected_scans_for_attached_resonance_editor()
        if not rows:
            messagebox.showwarning("No plottable scans", "No selected scans with normalized data are available.")
            return

        try:
            overlay_state = self._resonator_neighbor_scan_overlay_state(
                threshold_rel,
                initial_date_text=initial_date_text,
            )
        except Exception as exc:
            messagebox.showerror("Cannot save PDF", str(exc), parent=self.res_neighbor_scan_window)
            return

        pair_series = overlay_state["data"]["pair_series"]
        pair_mid_freqs_hz = sorted(float(pair["mean_freq_hz"]) for pair in pair_series)
        if not pair_mid_freqs_hz:
            messagebox.showwarning("No active pairs", "No active neighboring pairs are available for PDF export.")
            return

        out_dir = _dataset_dir(self.dataset) / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_neighbor_pair_scan_plots"
        out_dir.mkdir(parents=True, exist_ok=True)
        pdf_path = out_dir / "neighbor_pair_scan_overlay.pdf"

        freq_min_hz = min(float(row["freq"][0]) for row in rows)
        freq_max_hz = max(float(row["freq"][-1]) for row in rows)
        zoom_span_hz = 100.0e6
        zoom_step_hz = 80.0e6
        zoom_windows: list[tuple[float, float]] = []
        if freq_max_hz - freq_min_hz > zoom_span_hz:
            start_hz = freq_min_hz
            seen_starts: set[float] = set()
            while True:
                rounded_start = round(start_hz, 3)
                if rounded_start in seen_starts:
                    break
                seen_starts.add(rounded_start)
                end_hz = min(start_hz + zoom_span_hz, freq_max_hz)
                if any(start_hz <= pair_hz <= end_hz for pair_hz in pair_mid_freqs_hz):
                    zoom_windows.append((start_hz, end_hz))
                if end_hz >= freq_max_hz:
                    break
                next_start = start_hz + zoom_step_hz
                if next_start + zoom_span_hz > freq_max_hz:
                    next_start = max(freq_min_hz, freq_max_hz - zoom_span_hz)
                if next_start <= start_hz:
                    break
                start_hz = next_start

        page_count = 0
        marker_count = 0
        connector_count = 0
        with PdfPages(pdf_path) as pdf:
            fig = Figure(figsize=(12, 7))
            ax = fig.add_subplot(111)
            stats = self._draw_resonator_neighbor_scan_overlay(
                ax,
                rows=rows,
                overlay_state=overlay_state,
                title="Neighbor Pair Scan Overlay | Full Span",
            )
            fig.tight_layout()
            pdf.savefig(fig)
            plt.close(fig)
            page_count += 1
            marker_count += int(stats["marker_count"])
            connector_count += int(stats["connector_count"])

            for page_idx, (lo_hz, hi_hz) in enumerate(zoom_windows, start=1):
                fig = Figure(figsize=(12, 7))
                ax = fig.add_subplot(111)
                stats = self._draw_resonator_neighbor_scan_overlay(
                    ax,
                    rows=rows,
                    overlay_state=overlay_state,
                    xlim_ghz=(lo_hz / 1.0e9, hi_hz / 1.0e9),
                    title=(
                        f"Neighbor Pair Scan Overlay | Zoom {page_idx} | "
                        f"{lo_hz / 1.0e9:.6f} to {hi_hz / 1.0e9:.6f} GHz"
                    ),
                )
                if not stats["visible_pair_labels"]:
                    plt.close(fig)
                    continue
                fig.tight_layout()
                pdf.savefig(fig)
                plt.close(fig)
                page_count += 1
                marker_count += int(stats["marker_count"])
                connector_count += int(stats["connector_count"])

        self.dataset.processing_history.append(
            _make_event(
                "save_neighbor_pair_scan_overlay_pdf",
                {
                    "output_pdf": str(pdf_path),
                    "page_count": int(page_count),
                    "marker_count": int(marker_count),
                    "connector_count": int(connector_count),
                    "threshold_rel": float(threshold_rel),
                    "zoom_span_mhz": 100.0,
                    "zoom_overlap_mhz": 20.0,
                },
            )
        )
        self._mark_dirty()
        self._autosave_dataset()
        for warning in warnings[:20]:
            self._log(f"Neighbor pair scan overlay warning: {warning}")
        self._log(f"Saved neighbor pair scan overlay PDF: {pdf_path}")
        if self.res_neighbor_scan_status_var is not None:
            self.res_neighbor_scan_status_var.set(
                f"Saved overlay PDF with {page_count} page(s) to {pdf_path}"
            )

    def _resonator_shift_correlation_data(self) -> dict:
        tests = self._resonator_shift_test_units()
        if len(tests) < 3:
            raise ValueError("At least three selected test dates with marked resonators are required for correlation analysis.")

        values_by_resonator: dict[str, list[float]] = {}
        for test in tests:
            for resonator_label, freq_hz in test["resonators"].items():
                values_by_resonator.setdefault(resonator_label, []).append(float(freq_hz))

        mean_freq_by_resonator: dict[str, float] = {}
        for resonator_label, values in values_by_resonator.items():
            arr = np.asarray(values, dtype=float)
            arr = arr[np.isfinite(arr)]
            if arr.size:
                mean_freq_by_resonator[resonator_label] = float(np.mean(arr))

        pair_count = len(tests) - 1
        resonator_labels = sorted(mean_freq_by_resonator, key=lambda label: (mean_freq_by_resonator[label], self._resonator_sort_key(label)))
        if not resonator_labels:
            raise ValueError("No marked resonators were found for the selected tests.")

        label_to_index = {label: idx for idx, label in enumerate(resonator_labels)}
        shift_matrix = np.full((len(resonator_labels), pair_count), np.nan, dtype=float)
        pair_labels: list[str] = []
        for pair_idx in range(pair_count):
            left = tests[pair_idx]
            right = tests[pair_idx + 1]
            pair_labels.append(f"{left['date_label']} -> {right['date_label']}")
            shared = set(left["resonators"]).intersection(right["resonators"])
            for resonator_label in shared:
                row_idx = label_to_index.get(resonator_label)
                mean_freq_hz = mean_freq_by_resonator.get(resonator_label)
                if row_idx is None or mean_freq_hz is None or mean_freq_hz == 0.0:
                    continue
                delta_rel = (float(right["resonators"][resonator_label]) - float(left["resonators"][resonator_label])) / float(mean_freq_hz)
                shift_matrix[row_idx, pair_idx] = float(delta_rel)

        valid_counts = np.sum(np.isfinite(shift_matrix), axis=1)
        keep_mask = valid_counts >= 2
        if np.count_nonzero(keep_mask) < 2:
            raise ValueError("Need at least two resonators with shifts on two or more consecutive test pairs.")
        resonator_labels = [label for label, keep in zip(resonator_labels, keep_mask) if keep]
        mean_freqs_hz = np.asarray([mean_freq_by_resonator[label] for label in resonator_labels], dtype=float)
        shift_matrix = shift_matrix[keep_mask, :]

        corr_matrix = np.full((len(resonator_labels), len(resonator_labels)), np.nan, dtype=float)
        pair_points: list[dict] = []
        for i in range(len(resonator_labels)):
            corr_matrix[i, i] = 1.0
            for j in range(i + 1, len(resonator_labels)):
                xi = shift_matrix[i, :]
                xj = shift_matrix[j, :]
                mask = np.isfinite(xi) & np.isfinite(xj)
                if np.count_nonzero(mask) < 2:
                    continue
                vi = xi[mask]
                vj = xj[mask]
                std_i = float(np.std(vi))
                std_j = float(np.std(vj))
                if std_i == 0.0 and std_j == 0.0:
                    corr = 1.0 if np.allclose(vi, vj) else np.nan
                elif std_i == 0.0 or std_j == 0.0:
                    corr = np.nan
                else:
                    corr = float(np.corrcoef(vi, vj)[0, 1])
                corr_matrix[i, j] = corr
                corr_matrix[j, i] = corr
                if np.isfinite(corr):
                    pair_points.append(
                        {
                            "separation_hz": float(abs(mean_freqs_hz[j] - mean_freqs_hz[i])),
                            "correlation": corr,
                        }
                    )

        if not pair_points:
            raise ValueError("No resonator pairs had enough overlapping shift measurements to correlate.")

        separations_hz = np.asarray([pt["separation_hz"] for pt in pair_points], dtype=float)
        max_sep = float(np.max(separations_hz))
        if max_sep <= 0.0:
            bin_edges = np.asarray([0.0, 1.0], dtype=float)
        else:
            nbins = min(16, max(4, int(np.sqrt(len(pair_points)))))
            bin_edges = np.linspace(0.0, max_sep, nbins + 1)
            if not np.all(np.diff(bin_edges) > 0):
                bin_edges = np.asarray([0.0, max_sep], dtype=float)
        bin_centers: list[float] = []
        bin_means: list[float] = []
        bin_counts: list[int] = []
        for lo, hi in zip(bin_edges[:-1], bin_edges[1:]):
            if hi <= lo:
                continue
            if hi == bin_edges[-1]:
                mask = (separations_hz >= lo) & (separations_hz <= hi)
            else:
                mask = (separations_hz >= lo) & (separations_hz < hi)
            if not np.any(mask):
                continue
            corr_vals = np.asarray([pair_points[idx]["correlation"] for idx in np.flatnonzero(mask)], dtype=float)
            corr_vals = corr_vals[np.isfinite(corr_vals)]
            if corr_vals.size == 0:
                continue
            bin_centers.append(0.5 * (lo + hi))
            bin_means.append(float(np.mean(corr_vals)))
            bin_counts.append(int(corr_vals.size))

        return {
            "tests": tests,
            "pair_labels": pair_labels,
            "resonator_labels": resonator_labels,
            "mean_freqs_hz": mean_freqs_hz,
            "shift_matrix": shift_matrix,
            "corr_matrix": corr_matrix,
            "pair_points": pair_points,
            "bin_centers_hz": np.asarray(bin_centers, dtype=float),
            "bin_means": np.asarray(bin_means, dtype=float),
            "bin_counts": np.asarray(bin_counts, dtype=int),
        }

    def open_resonator_shift_correlation_window(self) -> None:
        if self.res_shift_corr_window is not None and self.res_shift_corr_window.winfo_exists():
            self.res_shift_corr_window.lift()
            self._render_resonator_shift_correlation_window()
            return

        self.res_shift_corr_window = tk.Toplevel(self.root)
        self.res_shift_corr_window.title("Resonator Shift Correlation")
        self.res_shift_corr_window.geometry("1460x900")
        self.res_shift_corr_window.protocol("WM_DELETE_WINDOW", self._close_resonator_shift_correlation_window)

        controls = tk.Frame(self.res_shift_corr_window, padx=8, pady=8)
        controls.pack(side="top", fill="x")
        self.res_shift_corr_status_var = tk.StringVar(
            value="Showing correlation of resonator df/f across consecutive selected tests."
        )
        tk.Label(controls, textvariable=self.res_shift_corr_status_var, anchor="w").pack(
            side="left", fill="x", expand=True
        )
        tk.Button(controls, text="Refresh", width=10, command=self._render_resonator_shift_correlation_window).pack(
            side="right"
        )

        self.res_shift_corr_figure = Figure(figsize=(13, 8))
        self.res_shift_corr_canvas = FigureCanvasTkAgg(self.res_shift_corr_figure, master=self.res_shift_corr_window)
        self.res_shift_corr_toolbar = NavigationToolbar2Tk(self.res_shift_corr_canvas, self.res_shift_corr_window)
        self.res_shift_corr_toolbar.update()
        self.res_shift_corr_toolbar.pack(side="top", fill="x")
        self.res_shift_corr_canvas.get_tk_widget().pack(fill="both", expand=True)

        self._render_resonator_shift_correlation_window()

    def _close_resonator_shift_correlation_window(self) -> None:
        if self.res_shift_corr_window is not None and self.res_shift_corr_window.winfo_exists():
            self.res_shift_corr_window.destroy()
        self.res_shift_corr_window = None
        self.res_shift_corr_canvas = None
        self.res_shift_corr_toolbar = None
        self.res_shift_corr_figure = None
        self.res_shift_corr_status_var = None
        self._res_shift_corr_axes = None

    def _render_resonator_shift_correlation_window(self) -> None:
        if self.res_shift_corr_figure is None or self.res_shift_corr_canvas is None:
            return

        prior_scatter_xlim = None
        prior_scatter_ylim = None
        if self._res_shift_corr_axes is not None:
            try:
                prior_scatter_xlim = self._res_shift_corr_axes[1].get_xlim()
                prior_scatter_ylim = self._res_shift_corr_axes[1].get_ylim()
            except Exception:
                prior_scatter_xlim = None
                prior_scatter_ylim = None

        self.res_shift_corr_figure.clear()
        ax_heat = self.res_shift_corr_figure.add_subplot(1, 2, 1)
        ax_scatter = self.res_shift_corr_figure.add_subplot(1, 2, 2)
        self._res_shift_corr_axes = (ax_heat, ax_scatter)

        try:
            data = self._resonator_shift_correlation_data()
        except Exception as exc:
            ax_heat.text(0.5, 0.5, str(exc), ha="center", va="center", transform=ax_heat.transAxes)
            ax_heat.set_axis_off()
            ax_scatter.set_axis_off()
            if self.res_shift_corr_status_var is not None:
                self.res_shift_corr_status_var.set(str(exc))
            self.res_shift_corr_canvas.draw_idle()
            return

        mean_freqs_ghz = np.asarray(data["mean_freqs_hz"], dtype=float) / 1.0e9
        corr_matrix = np.asarray(data["corr_matrix"], dtype=float)
        pair_points = data["pair_points"]
        bin_centers_mhz = np.asarray(data["bin_centers_hz"], dtype=float) / 1.0e6
        bin_means = np.asarray(data["bin_means"], dtype=float)

        freq_min = float(np.min(mean_freqs_ghz))
        freq_max = float(np.max(mean_freqs_ghz))
        im = ax_heat.imshow(
            corr_matrix,
            origin="lower",
            aspect="auto",
            extent=[freq_min, freq_max, freq_min, freq_max],
            vmin=-1.0,
            vmax=1.0,
            cmap="coolwarm",
        )
        ax_heat.set_xlabel("Mean Resonator Frequency (GHz)")
        ax_heat.set_ylabel("Mean Resonator Frequency (GHz)")
        ax_heat.set_title("Shift Correlation Heatmap")
        colorbar = self.res_shift_corr_figure.colorbar(im, ax=ax_heat, fraction=0.046, pad=0.04)
        colorbar.set_label("Correlation")

        sep_mhz = np.asarray([float(pt["separation_hz"]) / 1.0e6 for pt in pair_points], dtype=float)
        corr_vals = np.asarray([float(pt["correlation"]) for pt in pair_points], dtype=float)
        ax_scatter.scatter(
            sep_mhz,
            corr_vals,
            s=18,
            color="tab:blue",
            alpha=0.45,
            edgecolors="none",
            label="Resonator pairs",
        )
        if bin_centers_mhz.size and bin_means.size:
            ax_scatter.plot(
                bin_centers_mhz,
                bin_means,
                color="black",
                linewidth=2.0,
                marker="o",
                markersize=4,
                label="Binned mean",
            )
        ax_scatter.axhline(0.0, color="0.4", linewidth=0.8, linestyle="--")
        ax_scatter.grid(True, alpha=0.3)
        ax_scatter.set_xlabel("Resonator Frequency Separation (MHz)")
        ax_scatter.set_ylabel("Shift Correlation")
        ax_scatter.set_title("Correlation vs Frequency Separation")
        ax_scatter.legend(loc="best", fontsize=8)

        if prior_scatter_xlim is not None and prior_scatter_ylim is not None:
            ax_scatter.set_xlim(prior_scatter_xlim)
            ax_scatter.set_ylim(prior_scatter_ylim)

        self.res_shift_corr_figure.suptitle("Resonator Shift Correlations Across Tests", fontsize=12)
        self.res_shift_corr_figure.tight_layout()
        if self.res_shift_corr_status_var is not None:
            self.res_shift_corr_status_var.set(
                f"Showing {len(data['resonator_labels'])} resonators and {len(pair_points)} resonator-pair correlations across {len(data['pair_labels'])} consecutive test pair(s)."
            )
        self.res_shift_corr_canvas.draw_idle()

    def _resonator_pair_dfdiff_hist_data(self, max_separation_mhz: float) -> dict:
        tests = self._resonator_shift_test_units()
        if len(tests) < 2:
            raise ValueError("At least two selected test dates with marked resonators are required.")

        max_separation_hz = max(0.0, float(max_separation_mhz)) * 1.0e6

        values_by_resonator: dict[str, list[float]] = {}
        for test in tests:
            for resonator_label, freq_hz in test["resonators"].items():
                values_by_resonator.setdefault(resonator_label, []).append(float(freq_hz))

        mean_freq_by_resonator: dict[str, float] = {}
        for resonator_label, values in values_by_resonator.items():
            arr = np.asarray(values, dtype=float)
            arr = arr[np.isfinite(arr)]
            if arr.size:
                mean_freq_by_resonator[resonator_label] = float(np.mean(arr))
        if not mean_freq_by_resonator:
            raise ValueError("No marked resonators were found for the selected tests.")

        pair_labels: list[str] = []
        delta_diff_records: list[dict] = []
        for pair_idx in range(len(tests) - 1):
            left = tests[pair_idx]
            right = tests[pair_idx + 1]
            pair_label = f"{left['date_label']} -> {right['date_label']}"
            pair_labels.append(pair_label)

            shared_labels = [
                label
                for label in left["resonators"]
                if label in right["resonators"] and label in mean_freq_by_resonator
            ]
            shared_labels.sort(
                key=lambda label: (float(mean_freq_by_resonator[label]), self._resonator_sort_key(label))
            )
            if len(shared_labels) < 2:
                continue

            deltas_hz: dict[str, float] = {}
            for label in shared_labels:
                delta_hz = float(right["resonators"][label]) - float(left["resonators"][label])
                if np.isfinite(delta_hz):
                    deltas_hz[label] = delta_hz

            ordered = [label for label in shared_labels if label in deltas_hz]
            for left_idx in range(len(ordered) - 1):
                label1 = ordered[left_idx]
                f1_hz = float(mean_freq_by_resonator[label1])
                df1_hz = float(deltas_hz[label1])
                for right_idx in range(left_idx + 1, len(ordered)):
                    label2 = ordered[right_idx]
                    f2_hz = float(mean_freq_by_resonator[label2])
                    separation_hz = f2_hz - f1_hz
                    if separation_hz < 0.0:
                        continue
                    if separation_hz > max_separation_hz:
                        break
                    left_f1_hz = float(left["resonators"][label1])
                    left_f2_hz = float(left["resonators"][label2])
                    right_f1_hz = float(right["resonators"][label1])
                    right_f2_hz = float(right["resonators"][label2])
                    df2_hz = float(deltas_hz[label2])
                    rel_delta_diff = (df2_hz - df1_hz) / f1_hz if f1_hz != 0.0 else np.nan
                    start_rel_sep = (left_f2_hz - left_f1_hz) / left_f1_hz if left_f1_hz != 0.0 else np.nan
                    end_rel_sep = (right_f2_hz - right_f1_hz) / right_f1_hz if right_f1_hz != 0.0 else np.nan
                    delta_diff_records.append(
                        {
                            "pair_idx": pair_idx,
                            "pair_label": pair_label,
                            "res1": label1,
                            "res2": label2,
                            "f1_hz": f1_hz,
                            "f2_hz": f2_hz,
                            "left_f1_hz": left_f1_hz,
                            "left_f2_hz": left_f2_hz,
                            "right_f1_hz": right_f1_hz,
                            "right_f2_hz": right_f2_hz,
                            "separation_hz": separation_hz,
                            "df1_hz": df1_hz,
                            "df2_hz": df2_hz,
                            "delta_diff_hz": df2_hz - df1_hz,
                            "rel_delta_diff": rel_delta_diff,
                            "start_rel_sep": start_rel_sep,
                            "end_rel_sep": end_rel_sep,
                        }
                    )

        if not delta_diff_records:
            raise ValueError(
                f"No resonator pairs within {max_separation_mhz:.3g} MHz were found across consecutive selected tests."
            )

        return {
            "tests": tests,
            "pair_labels": pair_labels,
            "records": delta_diff_records,
        }

    def open_resonator_pair_dfdiff_hist_window(self) -> None:
        if self.res_pair_dfdiff_hist_window is not None and self.res_pair_dfdiff_hist_window.winfo_exists():
            self.res_pair_dfdiff_hist_window.lift()
            self._render_resonator_pair_dfdiff_hist_window()
            return

        self.res_pair_dfdiff_hist_window = tk.Toplevel(self.root)
        self.res_pair_dfdiff_hist_window.title("Histogram of df2-df1")
        self.res_pair_dfdiff_hist_window.geometry("1320x860")
        self.res_pair_dfdiff_hist_window.protocol("WM_DELETE_WINDOW", self._close_resonator_pair_dfdiff_hist_window)

        controls = tk.Frame(self.res_pair_dfdiff_hist_window, padx=8, pady=8)
        controls.pack(side="top", fill="x")
        self.res_pair_dfdiff_hist_status_var = tk.StringVar(
            value="Showing counts of (df2-df1)/f1 for resonator pairs from consecutive selected tests."
        )
        self.res_pair_dfdiff_hist_sep_mhz_var = tk.DoubleVar(value=10.0)
        self.res_pair_dfdiff_hist_bin_mhz_var = tk.DoubleVar(value=0.0001)
        self.res_pair_dfdiff_hist_capture_var = tk.DoubleVar(value=1.0 / 20000.0)
        self.res_pair_dfdiff_hist_fit_mode_var = tk.StringVar(value="gennorm")
        self.res_pair_dfdiff_hist_num_res_var = tk.IntVar(value=1000)
        self.res_pair_dfdiff_hist_center_ghz_var = tk.DoubleVar(value=0.8)
        self.res_pair_dfdiff_hist_dfrel_var = tk.DoubleVar(value=1.8e-3)
        self.res_pair_dfdiff_hist_freq_jitter_khz_var = tk.DoubleVar(value=0.0)

        scale_wrap = tk.Frame(controls)
        scale_wrap.pack(side="top", fill="x")
        tk.Label(scale_wrap, text="N").pack(side="left", padx=(0, 4))
        num_res_entry = tk.Entry(scale_wrap, width=7, textvariable=self.res_pair_dfdiff_hist_num_res_var)
        num_res_entry.pack(side="left")
        num_res_entry.bind("<Return>", lambda _event: self._render_resonator_pair_dfdiff_hist_window())
        num_res_entry.bind("<FocusOut>", lambda _event: self._render_resonator_pair_dfdiff_hist_window())
        tk.Label(scale_wrap, text="Center (GHz)").pack(side="left", padx=(10, 4))
        center_entry = tk.Entry(scale_wrap, width=6, textvariable=self.res_pair_dfdiff_hist_center_ghz_var)
        center_entry.pack(side="left")
        center_entry.bind("<Return>", lambda _event: self._render_resonator_pair_dfdiff_hist_window())
        center_entry.bind("<FocusOut>", lambda _event: self._render_resonator_pair_dfdiff_hist_window())
        tk.Label(scale_wrap, text="df/f").pack(side="left", padx=(10, 4))
        dfrel_entry = tk.Entry(scale_wrap, width=8, textvariable=self.res_pair_dfdiff_hist_dfrel_var)
        dfrel_entry.pack(side="left")
        dfrel_entry.bind("<Return>", lambda _event: self._render_resonator_pair_dfdiff_hist_window())
        dfrel_entry.bind("<FocusOut>", lambda _event: self._render_resonator_pair_dfdiff_hist_window())
        tk.Button(
            scale_wrap, text="Refresh", width=10, command=self._render_resonator_pair_dfdiff_hist_window
        ).pack(side="right", padx=(8, 0))
        tk.Label(controls, textvariable=self.res_pair_dfdiff_hist_status_var, anchor="w", justify="left").pack(
            side="top", fill="x", pady=(6, 0)
        )
        tk.Label(scale_wrap, text="Max |f2-f1| (MHz)").pack(side="left", padx=(0, 4))
        self.res_pair_dfdiff_hist_sep_scale = tk.Scale(
            scale_wrap,
            from_=0.1,
            to=100.0,
            resolution=0.1,
            orient="horizontal",
            length=170,
            showvalue=True,
            variable=self.res_pair_dfdiff_hist_sep_mhz_var,
        )
        self.res_pair_dfdiff_hist_sep_scale.pack(side="left")
        self.res_pair_dfdiff_hist_sep_scale.bind(
            "<ButtonRelease-1>",
            lambda _event: self._render_resonator_pair_dfdiff_hist_window(),
        )
        self.res_pair_dfdiff_hist_sep_scale.bind(
            "<KeyRelease>",
            lambda _event: self._render_resonator_pair_dfdiff_hist_window(),
        )
        tk.Label(scale_wrap, text="Bin width").pack(side="left", padx=(10, 4))
        self.res_pair_dfdiff_hist_bin_scale = tk.Scale(
            scale_wrap,
            from_=0.00001,
            to=5.0,
            resolution=0.00001,
            orient="horizontal",
            length=160,
            showvalue=True,
            variable=self.res_pair_dfdiff_hist_bin_mhz_var,
        )
        self.res_pair_dfdiff_hist_bin_scale.pack(side="left")
        self.res_pair_dfdiff_hist_bin_scale.bind(
            "<ButtonRelease-1>",
            lambda _event: self._render_resonator_pair_dfdiff_hist_window(),
        )
        self.res_pair_dfdiff_hist_bin_scale.bind(
            "<KeyRelease>",
            lambda _event: self._render_resonator_pair_dfdiff_hist_window(),
        )
        tk.Label(scale_wrap, text="Collision +/-").pack(side="left", padx=(10, 4))
        self.res_pair_dfdiff_hist_capture_scale = tk.Scale(
            scale_wrap,
            from_=0.00001,
            to=0.0005,
            resolution=0.00001,
            orient="horizontal",
            length=160,
            showvalue=True,
            variable=self.res_pair_dfdiff_hist_capture_var,
        )
        self.res_pair_dfdiff_hist_capture_scale.pack(side="left")
        self.res_pair_dfdiff_hist_capture_scale.bind(
            "<ButtonRelease-1>",
            lambda _event: self._render_resonator_pair_dfdiff_hist_window(),
        )
        self.res_pair_dfdiff_hist_capture_scale.bind(
            "<KeyRelease>",
            lambda _event: self._render_resonator_pair_dfdiff_hist_window(),
        )
        tk.Label(scale_wrap, text="Offset std (kHz)").pack(side="left", padx=(10, 4))
        self.res_pair_dfdiff_hist_freq_jitter_scale = tk.Scale(
            scale_wrap,
            from_=0.0,
            to=1000.0,
            resolution=1.0,
            orient="horizontal",
            length=160,
            showvalue=True,
            variable=self.res_pair_dfdiff_hist_freq_jitter_khz_var,
        )
        self.res_pair_dfdiff_hist_freq_jitter_scale.pack(side="left")
        self.res_pair_dfdiff_hist_freq_jitter_scale.bind(
            "<ButtonRelease-1>",
            lambda _event: self._render_resonator_pair_dfdiff_hist_window(),
        )
        self.res_pair_dfdiff_hist_freq_jitter_scale.bind(
            "<KeyRelease>",
            lambda _event: self._render_resonator_pair_dfdiff_hist_window(),
        )
        tk.Label(scale_wrap, text="Fit").pack(side="left", padx=(10, 4))
        tk.Radiobutton(
            scale_wrap,
            text="Gaussian",
            value="gaussian",
            variable=self.res_pair_dfdiff_hist_fit_mode_var,
            command=self._render_resonator_pair_dfdiff_hist_window,
        ).pack(side="left")
        tk.Radiobutton(
            scale_wrap,
            text="Gen. normal",
            value="gennorm",
            variable=self.res_pair_dfdiff_hist_fit_mode_var,
            command=self._render_resonator_pair_dfdiff_hist_window,
        ).pack(side="left")

        self.res_pair_dfdiff_hist_figure = Figure(figsize=(12, 7))
        self.res_pair_dfdiff_hist_canvas = FigureCanvasTkAgg(
            self.res_pair_dfdiff_hist_figure,
            master=self.res_pair_dfdiff_hist_window,
        )
        self.res_pair_dfdiff_hist_toolbar = NavigationToolbar2Tk(
            self.res_pair_dfdiff_hist_canvas,
            self.res_pair_dfdiff_hist_window,
        )
        self.res_pair_dfdiff_hist_toolbar.update()
        self.res_pair_dfdiff_hist_toolbar.pack(side="top", fill="x")
        self.res_pair_dfdiff_hist_canvas.get_tk_widget().pack(fill="both", expand=True)

        self._render_resonator_pair_dfdiff_hist_window()

    def _close_resonator_pair_dfdiff_hist_window(self) -> None:
        if self.res_pair_dfdiff_hist_window is not None and self.res_pair_dfdiff_hist_window.winfo_exists():
            self.res_pair_dfdiff_hist_window.destroy()
        self.res_pair_dfdiff_hist_window = None
        self.res_pair_dfdiff_hist_canvas = None
        self.res_pair_dfdiff_hist_toolbar = None
        self.res_pair_dfdiff_hist_figure = None
        self.res_pair_dfdiff_hist_status_var = None
        self.res_pair_dfdiff_hist_sep_mhz_var = None
        self.res_pair_dfdiff_hist_bin_mhz_var = None
        self.res_pair_dfdiff_hist_capture_var = None
        self.res_pair_dfdiff_hist_fit_mode_var = None
        self.res_pair_dfdiff_hist_num_res_var = None
        self.res_pair_dfdiff_hist_center_ghz_var = None
        self.res_pair_dfdiff_hist_dfrel_var = None
        self.res_pair_dfdiff_hist_freq_jitter_khz_var = None
        self.res_pair_dfdiff_hist_sep_scale = None
        self.res_pair_dfdiff_hist_bin_scale = None
        self.res_pair_dfdiff_hist_capture_scale = None
        self.res_pair_dfdiff_hist_freq_jitter_scale = None
        self._res_pair_dfdiff_hist_ax = None

    def _render_resonator_pair_dfdiff_hist_window(self) -> None:
        if self.res_pair_dfdiff_hist_figure is None or self.res_pair_dfdiff_hist_canvas is None:
            return

        prior_xlim = None
        prior_ylim = None
        if self._res_pair_dfdiff_hist_ax is not None:
            try:
                prior_xlim = self._res_pair_dfdiff_hist_ax.get_xlim()
                prior_ylim = self._res_pair_dfdiff_hist_ax.get_ylim()
            except Exception:
                prior_xlim = None
                prior_ylim = None

        self.res_pair_dfdiff_hist_figure.clear()
        gs = self.res_pair_dfdiff_hist_figure.add_gridspec(
            2,
            2,
            width_ratios=[3.0, 2.0],
            height_ratios=[3.0, 2.0],
        )
        ax_hist = self.res_pair_dfdiff_hist_figure.add_subplot(gs[0, 0])
        ax_prob = self.res_pair_dfdiff_hist_figure.add_subplot(gs[0, 1])
        ax_cdf = self.res_pair_dfdiff_hist_figure.add_subplot(gs[1, :])
        ax = ax_hist
        self._res_pair_dfdiff_hist_ax = ax_hist

        max_sep_mhz = (
            float(self.res_pair_dfdiff_hist_sep_mhz_var.get())
            if self.res_pair_dfdiff_hist_sep_mhz_var is not None
            else 10.0
        )
        bin_width_mhz = (
            float(self.res_pair_dfdiff_hist_bin_mhz_var.get())
            if self.res_pair_dfdiff_hist_bin_mhz_var is not None
            else 0.1
        )
        bin_width_mhz = max(bin_width_mhz, 1.0e-6)
        capture_threshold = (
            float(self.res_pair_dfdiff_hist_capture_var.get())
            if self.res_pair_dfdiff_hist_capture_var is not None
            else (1.0 / 20000.0)
        )
        capture_threshold = max(capture_threshold, 0.0)
        num_resonators = (
            max(2, int(self.res_pair_dfdiff_hist_num_res_var.get()))
            if self.res_pair_dfdiff_hist_num_res_var is not None
            else 1000
        )
        center_ghz = (
            float(self.res_pair_dfdiff_hist_center_ghz_var.get())
            if self.res_pair_dfdiff_hist_center_ghz_var is not None
            else 0.8
        )
        if not np.isfinite(center_ghz) or center_ghz <= 0.0:
            center_ghz = 0.8
        dfrel_nominal = (
            float(self.res_pair_dfdiff_hist_dfrel_var.get())
            if self.res_pair_dfdiff_hist_dfrel_var is not None
            else 1.8e-3
        )
        if not np.isfinite(dfrel_nominal) or dfrel_nominal <= 0.0:
            dfrel_nominal = 1.8e-3
        freq_jitter_khz = (
            float(self.res_pair_dfdiff_hist_freq_jitter_khz_var.get())
            if self.res_pair_dfdiff_hist_freq_jitter_khz_var is not None
            else 0.0
        )
        freq_jitter_hz = max(0.0, freq_jitter_khz) * 1.0e3

        try:
            data = self._resonator_pair_dfdiff_hist_data(max_sep_mhz)
        except Exception as exc:
            ax.text(0.5, 0.5, str(exc), ha="center", va="center", transform=ax.transAxes)
            ax.set_axis_off()
            ax_prob.set_axis_off()
            ax_cdf.set_axis_off()
            if self.res_pair_dfdiff_hist_status_var is not None:
                self.res_pair_dfdiff_hist_status_var.set(str(exc))
            self.res_pair_dfdiff_hist_canvas.draw_idle()
            return

        values_mhz = np.asarray(
            [float(record["rel_delta_diff"]) for record in data["records"]],
            dtype=float,
        )
        values_mhz = values_mhz[np.isfinite(values_mhz)]
        if values_mhz.size == 0:
            ax.text(0.5, 0.5, "No finite (df2-df1)/f1 values were available.", ha="center", va="center", transform=ax.transAxes)
            ax.set_axis_off()
            if self.res_pair_dfdiff_hist_status_var is not None:
                self.res_pair_dfdiff_hist_status_var.set("No finite (df2-df1)/f1 values were available.")
            self.res_pair_dfdiff_hist_canvas.draw_idle()
            return

        vmin = float(np.min(values_mhz))
        vmax = float(np.max(values_mhz))
        if np.isclose(vmin, vmax):
            pad = max(0.5 * bin_width_mhz, 1.0e-3)
            half_width = max(bin_width_mhz, pad)
            bin_edges = np.asarray([-half_width, 0.0, half_width], dtype=float)
        else:
            start = bin_width_mhz * (np.floor(vmin / bin_width_mhz - 0.5) + 0.5)
            stop = bin_width_mhz * (np.ceil(vmax / bin_width_mhz - 0.5) + 0.5)
            if np.isclose(start, stop):
                stop = start + bin_width_mhz
            bin_edges = np.arange(start, stop + bin_width_mhz * 1.0001, bin_width_mhz, dtype=float)
            if bin_edges.size < 2:
                bin_edges = np.asarray([start, start + bin_width_mhz], dtype=float)

        pair_labels = list(data["pair_labels"])
        values_by_pair: list[np.ndarray] = []
        plotted_pair_labels: list[str] = []
        for pair_idx, pair_label in enumerate(pair_labels):
            pair_values = np.asarray(
                [
                    float(record["rel_delta_diff"])
                    for record in data["records"]
                    if int(record["pair_idx"]) == pair_idx and np.isfinite(float(record["rel_delta_diff"]))
                ],
                dtype=float,
            )
            if pair_values.size == 0:
                continue
            values_by_pair.append(pair_values)
            plotted_pair_labels.append(pair_label)

        colors = plt.cm.rainbow(np.linspace(0.0, 1.0, max(len(values_by_pair), 1)))
        counts_list, bins, _patches = ax.hist(
            values_by_pair,
            bins=bin_edges,
            stacked=True,
            color=colors[: len(values_by_pair)],
            alpha=0.8,
            edgecolor="black",
            linewidth=0.5,
            label=plotted_pair_labels,
        )
        counts = np.asarray(counts_list[-1], dtype=float) if len(counts_list) else np.asarray([], dtype=float)
        ax.grid(True, axis="y", alpha=0.3)
        ax.set_xlabel("(df2 - df1) / f1")
        ax.set_ylabel("Count")
        ax.set_title("Histogram of (df2-df1)/f1 for Nearby Resonator Pairs")

        fit_mode = (
            str(self.res_pair_dfdiff_hist_fit_mode_var.get())
            if self.res_pair_dfdiff_hist_fit_mode_var is not None
            else "gaussian"
        ).strip().lower()
        x_fit = np.linspace(float(bins[0]), float(bins[-1]), 600)
        fit_summary = " Fit unavailable."
        fit_label = None
        y_fit = None
        fit_dist = None

        try:
            if fit_mode == "gennorm":
                beta_hat, mu_hat, alpha_hat = stats.gennorm.fit(values_mhz)
                if np.isfinite(beta_hat) and np.isfinite(mu_hat) and np.isfinite(alpha_hat) and beta_hat > 0.0 and alpha_hat > 0.0:
                    pdf = stats.gennorm.pdf(x_fit, beta_hat, loc=mu_hat, scale=alpha_hat)
                    y_fit = values_mhz.size * bin_width_mhz * pdf
                    fit_label = f"Gen. normal fit: mu={mu_hat:.3g}, alpha={alpha_hat:.3g}, beta={beta_hat:.3g}"
                    fit_summary = f" Gen. normal fit: mu={mu_hat:.3g}, alpha={alpha_hat:.3g}, beta={beta_hat:.3g}."
                    fit_dist = stats.gennorm(beta_hat, loc=mu_hat, scale=alpha_hat)
            else:
                mu_hat, sigma_hat = stats.norm.fit(values_mhz)
                if np.isfinite(mu_hat) and np.isfinite(sigma_hat) and sigma_hat > 0.0:
                    pdf = stats.norm.pdf(x_fit, loc=mu_hat, scale=sigma_hat)
                    y_fit = values_mhz.size * bin_width_mhz * pdf
                    fit_label = f"Gaussian fit: mu={mu_hat:.3g}, sigma={sigma_hat:.3g}"
                    fit_summary = f" Gaussian fit: mu={mu_hat:.3g}, sigma={sigma_hat:.3g}."
                    fit_dist = stats.norm(loc=mu_hat, scale=sigma_hat)
        except Exception:
            y_fit = None
            fit_dist = None

        if y_fit is not None and fit_label is not None:
            ax.plot(
                x_fit,
                y_fit,
                color="black",
                linewidth=2.0,
                linestyle="-",
                label=fit_label,
                zorder=5,
            )
        if plotted_pair_labels and len(plotted_pair_labels) <= 12:
            ax.legend(loc="best", fontsize=8, title="Consecutive tests")
        elif y_fit is not None:
            ax.legend(loc="best", fontsize=8)

        if prior_xlim is not None:
            ax.set_xlim(prior_xlim)
        else:
            ax.set_xlim(float(bins[0]), float(bins[-1]))

        y_candidates = [0.0]
        if counts.size:
            y_candidates.append(float(np.max(counts)))
        if y_fit is not None:
            fit_max = np.asarray(y_fit, dtype=float)
            fit_max = fit_max[np.isfinite(fit_max)]
            if fit_max.size:
                y_candidates.append(float(np.max(fit_max)))
        y_top = max(y_candidates)
        y_pad = 1.0 if y_top <= 0.0 else 0.08 * y_top
        ax.set_ylim(0.0, y_top + y_pad)

        start_rel_values = np.asarray(
            [float(record["start_rel_sep"]) for record in data["records"]],
            dtype=float,
        )
        end_rel_values = np.asarray(
            [float(record["end_rel_sep"]) for record in data["records"]],
            dtype=float,
        )
        valid_prob_mask = np.isfinite(start_rel_values) & np.isfinite(end_rel_values)
        start_rel_values = start_rel_values[valid_prob_mask]
        end_rel_values = end_rel_values[valid_prob_mask]
        if start_rel_values.size:
            prob_bin_width = bin_width_mhz
            prob_start = 0.0
            prob_stop = 0.001
            prob_edges = np.arange(
                prob_start,
                prob_stop + prob_bin_width * 1.0001,
                prob_bin_width,
                dtype=float,
            )
            if prob_edges.size < 2:
                prob_edges = np.asarray([prob_start, prob_start + prob_bin_width], dtype=float)

            prob_x: list[float] = []
            prob_y: list[float] = []
            prob_n: list[int] = []
            for lo, hi in zip(prob_edges[:-1], prob_edges[1:]):
                if hi <= lo:
                    continue
                if hi == prob_edges[-1]:
                    mask = (start_rel_values >= lo) & (start_rel_values <= hi)
                else:
                    mask = (start_rel_values >= lo) & (start_rel_values < hi)
                if not np.any(mask):
                    continue
                total = int(np.count_nonzero(mask))
                success = int(np.count_nonzero(np.abs(end_rel_values[mask]) <= capture_threshold))
                prob_x.append(0.5 * (lo + hi))
                prob_y.append(success / total if total > 0 else np.nan)
                prob_n.append(total)

            if prob_x:
                if fit_dist is not None:
                    x_model = np.linspace(0.0, 0.001, 500)
                    lower = -capture_threshold - x_model
                    upper = capture_threshold - x_model
                    y_model = fit_dist.cdf(upper) - fit_dist.cdf(lower)
                    ax_prob.fill_between(
                        x_model,
                        0.0,
                        y_model,
                        color="tab:blue",
                        alpha=0.2,
                        zorder=1,
                    )
                    ax_prob.plot(
                        x_model,
                        y_model,
                        color="black",
                        linewidth=2.0,
                        linestyle="-",
                        label="Model implied",
                        zorder=2,
                    )
                else:
                    ax_prob.text(
                        0.5,
                        0.5,
                        "Model fit unavailable.",
                        ha="center",
                        va="center",
                        transform=ax_prob.transAxes,
                    )
            else:
                ax_prob.text(
                    0.5,
                    0.5,
                    "No populated start-separation bins.",
                    ha="center",
                    va="center",
                    transform=ax_prob.transAxes,
                )
        else:
            ax_prob.text(
                0.5,
                0.5,
                "No finite start/end separations were available.",
                ha="center",
                va="center",
                transform=ax_prob.transAxes,
            )

        ax_prob.axhline(0.0, color="0.5", linewidth=0.8, linestyle="--")
        ax_prob.axhline(1.0, color="0.5", linewidth=0.8, linestyle="--")
        ax_prob.grid(True, alpha=0.3)
        ax_prob.set_xlim(0.0, 0.001)
        ax_prob.set_ylim(0.0, 1.0)
        ax_prob.margins(x=0.0, y=0.0)
        ax_prob.set_xlabel("Starting relative separation (f2-f1) / f1")
        ax_prob.set_ylabel("Collision probability")
        ax_prob.set_title(f"P(|final separation| <= {capture_threshold:.3g}) vs starting separation")
        if fit_dist is not None:
            ax_prob.legend(loc="best", fontsize=8)

        if fit_dist is not None:
            pair_count = num_resonators * (num_resonators - 1) / 2.0
            rng = np.random.default_rng(12345)
            idx = np.arange(num_resonators, dtype=float) - 0.5 * (num_resonators - 1)
            base_freqs_hz = center_ghz * 1.0e9 * np.power(1.0 + dfrel_nominal, idx)
            jitter_hz = rng.normal(0.0, freq_jitter_hz, size=num_resonators)
            grid_hz = np.sort(base_freqs_hz + jitter_hz)
            ii, jj = np.triu_indices(num_resonators, k=1)
            f_lo = grid_hz[ii]
            f_hi = grid_hz[jj]
            valid_pairs = f_lo > 0.0
            rel_sep_samples = (f_hi[valid_pairs] - f_lo[valid_pairs]) / f_lo[valid_pairs]
            f_min_ghz = float(np.min(grid_hz)) / 1.0e9
            f_max_ghz = float(np.max(grid_hz)) / 1.0e9
            dist_label = (
                f"center {center_ghz:.3g} GHz, df/f {dfrel_nominal:.3g}"
                f", Gaussian offset std {freq_jitter_khz:.3g} kHz"
                f" | span {f_min_ghz:.3g}-{f_max_ghz:.3g} GHz"
            )
            if rel_sep_samples.size:
                initial_pair_probs = (rel_sep_samples <= capture_threshold).astype(float)
                lambda_initial = float(pair_count * np.mean(initial_pair_probs))
                cdf_x_max = int(max(1, num_resonators))
                x_counts = np.arange(1, cdf_x_max + 1, dtype=int)
                y_cdf_initial = stats.poisson.sf(x_counts - 1, lambda_initial)
                ax_cdf.plot(
                    x_counts,
                    y_cdf_initial,
                    color="tab:gray",
                    linewidth=1.8,
                    linestyle="--",
                    label=f"Initial model, E[K]={lambda_initial:.3g}",
                )
                step_colors = ["tab:purple", "tab:blue", "tab:green", "tab:orange", "tab:red"]
                lambda_by_step: dict[int, float] = {}
                delta_sample_count = 40000
                base_delta_samples = np.asarray(
                    fit_dist.rvs(size=(5, delta_sample_count), random_state=rng),
                    dtype=float,
                )
                for step_count, color in zip(range(1, 6), step_colors):
                    summed_delta = np.sum(base_delta_samples[:step_count, :], axis=0)
                    summed_delta = np.sort(summed_delta[np.isfinite(summed_delta)])
                    if summed_delta.size == 0:
                        continue
                    upper = capture_threshold - rel_sep_samples
                    lower = -capture_threshold - rel_sep_samples
                    upper_idx = np.searchsorted(summed_delta, upper, side="right")
                    lower_idx = np.searchsorted(summed_delta, lower, side="left")
                    model_pair_probs = (upper_idx - lower_idx) / float(summed_delta.size)
                    model_pair_probs = np.clip(np.asarray(model_pair_probs, dtype=float), 0.0, 1.0)
                    lambda_step = float(pair_count * np.mean(model_pair_probs))
                    lambda_by_step[step_count] = lambda_step
                    y_cdf = stats.poisson.sf(x_counts - 1, lambda_step)
                    ax_cdf.plot(
                        x_counts,
                        y_cdf,
                        color=color,
                        linewidth=2.0,
                        label=f"{step_count} step{'s' if step_count != 1 else ''}, E[K]={lambda_step:.3g}",
                    )
                ax_cdf.set_xlim(0, cdf_x_max)
                ax_cdf.set_xscale("log")
                ax_cdf.set_xlim(1, cdf_x_max)
                ax_cdf.set_ylim(0.0, 1.0)
                ax_cdf.set_xlabel("x collisions")
                ax_cdf.set_ylabel("P(K >= x)")
                ax_cdf.set_title(
                    f"Collision Count Exceedance | N={num_resonators}, {dist_label}"
                )
                ax_cdf.grid(True, alpha=0.3)
                ax_cdf.legend(loc="best", fontsize=8)
                ax_cdf.text(
                    0.01,
                    0.02,
                    "Note: E[K] from pair Monte Carlo/exact pair set; P(K>=x) from Poisson approximation.",
                    transform=ax_cdf.transAxes,
                    ha="left",
                    va="bottom",
                    fontsize=8,
                    color="0.35",
                )
                lambda_expected = float(lambda_by_step.get(1, np.nan))
            else:
                lambda_initial = np.nan
                lambda_expected = np.nan
                ax_cdf.text(
                    0.5,
                    0.5,
                    "No valid random pair samples were available.",
                    ha="center",
                    va="center",
                    transform=ax_cdf.transAxes,
                )
                ax_cdf.set_axis_off()
        else:
            lambda_initial = np.nan
            lambda_expected = np.nan
            ax_cdf.text(
                0.5,
                0.5,
                "Fit model unavailable for collision-count estimate.",
                ha="center",
                va="center",
                transform=ax_cdf.transAxes,
            )
            ax_cdf.set_axis_off()

        self.res_pair_dfdiff_hist_figure.tight_layout()
        if self.res_pair_dfdiff_hist_status_var is not None:
            pileup_text = (
                f" Expected initial collisions: {lambda_initial:.3g}. Expected post-shift collisions: {lambda_expected:.3g}. Layout: {dist_label}."
                if np.isfinite(lambda_expected)
                else " Expected collisions unavailable."
            )
            self.res_pair_dfdiff_hist_status_var.set(
                f"Showing {int(values_mhz.size)} (df2-df1)/f1 value(s) from {len(data['records'])} resonator pair(s) across {len(data['pair_labels'])} consecutive test pair(s). Max |f2-f1|={max_sep_mhz:.3g} MHz, bin width={bin_width_mhz:.3g}, collision +/-={capture_threshold:.3g}, peak count={int(np.max(counts)) if counts.size else 0}.{fit_summary}{pileup_text}"
            )
        self.res_pair_dfdiff_hist_canvas.draw_idle()

    @staticmethod
    def _coerce_frequency_to_scan_hz(value: float, scan: VNAScan) -> Optional[float]:
        freq = np.asarray(scan.freq, dtype=float)
        if freq.size == 0 or not np.isfinite(value):
            return None
        lo = float(np.min(freq))
        hi = float(np.max(freq))
        candidates = [float(value), float(value) * 1e9, float(value) * 1e6, float(value) * 1e3]
        for cand in candidates:
            if lo <= cand <= hi:
                return cand
        return None

    def _find_scan_for_sheet_identifier(self, identifier: str, target_value: float) -> tuple[VNAScan, float]:
        text = identifier.strip()
        lower = text.lower()
        scans = list(self.dataset.vna_scans)

        if lower.startswith("group "):
            try:
                group_num = int(lower.split()[1])
            except Exception as exc:
                raise ValueError(f"Invalid group name: {identifier}") from exc
            grouped = [scan for scan in scans if scan.plot_group == group_num]
            if not grouped:
                raise ValueError(f"No scans found for {identifier}.")
            for scan in grouped:
                target_hz = self._coerce_frequency_to_scan_hz(target_value, scan)
                if target_hz is not None:
                    return scan, target_hz
            raise ValueError(f"No scan in {identifier} contains frequency value {target_value}.")

        for scan in scans:
            name = Path(scan.filename).name.lower()
            stem = Path(scan.filename).stem.lower()
            if lower == name or lower == stem:
                target_hz = self._coerce_frequency_to_scan_hz(target_value, scan)
                if target_hz is None:
                    raise ValueError(f"Frequency value {target_value} is outside {Path(scan.filename).name}.")
                return scan, target_hz
        raise ValueError(f"No scan found matching filename {identifier}.")

    @staticmethod
    def _resonance_plot_window_hz(scan: VNAScan, target_hz: float) -> tuple[float, float]:
        freq = np.sort(np.asarray(scan.freq, dtype=float))
        if freq.size == 0:
            return target_hz, target_hz
        span = float(freq[-1] - freq[0])
        diffs = np.diff(freq)
        diffs = diffs[np.isfinite(diffs) & (diffs > 0)]
        step = float(np.median(diffs)) if diffs.size else max(span / max(freq.size - 1, 1), 1.0)
        half_width = max(250.0 * step, 0.002 * span)
        half_width = min(half_width, 0.05 * span if span > 0 else half_width)
        lo = max(float(freq[0]), target_hz - half_width)
        hi = min(float(freq[-1]), target_hz + half_width)
        if hi <= lo:
            half_width = max(1000.0 * step, 1.0)
            lo = max(float(freq[0]), target_hz - half_width)
            hi = min(float(freq[-1]), target_hz + half_width)
        return lo, hi

    @staticmethod
    def _interpolate_y(freq_hz: np.ndarray, y: np.ndarray, target_hz: float) -> float:
        x = np.asarray(freq_hz, dtype=float)
        vals = np.asarray(y, dtype=float)
        order = np.argsort(x)
        x = x[order]
        vals = vals[order]
        return float(np.interp(target_hz, x, vals))

    @staticmethod
    def _sheet_resonator_label(value: object) -> str:
        if value is None:
            return ""
        try:
            numeric = float(value)
            if np.isfinite(numeric) and numeric.is_integer():
                return str(int(numeric))
        except Exception:
            pass
        return str(value).strip()

    @staticmethod
    def _sheet_resonance_attachment(scan: VNAScan) -> dict[str, object]:
        payload = scan.candidate_resonators.get("sheet_resonances")
        if not isinstance(payload, dict):
            payload = {"assignments": {}}
            scan.candidate_resonators["sheet_resonances"] = payload
        assignments = payload.get("assignments")
        if not isinstance(assignments, dict):
            payload["assignments"] = {}
        return payload

    @staticmethod
    def _resonator_sort_key(label: str) -> tuple[int, object]:
        text = str(label).strip()
        try:
            value = int(text)
        except Exception:
            return (1, text)
        return (0, value)

    @staticmethod
    def _sheet_identifier_for_scan(scan: VNAScan, record: Optional[dict] = None) -> str:
        if scan.plot_group is not None:
            return f"Group {int(scan.plot_group)}"
        if isinstance(record, dict):
            identifier = str(record.get("identifier") or "").strip()
            if identifier:
                return identifier
        return Path(scan.filename).name

    def _normalize_sheet_resonance_identifiers(self) -> int:
        updated_count = 0
        for scan in self.dataset.vna_scans:
            payload = scan.candidate_resonators.get("sheet_resonances")
            if not isinstance(payload, dict):
                continue
            assignments = payload.get("assignments")
            if not isinstance(assignments, dict):
                continue
            for record in assignments.values():
                if not isinstance(record, dict):
                    continue
                expected = self._sheet_identifier_for_scan(scan, record)
                current = str(record.get("identifier") or "").strip()
                if current != expected:
                    record["identifier"] = expected
                    updated_count += 1
        return updated_count

    def _save_resonances_to_sheet(self, sheet_path: Path) -> int:
        normalized_count = self._normalize_sheet_resonance_identifiers()
        if normalized_count > 0:
            self.dataset.processing_history.append(
                _make_event(
                    "normalize_sheet_resonance_identifiers",
                    {"updated_count": int(normalized_count)},
                )
            )
        row_map: dict[str, dict[str, float]] = {}
        row_order: list[str] = []
        resonator_numbers: set[str] = set()
        assignment_count = 0

        for scan in self.dataset.vna_scans:
            payload = scan.candidate_resonators.get("sheet_resonances")
            if not isinstance(payload, dict):
                continue
            assignments = payload.get("assignments")
            if not isinstance(assignments, dict):
                continue
            for resonator_number, record in assignments.items():
                if not isinstance(record, dict):
                    continue
                try:
                    target_hz = float(record.get("frequency_hz"))
                except Exception:
                    continue
                identifier = self._sheet_identifier_for_scan(scan, record)
                if identifier not in row_map:
                    row_map[identifier] = {}
                    row_order.append(identifier)
                resonator_label = str(resonator_number).strip()
                row_map[identifier][resonator_label] = target_hz
                resonator_numbers.add(resonator_label)
                assignment_count += 1

        if not row_map or not resonator_numbers:
            raise ValueError("No marked resonator assignments were found on the loaded VNA scans.")

        wb = Workbook()
        ws = wb.active
        ws.title = "Resonators"
        ws.cell(row=1, column=1, value="")
        ws.cell(row=2, column=1, value="Identifier")

        ordered_resonators = sorted(resonator_numbers, key=self._resonator_sort_key)
        for col_idx, resonator_label in enumerate(ordered_resonators, start=2):
            ws.cell(row=1, column=col_idx, value=resonator_label)
            ws.cell(row=2, column=col_idx, value="Frequency (Hz)")

        for row_idx, identifier in enumerate(row_order, start=3):
            ws.cell(row=row_idx, column=1, value=identifier)
            for col_idx, resonator_label in enumerate(ordered_resonators, start=2):
                value = row_map[identifier].get(resonator_label)
                if value is not None:
                    ws.cell(row=row_idx, column=col_idx, value=float(value))

        sheet_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(sheet_path)
        wb.close()

        self.dataset.processing_history.append(
            _make_event(
                "save_resonances_to_sheet",
                {
                    "sheet_path": str(sheet_path),
                    "sheet_name": str(ws.title),
                    "assignment_count": int(assignment_count),
                    "normalized_count": int(normalized_count),
                },
            )
        )
        self._mark_dirty()
        self._autosave_dataset()
        if normalized_count > 0:
            self._log(f"Normalized {normalized_count} marked resonator identifier(s) before sheet export.")
        self._log(f"Saved {assignment_count} resonator assignment(s) to {sheet_path}")
        return assignment_count

    def _load_resonances_from_sheet(self, sheet_path: Path) -> int:
        wb = load_workbook(sheet_path, data_only=True, read_only=True)
        try:
            ws = wb.active

            column_headers: dict[int, str] = {}
            for col_idx in range(2, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx).value
                header = self._sheet_resonator_label(cell)
                column_headers[col_idx] = header

            row_records: list[dict] = []
            warnings: list[str] = []
            assignment_count = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                if not row:
                    continue
                identifier_raw = row[0]
                if identifier_raw is None or str(identifier_raw).strip() == "":
                    continue
                identifier = str(identifier_raw).strip()
                freq_entries: dict[int, dict] = {}
                for col_offset, cell in enumerate(row[1:], start=2):
                    resonator_number = column_headers.get(col_offset, "")
                    if cell is None or str(cell).strip() == "":
                        freq_entries[col_offset] = None
                        continue
                    try:
                        target_value = float(cell)
                    except Exception:
                        warnings.append(f"Row {row_idx}, column {col_offset}: skipped non-numeric frequency cell {cell!r}.")
                        freq_entries[col_offset] = None
                        continue
                    try:
                        scan, target_hz = self._find_scan_for_sheet_identifier(identifier, target_value)
                    except Exception as exc:
                        warnings.append(f"Row {row_idx}, column {col_offset}: {exc}")
                        freq_entries[col_offset] = None
                        continue
                    if resonator_number:
                        payload = self._sheet_resonance_attachment(scan)
                        assignments = payload["assignments"]
                        assignments[str(resonator_number)] = {
                            "frequency_hz": float(target_hz),
                            "sheet_path": str(sheet_path),
                            "sheet_name": str(ws.title),
                            "row": int(row_idx),
                            "column": int(col_offset),
                            "identifier": identifier,
                        }
                        assignment_count += 1
                    else:
                        warnings.append(
                            f"Row 1, column {col_offset}: blank resonator number header; frequency was plotted but not attached."
                        )
                    freq_entries[col_offset] = {
                        "scan": scan,
                        "target_hz": target_hz,
                        "target_value": target_value,
                        "resonator_number": resonator_number,
                    }
                row_records.append(
                    {
                        "row_idx": row_idx,
                        "identifier": identifier,
                        "entries": freq_entries,
                    }
                )

            data_columns = [col for col in range(2, ws.max_column + 1)]
        finally:
            wb.close()

        if not row_records or not data_columns:
            raise ValueError("No loadable resonance entries were found in the selected spreadsheet.")

        self.dataset.processing_history.append(
            _make_event(
                "load_resonances_from_sheet",
                {
                    "sheet_path": str(sheet_path),
                    "sheet_name": str(ws.title),
                    "assignment_count": int(assignment_count),
                },
            )
        )
        self._mark_dirty()
        self._autosave_dataset()
        for warning in warnings[:20]:
            self._log(f"Sheet resonance load warning: {warning}")
        self._log(f"Loaded {assignment_count} resonator assignment(s) from {sheet_path}")
        return assignment_count

    def _collect_attached_resonance_rows(self) -> tuple[list[dict], dict[int, str], list[int], list[str], str]:
        self._normalize_sheet_resonance_identifiers()
        grouped_rows: dict[int, dict] = {}
        column_headers: dict[int, str] = {}
        warnings: list[str] = []
        sheet_labels: set[str] = set()

        for scan in self.dataset.vna_scans:
            payload = scan.candidate_resonators.get("sheet_resonances")
            if not isinstance(payload, dict):
                continue
            assignments = payload.get("assignments")
            if not isinstance(assignments, dict):
                continue
            for resonator_number, record in assignments.items():
                if not isinstance(record, dict):
                    continue
                try:
                    row_idx = int(record.get("row"))
                    col_idx = int(record.get("column"))
                    target_hz = float(record.get("frequency_hz"))
                except Exception:
                    warnings.append(f"{Path(scan.filename).name}: skipped malformed resonator marker record.")
                    continue
                identifier = str(record.get("identifier") or Path(scan.filename).name)
                sheet_path = str(record.get("sheet_path") or "")
                sheet_name = str(record.get("sheet_name") or "")
                if sheet_path or sheet_name:
                    sheet_labels.add(f"{Path(sheet_path).name} | {sheet_name}".strip(" |"))
                column_headers[col_idx] = str(resonator_number).strip()
                row_record = grouped_rows.setdefault(
                    row_idx,
                    {
                        "row_idx": row_idx,
                        "identifier": identifier,
                        "entries": {},
                    },
                )
                row_record["entries"][col_idx] = {
                    "scan": scan,
                    "target_hz": target_hz,
                    "target_value": target_hz,
                    "resonator_number": str(resonator_number).strip(),
                }

        if not grouped_rows or not column_headers:
            raise ValueError("No marked resonator assignments were found on the loaded VNA scans.")

        row_records = [grouped_rows[idx] for idx in sorted(grouped_rows)]
        data_columns = sorted(column_headers)
        source_label = "; ".join(sorted(label for label in sheet_labels if label))
        return row_records, column_headers, data_columns, warnings, source_label

    def _plot_resonance_rows(
        self,
        *,
        row_records: list[dict],
        column_headers: dict[int, str],
        data_columns: list[int],
        data_mode: str,
        warnings: list[str],
        source_label: str,
    ) -> list[Path]:
        if not row_records or not data_columns:
            raise ValueError("No plottable resonance entries are available.")

        out_dir = _dataset_dir(self.dataset) / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_sheet_resonance_plots"
        out_dir.mkdir(parents=True, exist_ok=True)
        saved_paths: list[Path] = []

        column_chunks = [data_columns[i : i + 3] for i in range(0, len(data_columns), 3)]
        half_window_hz = 0.5e6

        for page_idx, page_columns in enumerate(column_chunks, start=1):
            nrows = len(row_records)
            ncols = len(page_columns)
            fig, axes = plt.subplots(
                nrows,
                ncols,
                figsize=(5.2 * ncols, max(2.2 * nrows, 4.0)),
                squeeze=False,
                sharex="col",
            )

            column_ranges: dict[int, tuple[float, float]] = {}
            for col in page_columns:
                targets = []
                for row_record in row_records:
                    entry = row_record["entries"].get(col)
                    if isinstance(entry, dict):
                        targets.append(float(entry["target_hz"]))
                if targets:
                    column_ranges[col] = (min(targets) - half_window_hz, max(targets) + half_window_hz)
                else:
                    column_ranges[col] = (0.0, 1.0)

            for row_pos, row_record in enumerate(row_records):
                for col_pos, col in enumerate(page_columns):
                    ax = axes[row_pos][col_pos]
                    entry = row_record["entries"].get(col)
                    if entry is None:
                        ax.axis("off")
                        continue

                    scan = entry["scan"]
                    freq = np.asarray(scan.freq, dtype=float)
                    if data_mode == "normalized":
                        norm = scan.baseline_filter.get("normalized", {})
                        amp, _phase = _read_polar_series(
                            norm,
                            amplitude_key="norm_amp",
                            phase_key="norm_phase_deg_unwrapped",
                        )
                        if amp.shape == scan.freq.shape:
                            y = np.asarray(amp, dtype=float)
                            y_label = "Normalized |S21|"
                        else:
                            y = np.asarray(scan.amplitude(), dtype=float)
                            y_label = "|S21|"
                            warnings.append(
                                f"{Path(scan.filename).name}: normalized data missing; used raw amplitude instead."
                            )
                    else:
                        y = np.asarray(scan.amplitude(), dtype=float)
                        y_label = "|S21|"

                    order = np.argsort(freq)
                    freq_sorted = freq[order]
                    y_sorted = y[order]
                    col_lo, col_hi = column_ranges[col]
                    full_mask = (freq_sorted >= col_lo) & (freq_sorted <= col_hi)
                    if not np.any(full_mask):
                        ax.text(0.5, 0.5, "No data in column range", ha="center", va="center")
                        ax.axis("off")
                        continue

                    ax.plot(
                        freq_sorted[full_mask] / 1.0e9,
                        y_sorted[full_mask],
                        color="0.65",
                        linewidth=0.8,
                        zorder=1,
                    )

                    target_hz = float(entry["target_hz"])
                    local_lo = target_hz - half_window_hz
                    local_hi = target_hz + half_window_hz
                    local_mask = (freq_sorted >= local_lo) & (freq_sorted <= local_hi)
                    if np.any(local_mask):
                        ax.plot(
                            freq_sorted[local_mask] / 1.0e9,
                            y_sorted[local_mask],
                            color="tab:blue",
                            linewidth=2.2,
                            zorder=2,
                        )

                    show_labels = row_pos == 0 and col_pos == 0
                    cand = scan.candidate_resonators
                    gaussian_freqs = np.asarray(
                        cand.get("gaussian_convolution", {}).get("candidate_freq", np.array([])),
                        dtype=float,
                    )
                    dsdf_freqs = np.asarray(
                        cand.get("dsdf_gaussian_convolution", {}).get("candidate_freq", np.array([])),
                        dtype=float,
                    )
                    for freqs_hz, style, label in (
                        (
                            gaussian_freqs,
                            dict(
                                linestyle="none",
                                marker="o",
                                markersize=7,
                                markerfacecolor="none",
                                markeredgewidth=1.5,
                                color="green",
                            ),
                            "Gaussian candidates",
                        ),
                        (
                            dsdf_freqs,
                            dict(linestyle="none", marker="D", markersize=5, color="red"),
                            "dS21/df peaks",
                        ),
                    ):
                        in_win = freqs_hz[(freqs_hz >= col_lo) & (freqs_hz <= col_hi)]
                        if in_win.size == 0:
                            continue
                        y_pts = [self._interpolate_y(freq_sorted, y_sorted, float(f)) for f in in_win]
                        ax.plot(in_win / 1.0e9, y_pts, label=(label if show_labels else None), **style)

                    target_y = self._interpolate_y(freq_sorted, y_sorted, target_hz)
                    ax.plot(
                        [target_hz / 1.0e9],
                        [target_y],
                        linestyle="none",
                        marker="s",
                        markersize=7,
                        color="black",
                        label=("Spreadsheet frequency" if show_labels else None),
                    )
                    ax.set_xlim(col_lo / 1.0e9, col_hi / 1.0e9)
                    ax.grid(True, alpha=0.3)
                    ax.set_title(
                        f"{row_record['identifier']} | {Path(scan.filename).name}\n{target_hz / 1.0e9:.9g} GHz",
                        fontsize=8,
                    )
                    if col_pos == 0:
                        ax.set_ylabel(y_label)
                    if row_pos == nrows - 1:
                        ax.set_xlabel("Frequency (GHz)")
                    if show_labels:
                        ax.legend(loc="upper right", fontsize=8)

            for col_pos, col in enumerate(page_columns):
                resonator_number = column_headers.get(col, "")
                resonator_label = (
                    f"Resonator {resonator_number}"
                    if resonator_number
                    else f"Sheet column {col}"
                )
                axes[0][col_pos].set_title(
                    f"{resonator_label} | {column_ranges[col][0] / 1.0e9:.9g} to {column_ranges[col][1] / 1.0e9:.9g} GHz",
                    fontsize=9,
                )

            fig.suptitle(
                f"Resonator Marker Plots | mode={'baseline normalized' if data_mode == 'normalized' else 'raw'} | page {page_idx}"
                + (f" | {source_label}" if source_label else ""),
                fontsize=12,
            )
            fig.tight_layout()
            out_path = out_dir / f"resonance_sheet_page_{page_idx:02d}.pdf"
            fig.savefig(out_path)
            plt.close(fig)
            saved_paths.append(out_path)

        self.dataset.processing_history.append(
            _make_event(
                "plot_attached_resonances",
                {
                    "data_mode": data_mode,
                    "plot_count": int(
                        sum(sum(1 for entry in r["entries"].values() if isinstance(entry, dict)) for r in row_records)
                    ),
                    "page_count": len(saved_paths),
                    "output_dir": str(out_dir),
                    "source_label": source_label,
                },
            )
        )
        self._mark_dirty()
        self._autosave_dataset()
        for warning in warnings[:20]:
            self._log(f"Attached resonance plot warning: {warning}")
        return saved_paths

    def _plot_attached_resonances_draw_page(
        self,
        ax,
        rows: list[dict],
        *,
        spacing: float = 1.5,
        truncate_threshold: float = 1.5,
        xlim_ghz: Optional[tuple[float, float]] = None,
        title: str = "",
    ) -> int:
        offset_by_scan_key, tick_info = self._attached_resonance_editor_offset_map(rows, spacing)
        trace_colors = self._attached_resonance_editor_trace_colors()
        resonator_tracks: dict[str, list[tuple[float, float]]] = {}
        resonator_markers: list[dict] = []
        marker_count = 0
        x_range_ghz = xlim_ghz

        for row in rows:
            offset = float(offset_by_scan_key[str(row["scan_key"])])
            freq_ghz = np.asarray(row["freq"], dtype=float) / 1.0e9
            amp_display = np.minimum(np.asarray(row["amp"], dtype=float), truncate_threshold)
            y = amp_display + offset
            trace_color = trace_colors[int(row.get("scan_index", 0)) % len(trace_colors)]
            ax.plot(freq_ghz, y, linewidth=1.0, color=trace_color, alpha=0.9, zorder=1)

            for resonator in row["resonators"]:
                target_hz = float(resonator["target_hz"])
                target_ghz = target_hz / 1.0e9
                y_pt = self._interpolate_y(row["freq"], amp_display, target_hz) + offset
                resonator_number = str(resonator["resonator_number"])
                resonator_markers.append(
                    {
                        "resonator_number": resonator_number,
                        "x_ghz": target_ghz,
                        "y": y_pt,
                    }
                )
                resonator_tracks.setdefault(resonator_number, []).append((target_ghz, y_pt))
                marker_count += 1

        for resonator_number, points in resonator_tracks.items():
            if x_range_ghz is not None:
                points = [pt for pt in points if x_range_ghz[0] <= pt[0] <= x_range_ghz[1]]
            if len(points) < 2:
                continue
            points = sorted(points, key=lambda item: item[1], reverse=True)
            ax.plot(
                [pt[0] for pt in points],
                [pt[1] for pt in points],
                linestyle=":",
                linewidth=1.0,
                color="tab:red",
                alpha=0.9,
                zorder=2,
            )

        for point in resonator_markers:
            if x_range_ghz is not None and not (x_range_ghz[0] <= point["x_ghz"] <= x_range_ghz[1]):
                continue
            ax.plot(
                [point["x_ghz"]],
                [point["y"]],
                linestyle="none",
                marker="o",
                markersize=6,
                markerfacecolor="none",
                markeredgecolor="tab:red",
                markeredgewidth=1.5,
                zorder=4,
            )
            ax.text(
                point["x_ghz"],
                point["y"] - 0.18,
                point["resonator_number"],
                ha="center",
                va="top",
                fontsize=8,
                color="tab:red",
                zorder=5,
            )

        y_low = min(
            float(np.min(np.minimum(np.asarray(row["amp"], dtype=float), truncate_threshold)))
            + float(offset_by_scan_key[str(row["scan_key"])])
            for row in rows
        )
        y_high = max(
            float(np.max(np.minimum(np.asarray(row["amp"], dtype=float), truncate_threshold)))
            + float(offset_by_scan_key[str(row["scan_key"])])
            for row in rows
        )
        freq_min = min(float(row["freq"][0]) for row in rows) / 1.0e9
        freq_max = max(float(row["freq"][-1]) for row in rows) / 1.0e9
        x_pad = max((freq_max - freq_min) * 0.01, 1e-6)

        ax.set_xlabel("Frequency (GHz)")
        ax.set_ylabel("Normalized |S21| + vertical offset")
        ax.set_yticks([item[0] for item in tick_info])
        ax.set_yticklabels([item[1] for item in tick_info], fontsize=8)
        ax.set_ylim(y_low - 0.2, y_high + 0.2)
        if xlim_ghz is None:
            ax.set_xlim(freq_min - 0.5 * x_pad, freq_max + 2.0 * x_pad)
        else:
            ax.set_xlim(xlim_ghz)
        ax.grid(True, alpha=0.3)
        if title:
            ax.set_title(title, fontsize=11)
        return marker_count

    def _plot_attached_resonances(
        self,
        *,
        progress_callback: Optional[Callable[[int, int], None]] = None,
    ) -> list[Path]:
        rows, warnings = self._selected_scans_for_attached_resonance_editor()
        if not rows:
            raise ValueError("No selected scans with normalized data are available for resonator-marker plotting.")

        out_dir = _dataset_dir(self.dataset) / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_resonator_marker_plots"
        out_dir.mkdir(parents=True, exist_ok=True)
        pdf_path = out_dir / "resonator_markers.pdf"

        freq_min_hz = min(float(row["freq"][0]) for row in rows)
        freq_max_hz = max(float(row["freq"][-1]) for row in rows)
        marker_freqs_hz = sorted(
            float(resonator["target_hz"])
            for row in rows
            for resonator in row["resonators"]
        )
        zoom_span_hz = 100.0e6
        zoom_step_hz = 80.0e6
        zoom_windows: list[tuple[float, float]] = []
        if marker_freqs_hz and freq_max_hz - freq_min_hz > zoom_span_hz:
            start_hz = freq_min_hz
            seen_starts: set[float] = set()
            while True:
                rounded_start = round(start_hz, 3)
                if rounded_start in seen_starts:
                    break
                seen_starts.add(rounded_start)
                end_hz = min(start_hz + zoom_span_hz, freq_max_hz)
                if any(start_hz <= marker_hz <= end_hz for marker_hz in marker_freqs_hz):
                    zoom_windows.append((start_hz, end_hz))
                if end_hz >= freq_max_hz:
                    break
                next_start = start_hz + zoom_step_hz
                if next_start + zoom_span_hz > freq_max_hz:
                    next_start = max(freq_min_hz, freq_max_hz - zoom_span_hz)
                if next_start <= start_hz:
                    break
                start_hz = next_start

        total_pages = 1 + len(zoom_windows)
        page_count = 0
        marker_count = 0
        with PdfPages(pdf_path) as pdf:
            fig = Figure(figsize=(12, 7))
            ax = fig.add_subplot(111)
            marker_count += self._plot_attached_resonances_draw_page(
                ax,
                rows,
                title="Resonator Markers | Full Span",
            )
            fig.tight_layout()
            pdf.savefig(fig)
            plt.close(fig)
            page_count += 1
            if progress_callback is not None:
                progress_callback(page_count, total_pages)

            for page_idx, (lo_hz, hi_hz) in enumerate(zoom_windows, start=1):
                fig = Figure(figsize=(12, 7))
                ax = fig.add_subplot(111)
                marker_count += self._plot_attached_resonances_draw_page(
                    ax,
                    rows,
                    xlim_ghz=(lo_hz / 1.0e9, hi_hz / 1.0e9),
                    title=(
                        f"Resonator Markers | Zoom {page_idx} | "
                        f"{lo_hz / 1.0e9:.6f} to {hi_hz / 1.0e9:.6f} GHz"
                    ),
                )
                fig.tight_layout()
                pdf.savefig(fig)
                plt.close(fig)
                page_count += 1
                if progress_callback is not None:
                    progress_callback(page_count, total_pages)

        self.dataset.processing_history.append(
            _make_event(
                "plot_attached_resonances",
                {
                    "page_count": int(page_count),
                    "plot_count": int(marker_count),
                    "output_pdf": str(pdf_path),
                    "zoom_span_mhz": 100.0,
                    "zoom_overlap_mhz": 20.0,
                },
            )
        )
        self._mark_dirty()
        self._autosave_dataset()
        for warning in warnings[:20]:
            self._log(f"Resonator marker plot warning: {warning}")
        self._log(f"Saved resonator marker PDF: {pdf_path}")
        return [pdf_path]

    def open_attached_resonance_editor(self) -> None:
        scans = self._selected_scans()
        if not scans:
            messagebox.showwarning(
                "No selection",
                "No scans selected for analysis.\nUse 'Select Scans for Analysis' first.",
            )
            return

        if self.attached_res_edit_window is not None and self.attached_res_edit_window.winfo_exists():
            self.attached_res_edit_window.lift()
            self._render_attached_resonance_editor()
            return

        self.attached_res_edit_window = tk.Toplevel(self.root)
        self.attached_res_edit_window.title("Edit Resonator Markers")
        self.attached_res_edit_window.geometry("1320x900")
        self.attached_res_edit_window.protocol("WM_DELETE_WINDOW", self._attached_resonance_editor_exit)

        controls = tk.Frame(self.attached_res_edit_window, padx=8, pady=8)
        controls.pack(side="top", fill="x")
        self.attached_res_edit_status_var = tk.StringVar(value="Normalized selected scans will be plotted.")
        tk.Label(controls, textvariable=self.attached_res_edit_status_var, anchor="w").pack(
            side="left", fill="x", expand=True
        )
        number_controls = tk.Frame(controls)
        number_controls.pack(side="right")
        self.attached_res_edit_add_button = tk.Button(
            number_controls, text="Add Resonator", width=14, command=self._attached_resonance_editor_toggle_add
        )
        self.attached_res_edit_add_button.pack(side="left")
        tk.Label(number_controls, text="Spacing").pack(side="left", padx=(10, 4))
        self.attached_res_edit_spacing_var = tk.DoubleVar(value=1.5)
        self.attached_res_edit_spacing_scale = tk.Scale(
            number_controls,
            from_=0.0,
            to=3.0,
            resolution=0.05,
            orient="horizontal",
            length=120,
            showvalue=True,
            variable=self.attached_res_edit_spacing_var,
        )
        self.attached_res_edit_spacing_scale.pack(side="left")
        self.attached_res_edit_spacing_scale.bind(
            "<ButtonRelease-1>",
            self._attached_resonance_editor_on_spacing_release,
        )
        self.attached_res_edit_spacing_scale.bind(
            "<KeyRelease>",
            self._attached_resonance_editor_on_spacing_release,
        )
        tk.Label(number_controls, text="Search width (kHz)").pack(side="left", padx=(10, 4))
        self.attached_res_edit_search_window_khz_var = tk.DoubleVar(value=300.0)
        self.attached_res_edit_search_window_scale = tk.Scale(
            number_controls,
            from_=25.0,
            to=2000.0,
            resolution=25.0,
            orient="horizontal",
            length=140,
            showvalue=True,
            variable=self.attached_res_edit_search_window_khz_var,
        )
        self.attached_res_edit_search_window_scale.pack(side="left")
        self.attached_res_edit_truncate_var = tk.BooleanVar(value=True)
        tk.Checkbutton(
            number_controls,
            text="Truncate |S21|",
            variable=self.attached_res_edit_truncate_var,
            command=self._attached_resonance_editor_on_truncate_toggle,
        ).pack(side="left", padx=(10, 4))
        self.attached_res_edit_truncate_threshold_var = tk.DoubleVar(value=1.5)
        self.attached_res_edit_truncate_threshold_scale = tk.Scale(
            number_controls,
            from_=1.0,
            to=2.0,
            resolution=0.05,
            orient="horizontal",
            length=110,
            showvalue=True,
            variable=self.attached_res_edit_truncate_threshold_var,
        )
        self.attached_res_edit_truncate_threshold_scale.pack(side="left")
        self.attached_res_edit_truncate_threshold_scale.bind(
            "<ButtonRelease-1>",
            self._attached_resonance_editor_on_truncate_release,
        )
        self.attached_res_edit_truncate_threshold_scale.bind(
            "<KeyRelease>",
            self._attached_resonance_editor_on_truncate_release,
        )
        tk.Label(number_controls, text="Working #").pack(side="left", padx=(10, 4))
        self.attached_res_edit_working_number_var = tk.StringVar(value="1")
        self.attached_res_edit_working_number_spinbox = tk.Spinbox(
            number_controls,
            from_=1,
            to=9999,
            increment=1,
            width=6,
            textvariable=self.attached_res_edit_working_number_var,
        )
        self.attached_res_edit_working_number_spinbox.pack(side="left")
        self.attached_res_edit_renumber_button = tk.Button(
            controls,
            text="Renumber Low->High",
            width=18,
            command=self._attached_resonance_editor_renumber_low_to_high,
        )
        self.attached_res_edit_renumber_button.pack(side="right", padx=(8, 0))
        self.attached_res_edit_undo_button = tk.Button(
            controls, text="Undo", width=10, command=self._attached_resonance_editor_undo
        )
        self.attached_res_edit_undo_button.pack(side="right", padx=(8, 0))
        tk.Button(
            controls, text="Delete Selected", width=14, command=self._attached_resonance_editor_delete_selected
        ).pack(side="right", padx=(8, 0))
        tk.Button(
            controls,
            text="Clear Sel. Scan Markers",
            width=20,
            command=self._attached_resonance_editor_clear_selected_scan_markers,
        ).pack(side="right", padx=(8, 0))
        tk.Button(
            controls, text="Reset View", width=12, command=self._attached_resonance_editor_reset_view
        ).pack(side="right", padx=(8, 0))
        self.attached_res_edit_save_button = tk.Button(
            controls, text="Save", width=10, command=self._attached_resonance_editor_save
        )
        self.attached_res_edit_save_button.pack(side="right", padx=(8, 0))
        self.attached_res_edit_exit_button = tk.Button(
            controls, text="Exit", width=12, command=self._attached_resonance_editor_exit
        )
        self.attached_res_edit_exit_button.pack(side="right", padx=(8, 0))

        self.attached_res_edit_figure = Figure(figsize=(12, 7))
        self.attached_res_edit_canvas = FigureCanvasTkAgg(
            self.attached_res_edit_figure, master=self.attached_res_edit_window
        )
        self.attached_res_edit_toolbar = NavigationToolbar2Tk(
            self.attached_res_edit_canvas, self.attached_res_edit_window
        )
        self.attached_res_edit_toolbar.update()
        self.attached_res_edit_toolbar.pack(side="top", fill="x")
        self.attached_res_edit_canvas.get_tk_widget().pack(fill="both", expand=True)
        self.attached_res_edit_canvas.mpl_connect(
            "button_press_event", self._attached_resonance_editor_on_click
        )

        self._attached_res_edit_selected = None
        self._attached_res_edit_pending_add = False
        self._attached_res_edit_default_xlim = None
        self._attached_res_edit_missing_normalized_warned = None
        self._attached_res_edit_snapshot = self._attached_resonance_editor_capture_snapshot()
        self._attached_res_edit_undo_stack = []
        self._attached_res_edit_changed = False
        self._attached_resonance_editor_reset_working_number()
        self._render_attached_resonance_editor()

    def _close_attached_resonance_editor(self) -> None:
        if self.attached_res_edit_window is not None and self.attached_res_edit_window.winfo_exists():
            self.attached_res_edit_window.destroy()
        self.attached_res_edit_window = None
        self.attached_res_edit_canvas = None
        self.attached_res_edit_toolbar = None
        self.attached_res_edit_figure = None
        self.attached_res_edit_status_var = None
        self.attached_res_edit_working_number_var = None
        self.attached_res_edit_working_number_spinbox = None
        self.attached_res_edit_spacing_var = None
        self.attached_res_edit_spacing_scale = None
        self.attached_res_edit_search_window_khz_var = None
        self.attached_res_edit_search_window_scale = None
        self.attached_res_edit_truncate_var = None
        self.attached_res_edit_truncate_threshold_var = None
        self.attached_res_edit_truncate_threshold_scale = None
        self.attached_res_edit_add_button = None
        self.attached_res_edit_renumber_button = None
        self.attached_res_edit_undo_button = None
        self.attached_res_edit_save_button = None
        self.attached_res_edit_exit_button = None
        self.attached_res_edit_ax = None
        self._attached_res_edit_points = []
        self._attached_res_edit_rows_cache = []
        self._attached_res_edit_offset_by_scan_key = {}
        self._attached_res_edit_selected = None
        self._attached_res_edit_pending_add = False
        self._attached_res_edit_default_xlim = None
        self._attached_res_edit_missing_normalized_warned = None
        self._attached_res_edit_snapshot = None
        self._attached_res_edit_undo_stack = []
        self._attached_res_edit_changed = False

    def _attached_resonance_editor_capture_snapshot(self) -> dict:
        scan_payloads: dict[str, object] = {}
        scan_history_lengths: dict[str, int] = {}
        for scan in self._selected_scans():
            key = self._scan_key(scan)
            payload = scan.candidate_resonators.get("sheet_resonances")
            scan_payloads[key] = copy.deepcopy(payload) if isinstance(payload, dict) else None
            scan_history_lengths[key] = len(scan.processing_history)
        return {
            "scan_payloads": scan_payloads,
            "scan_history_lengths": scan_history_lengths,
            "dataset_processing_history_len": len(self.dataset.processing_history),
            "dataset_transcript_len": len(self.dataset.transcript),
            "was_dirty": self._dirty,
            "selected": copy.deepcopy(self._attached_res_edit_selected),
            "working_number": self._attached_resonance_editor_working_number(),
        }

    def _attached_resonance_editor_apply_snapshot(self, snapshot: Optional[dict]) -> None:
        if not isinstance(snapshot, dict):
            return
        payloads = snapshot.get("scan_payloads", {})
        history_lengths = snapshot.get("scan_history_lengths", {})
        for scan in self.dataset.vna_scans:
            key = self._scan_key(scan)
            if key not in payloads:
                continue
            payload = payloads[key]
            if isinstance(payload, dict):
                scan.candidate_resonators["sheet_resonances"] = copy.deepcopy(payload)
            else:
                scan.candidate_resonators.pop("sheet_resonances", None)
            prior_len = int(history_lengths.get(key, len(scan.processing_history)))
            if len(scan.processing_history) > prior_len:
                del scan.processing_history[prior_len:]
        dataset_processing_len = int(snapshot.get("dataset_processing_history_len", len(self.dataset.processing_history)))
        if len(self.dataset.processing_history) > dataset_processing_len:
            del self.dataset.processing_history[dataset_processing_len:]
        dataset_transcript_len = int(snapshot.get("dataset_transcript_len", len(self.dataset.transcript)))
        if len(self.dataset.transcript) > dataset_transcript_len:
            del self.dataset.transcript[dataset_transcript_len:]
        self._dirty = bool(snapshot.get("was_dirty", self._dirty))
        selected = snapshot.get("selected")
        if isinstance(selected, tuple) and len(selected) == 2:
            self._attached_res_edit_selected = (str(selected[0]), str(selected[1]))
        else:
            self._attached_res_edit_selected = None
        if self.attached_res_edit_working_number_var is not None:
            working_number = str(snapshot.get("working_number", self._attached_resonance_editor_next_number()))
            self.attached_res_edit_working_number_var.set(working_number)
        self._refresh_status()
        self._reload_transcript_ui()

    def _attached_resonance_editor_restore_snapshot(self) -> None:
        self._attached_resonance_editor_apply_snapshot(self._attached_res_edit_snapshot)

    def _attached_resonance_editor_push_undo_snapshot(self) -> None:
        self._attached_res_edit_undo_stack.append(self._attached_resonance_editor_capture_snapshot())
        self._attached_resonance_editor_update_undo_button()

    def _attached_resonance_editor_save(self) -> bool:
        if not self._attached_res_edit_changed:
            if self.attached_res_edit_status_var is not None:
                self.attached_res_edit_status_var.set("Attached resonator edits are already saved.")
            self._attached_resonance_editor_update_save_button()
            return True
        self._mark_dirty()
        if not self._autosave_dataset():
            self._attached_resonance_editor_update_save_button()
            return False
        self._attached_res_edit_changed = False
        self._attached_res_edit_snapshot = self._attached_resonance_editor_capture_snapshot()
        self._attached_res_edit_undo_stack = []
        if self.attached_res_edit_status_var is not None:
            self.attached_res_edit_status_var.set("Attached resonator edits saved.")
        self._attached_resonance_editor_update_undo_button()
        self._attached_resonance_editor_update_save_button()
        self._log("Saved resonator marker edits.")
        return True

    def _attached_resonance_editor_exit(self) -> None:
        if not self._attached_res_edit_changed:
            self._close_attached_resonance_editor()
            return
        dialog = messagebox.Message(
            parent=self.attached_res_edit_window,
            title="Save resonator marker edits?",
            message="Save resonator marker edits before exiting?",
            icon=messagebox.WARNING,
            type=messagebox.YESNOCANCEL,
            default=messagebox.YES,
        )
        response = str(dialog.show()).lower()
        if response == "cancel":
            return
        if response == "yes":
            if not self._attached_resonance_editor_save():
                return
        else:
            self._attached_resonance_editor_restore_snapshot()
        self._close_attached_resonance_editor()

    def _selected_scans_for_attached_resonance_editor(
        self,
    ) -> tuple[list[dict], list[str]]:
        rows: list[dict] = []
        warnings: list[str] = []
        selected_scans = self._selected_scans()
        for idx, scan in enumerate(selected_scans):
            if not self._has_valid_normalized_output(scan):
                warnings.append(Path(scan.filename).name)
                continue
            payload = scan.candidate_resonators.get("sheet_resonances")
            assignments = payload.get("assignments") if isinstance(payload, dict) else {}
            if not isinstance(assignments, dict):
                assignments = {}
            freq = np.asarray(scan.freq, dtype=float)
            norm = scan.baseline_filter.get("normalized", {})
            amp, _phase = _read_polar_series(
                norm,
                amplitude_key="norm_amp",
                phase_key="norm_phase_deg_unwrapped",
            )
            amp = np.asarray(amp, dtype=float)
            if freq.shape != amp.shape or freq.size == 0:
                warnings.append(Path(scan.filename).name)
                continue
            order = np.argsort(freq)
            freq = freq[order]
            amp = amp[order]
            resonators: list[dict] = []
            for resonator_number, record in assignments.items():
                if not isinstance(record, dict):
                    continue
                try:
                    target_hz = float(record.get("frequency_hz"))
                except Exception:
                    continue
                if not (float(freq[0]) <= target_hz <= float(freq[-1])):
                    continue
                resonators.append(
                    {
                        "resonator_number": str(resonator_number).strip(),
                        "target_hz": target_hz,
                    }
                )
            rows.append(
                {
                    "scan": scan,
                    "scan_key": self._scan_key(scan),
                    "plot_group": scan.plot_group,
                    "freq": freq,
                    "amp": amp,
                    "scan_index": idx,
                    "resonators": sorted(resonators, key=lambda item: item["target_hz"]),
                }
            )
        return rows, warnings

    @staticmethod
    def _attached_resonance_editor_offset_map(
        rows: list[dict],
        spacing: float,
    ) -> tuple[dict[str, float], list[tuple[float, str]]]:
        level_keys: list[tuple[str, int]] = []
        labels_by_level: dict[tuple[str, int], list[str]] = {}
        for row in rows:
            plot_group = row.get("plot_group")
            if plot_group is None:
                level_key = ("scan", int(row.get("scan_index", 0)))
            else:
                level_key = ("group", int(plot_group))
            if level_key not in labels_by_level:
                level_keys.append(level_key)
                labels_by_level[level_key] = []
            file_timestamp = str(getattr(row["scan"], "file_timestamp", "")).strip()
            date_label = file_timestamp.split("T", 1)[0] if file_timestamp else "unknown date"
            labels_by_level[level_key].append(date_label)

        offset_by_scan_key: dict[str, float] = {}
        tick_info: list[tuple[float, str]] = []
        nlevels = len(level_keys)
        for level_pos, level_key in enumerate(level_keys):
            offset = float((nlevels - 1 - level_pos) * spacing)
            tick_y = offset + 1.0
            label_names = labels_by_level[level_key]
            label = label_names[0] if label_names else "unknown date"
            tick_info.append((tick_y, label))
            for row in rows:
                plot_group = row.get("plot_group")
                row_level_key = ("group", int(plot_group)) if plot_group is not None else ("scan", int(row.get("scan_index", 0)))
                if row_level_key == level_key:
                    offset_by_scan_key[str(row["scan_key"])] = offset
        return offset_by_scan_key, tick_info

    @staticmethod
    def _attached_resonance_editor_trace_colors() -> list[str]:
        return [
            "#8c564b",
            "#9467bd",
            "#ff7f0e",
            "#17becf",
            "#bcbd22",
            "#7f7f7f",
            "#e377c2",
            "#1f77b4",
            "#d62728",
            "#2ca02c",
            "#aec7e8",
            "#ffbb78",
        ]

    def _render_attached_resonance_editor(self) -> None:
        if self.attached_res_edit_figure is None or self.attached_res_edit_canvas is None:
            return
        rows, warnings = self._selected_scans_for_attached_resonance_editor()
        warning_tuple = tuple(sorted(set(warnings)))
        if warning_tuple:
            if self._attached_res_edit_missing_normalized_warned != warning_tuple:
                self._attached_res_edit_missing_normalized_warned = warning_tuple
                detail_lines = list(warning_tuple[:12])
                if len(warning_tuple) > 12:
                    detail_lines.append(f"... and {len(warning_tuple) - 12} more")
                messagebox.showwarning(
                    "Missing normalized data",
                    "The following selected scans do not have normalized data and will not be shown:\n\n"
                    + "\n".join(detail_lines),
                    parent=self.attached_res_edit_window,
                )
        else:
            self._attached_res_edit_missing_normalized_warned = None
        if not rows:
            message = "No selected scans with normalized data."
            if warnings:
                message += " Missing normalized data for: " + ", ".join(warnings[:6])
                if len(warnings) > 6:
                    message += f", ... (+{len(warnings) - 6} more)"
            if self.attached_res_edit_status_var is not None:
                self.attached_res_edit_status_var.set(message)
            self.attached_res_edit_figure.clear()
            ax = self.attached_res_edit_figure.add_subplot(111)
            ax.text(0.5, 0.5, message, ha="center", va="center", transform=ax.transAxes)
            ax.set_axis_off()
            self._attached_resonance_editor_update_save_button()
            self.attached_res_edit_canvas.draw_idle()
            return

        prior_xlim = None
        if self.attached_res_edit_ax is not None:
            try:
                prior_xlim = self.attached_res_edit_ax.get_xlim()
            except Exception:
                prior_xlim = None

        self.attached_res_edit_figure.clear()
        ax = self.attached_res_edit_figure.add_subplot(111)
        self.attached_res_edit_ax = ax
        self._attached_res_edit_points = []
        self._attached_res_edit_rows_cache = rows
        offset_by_scan_key, tick_info = self._attached_resonance_editor_offset_map(
            rows,
            self._attached_resonance_editor_curve_spacing(),
        )
        self._attached_res_edit_offset_by_scan_key = dict(offset_by_scan_key)

        freq_min = min(float(row["freq"][0]) for row in rows) / 1.0e9
        freq_max = max(float(row["freq"][-1]) for row in rows) / 1.0e9
        x_pad = max((freq_max - freq_min) * 0.01, 1e-6)

        resonator_tracks: dict[str, list[tuple[float, float]]] = {}
        resonator_markers: list[dict] = []
        y_text_offset = 0.18
        trace_colors = self._attached_resonance_editor_trace_colors()
        for row in rows:
            scan = row["scan"]
            offset = float(offset_by_scan_key[str(row["scan_key"])])
            freq_ghz = np.asarray(row["freq"], dtype=float) / 1.0e9
            amp_display = self._attached_resonance_editor_display_amp(row["amp"])
            y = amp_display + offset
            trace_color = trace_colors[int(row.get("scan_index", 0)) % len(trace_colors)]
            ax.plot(freq_ghz, y, linewidth=1.0, color=trace_color, alpha=0.9, zorder=1)

            for resonator in row["resonators"]:
                target_hz = float(resonator["target_hz"])
                target_ghz = target_hz / 1.0e9
                y_pt = self._interpolate_y(row["freq"], amp_display, target_hz) + offset
                scan_key = str(row["scan_key"])
                resonator_number = str(resonator["resonator_number"])
                point = {
                    "scan": scan,
                    "scan_key": scan_key,
                    "resonator_number": resonator_number,
                    "x_ghz": target_ghz,
                    "y": y_pt,
                    "freq_hz": target_hz,
                }
                self._attached_res_edit_points.append(point)
                resonator_markers.append(point)
                resonator_tracks.setdefault(resonator_number, []).append((target_ghz, y_pt))

        for resonator_number, points in resonator_tracks.items():
            if len(points) < 2:
                continue
            points = sorted(points, key=lambda item: item[1], reverse=True)
            ax.plot(
                [pt[0] for pt in points],
                [pt[1] for pt in points],
                linestyle=":",
                linewidth=1.0,
                color="tab:red",
                alpha=0.9,
                zorder=2,
            )

        for point in resonator_markers:
            is_selected = self._attached_res_edit_selected == (point["scan_key"], point["resonator_number"])
            ax.plot(
                [point["x_ghz"]],
                [point["y"]],
                linestyle="none",
                marker="o",
                markersize=(9 if is_selected else 6),
                markerfacecolor="none",
                markeredgecolor=("black" if is_selected else "tab:red"),
                markeredgewidth=1.5,
                zorder=4,
            )
            ax.text(
                point["x_ghz"],
                point["y"] - y_text_offset,
                point["resonator_number"],
                ha="center",
                va="top",
                fontsize=8,
                color=("black" if is_selected else "tab:red"),
                zorder=5,
            )

        ax.set_xlabel("Frequency (GHz)")
        ax.set_ylabel("Normalized |S21| + vertical offset")
        ax.grid(True, alpha=0.3)
        ax.set_yticks([item[0] for item in tick_info])
        ax.set_yticklabels([item[1] for item in tick_info], fontsize=8)

        y_low = min(
            float(np.min(self._attached_resonance_editor_display_amp(row["amp"])))
            + float(offset_by_scan_key[str(row["scan_key"])])
            for row in rows
        )
        y_high = max(
            float(np.max(self._attached_resonance_editor_display_amp(row["amp"])))
            + float(offset_by_scan_key[str(row["scan_key"])])
            for row in rows
        )
        ax.set_ylim(y_low - 0.2, y_high + 0.2)
        self._attached_res_edit_default_xlim = (freq_min - 0.5 * x_pad, freq_max + 2.0 * x_pad)
        if prior_xlim is not None:
            ax.set_xlim(prior_xlim)
        else:
            ax.set_xlim(self._attached_res_edit_default_xlim)

        if self.attached_res_edit_status_var is not None:
            status = (
                "Left-click a resonator to select. Double-click on the same scan to move it. "
                "Use 'Add Resonator' then click a scan to add one."
            )
            if warnings:
                status += " Missing normalized data for: " + ", ".join(warnings[:6])
                if len(warnings) > 6:
                    status += f", ... (+{len(warnings) - 6} more)"
            if self._attached_res_edit_pending_add:
                resonator_number = self._attached_resonance_editor_working_number()
                status = f"Add mode: click near a scan trace to add resonator {resonator_number}."
                if warnings:
                    status += " Missing normalized data for: " + ", ".join(warnings[:4])
                    if len(warnings) > 4:
                        status += f", ... (+{len(warnings) - 4} more)"
            self.attached_res_edit_status_var.set(status)
        self._attached_resonance_editor_update_add_button()
        self._attached_resonance_editor_update_undo_button()
        self._attached_resonance_editor_update_save_button()

        self.attached_res_edit_figure.subplots_adjust(left=0.12, right=0.985, bottom=0.09, top=0.96)
        self.attached_res_edit_canvas.draw_idle()

    def _attached_resonance_editor_reset_view(self) -> None:
        if self.attached_res_edit_ax is None or self._attached_res_edit_default_xlim is None:
            return
        self.attached_res_edit_ax.set_xlim(self._attached_res_edit_default_xlim)
        if self.attached_res_edit_canvas is not None:
            self.attached_res_edit_canvas.draw_idle()

    def _attached_resonance_editor_curve_spacing(self) -> float:
        if self.attached_res_edit_spacing_var is None:
            return 1.5
        try:
            value = float(self.attached_res_edit_spacing_var.get())
        except Exception:
            value = 1.5
        value = min(max(value, 0.0), 3.0)
        if abs(value - float(self.attached_res_edit_spacing_var.get())) > 1e-12:
            self.attached_res_edit_spacing_var.set(value)
        return value

    def _attached_resonance_editor_truncate_enabled(self) -> bool:
        return bool(self.attached_res_edit_truncate_var.get()) if self.attached_res_edit_truncate_var is not None else True

    def _attached_resonance_editor_search_window_hz(self) -> float:
        if self.attached_res_edit_search_window_khz_var is None:
            return 3.0e5
        try:
            value_khz = float(self.attached_res_edit_search_window_khz_var.get())
        except Exception:
            value_khz = 300.0
        value_khz = min(max(value_khz, 25.0), 2000.0)
        if abs(value_khz - float(self.attached_res_edit_search_window_khz_var.get())) > 1e-12:
            self.attached_res_edit_search_window_khz_var.set(value_khz)
        return value_khz * 1.0e3

    def _attached_resonance_editor_truncate_threshold(self) -> float:
        if self.attached_res_edit_truncate_threshold_var is None:
            return 1.5
        try:
            value = float(self.attached_res_edit_truncate_threshold_var.get())
        except Exception:
            value = 1.5
        value = min(max(value, 1.0), 2.0)
        if abs(value - float(self.attached_res_edit_truncate_threshold_var.get())) > 1e-12:
            self.attached_res_edit_truncate_threshold_var.set(value)
        return value

    def _attached_resonance_editor_display_amp(self, amp: Sequence[float]) -> np.ndarray:
        amp_arr = np.asarray(amp, dtype=float)
        if not self._attached_resonance_editor_truncate_enabled():
            return amp_arr
        return np.minimum(amp_arr, self._attached_resonance_editor_truncate_threshold())

    def _attached_resonance_editor_on_spacing_release(self, _event) -> None:
        if self.attached_res_edit_window is None or not self.attached_res_edit_window.winfo_exists():
            return
        prior_xlim = None
        if self.attached_res_edit_ax is not None:
            try:
                prior_xlim = self.attached_res_edit_ax.get_xlim()
            except Exception:
                prior_xlim = None
        self._render_attached_resonance_editor()
        if prior_xlim is not None and self.attached_res_edit_ax is not None:
            self.attached_res_edit_ax.set_xlim(prior_xlim)
            if self.attached_res_edit_canvas is not None:
                self.attached_res_edit_canvas.draw_idle()

    def _attached_resonance_editor_on_truncate_toggle(self) -> None:
        self._render_attached_resonance_editor()

    def _attached_resonance_editor_on_truncate_release(self, _event) -> None:
        self._render_attached_resonance_editor()

    def _attached_resonance_editor_toggle_add(self) -> None:
        self._attached_res_edit_pending_add = not self._attached_res_edit_pending_add
        self._attached_res_edit_selected = None
        if self.attached_res_edit_status_var is not None:
            if self._attached_res_edit_pending_add:
                resonator_number = self._attached_resonance_editor_working_number()
                self.attached_res_edit_status_var.set(
                    f"Add mode: click near a selected normalized scan to add resonator {resonator_number}."
                )
            else:
                self.attached_res_edit_status_var.set("Add mode deactivated.")
        self._render_attached_resonance_editor()

    def _attached_resonance_editor_update_add_button(self) -> None:
        if self.attached_res_edit_add_button is None:
            return
        if self._attached_res_edit_pending_add:
            self.attached_res_edit_add_button.configure(
                relief="sunken",
                bg="light green",
                activebackground="light green",
            )
            return
        self.attached_res_edit_add_button.configure(
            relief="raised",
            bg=self._default_button_bg,
            activebackground=self._default_button_activebg,
        )

    def _attached_resonance_editor_update_save_button(self) -> None:
        if self.attached_res_edit_save_button is None:
            return
        if self._attached_res_edit_changed:
            self.attached_res_edit_save_button.configure(
                bg="pink",
                activebackground="pink",
            )
            return
        self.attached_res_edit_save_button.configure(
            bg=self._default_button_bg,
            activebackground=self._default_button_activebg,
        )

    def _attached_resonance_editor_update_undo_button(self) -> None:
        if self.attached_res_edit_undo_button is None:
            return
        if self._attached_res_edit_undo_stack:
            self.attached_res_edit_undo_button.configure(state="normal")
        else:
            self.attached_res_edit_undo_button.configure(state="disabled")

    def _attached_resonance_editor_undo(self) -> None:
        if not self._attached_res_edit_undo_stack:
            if self.attached_res_edit_status_var is not None:
                self.attached_res_edit_status_var.set("Nothing to undo.")
            self._attached_resonance_editor_update_undo_button()
            return
        snapshot = self._attached_res_edit_undo_stack.pop()
        self._attached_resonance_editor_apply_snapshot(snapshot)
        self._attached_res_edit_changed = bool(self._attached_res_edit_undo_stack)
        self._attached_resonance_editor_update_undo_button()
        self._attached_resonance_editor_update_save_button()
        if self.attached_res_edit_status_var is not None:
            self.attached_res_edit_status_var.set("Undid last resonator marker edit.")
        self._render_attached_resonance_editor()

    def _attached_resonance_editor_renumber_low_to_high(self) -> None:
        selected_scans = self._selected_scans()
        resonator_values: dict[str, list[float]] = {}
        for scan in selected_scans:
            payload = scan.candidate_resonators.get("sheet_resonances")
            assignments = payload.get("assignments") if isinstance(payload, dict) else {}
            if not isinstance(assignments, dict):
                continue
            for resonator_number, record in assignments.items():
                if not isinstance(record, dict):
                    continue
                try:
                    freq_hz = float(record.get("frequency_hz"))
                except Exception:
                    continue
                if not np.isfinite(freq_hz):
                    continue
                label = str(resonator_number).strip()
                if not label:
                    continue
                resonator_values.setdefault(label, []).append(freq_hz)

        ordered_labels = sorted(
            (label for label, values in resonator_values.items() if values),
            key=lambda label: (
                float(np.mean(np.asarray(resonator_values[label], dtype=float))),
                self._resonator_sort_key(label),
            ),
        )
        if not ordered_labels:
            if self.attached_res_edit_status_var is not None:
                self.attached_res_edit_status_var.set("No resonator markers are available to renumber.")
            return

        renumber_map = {
            old_label: str(new_idx)
            for new_idx, old_label in enumerate(ordered_labels, start=1)
        }
        if all(old_label == new_label for old_label, new_label in renumber_map.items()):
            if self.attached_res_edit_status_var is not None:
                self.attached_res_edit_status_var.set("Resonator markers are already numbered low to high.")
            return

        self._attached_resonance_editor_push_undo_snapshot()
        for scan in selected_scans:
            payload = scan.candidate_resonators.get("sheet_resonances")
            assignments = payload.get("assignments") if isinstance(payload, dict) else {}
            if not isinstance(assignments, dict) or not assignments:
                continue
            new_assignments: dict[str, dict] = {}
            for resonator_number, record in assignments.items():
                if not isinstance(record, dict):
                    continue
                new_label = renumber_map.get(str(resonator_number).strip())
                if new_label is None:
                    continue
                new_assignments[new_label] = record
            payload["assignments"] = new_assignments

        if self._attached_res_edit_selected is not None:
            selected_scan_key, selected_number = self._attached_res_edit_selected
            mapped = renumber_map.get(str(selected_number).strip())
            self._attached_res_edit_selected = (selected_scan_key, mapped) if mapped is not None else None
        self._attached_res_edit_changed = True
        self._attached_resonance_editor_reset_working_number()
        self._attached_resonance_editor_update_save_button()
        self._attached_resonance_editor_update_undo_button()
        if self.attached_res_edit_status_var is not None:
            self.attached_res_edit_status_var.set(
                f"Renumbered {len(ordered_labels)} resonator marker(s) from low to high mean frequency."
            )
        self._render_attached_resonance_editor()

    def _attached_resonance_editor_working_number(self) -> str:
        if self.attached_res_edit_working_number_var is None:
            return "1"
        raw_value = str(self.attached_res_edit_working_number_var.get()).strip()
        try:
            number = max(1, int(raw_value))
        except Exception:
            number = 1
        if raw_value != str(number):
            self.attached_res_edit_working_number_var.set(str(number))
        return str(number)

    def _attached_resonance_editor_next_number(self) -> str:
        max_number = 0
        for scan in self._selected_scans():
            payload = scan.candidate_resonators.get("sheet_resonances")
            assignments = payload.get("assignments") if isinstance(payload, dict) else {}
            if not isinstance(assignments, dict):
                continue
            for resonator_number in assignments.keys():
                try:
                    parsed = int(str(resonator_number).strip())
                except Exception:
                    continue
                if parsed >= 1:
                    max_number = max(max_number, parsed)
        return str(max_number + 1)

    def _attached_resonance_editor_reset_working_number(self) -> None:
        if self.attached_res_edit_working_number_var is None:
            return
        self.attached_res_edit_working_number_var.set(self._attached_resonance_editor_next_number())

    def _attached_resonance_editor_delete_selected(self) -> None:
        if self._attached_res_edit_selected is None:
            if self.attached_res_edit_status_var is not None:
                self.attached_res_edit_status_var.set("No resonator selected to delete.")
            return
        scan_key, resonator_number = self._attached_res_edit_selected
        for scan in self.dataset.vna_scans:
            if self._scan_key(scan) != scan_key:
                continue
            payload = scan.candidate_resonators.get("sheet_resonances")
            if not isinstance(payload, dict):
                break
            assignments = payload.get("assignments")
            if not isinstance(assignments, dict):
                break
            self._attached_resonance_editor_push_undo_snapshot()
            assignments.pop(resonator_number, None)
            if not assignments:
                scan.candidate_resonators.pop("sheet_resonances", None)
            self.dataset.processing_history.append(
                _make_event(
                    "delete_attached_resonator",
                    {"scan": scan.filename, "resonator_number": resonator_number},
                )
            )
            self._attached_res_edit_changed = True
            break
        self._attached_res_edit_selected = None
        self._attached_resonance_editor_reset_working_number()
        self._render_attached_resonance_editor()

    def _attached_resonance_editor_clear_selected_scan_markers(self) -> None:
        scans = self._selected_scans()
        if not scans:
            if self.attached_res_edit_status_var is not None:
                self.attached_res_edit_status_var.set("No selected scans are available.")
            return

        scans_with_markers = []
        for scan in scans:
            payload = scan.candidate_resonators.get("sheet_resonances")
            assignments = payload.get("assignments") if isinstance(payload, dict) else {}
            if isinstance(assignments, dict) and assignments:
                scans_with_markers.append(scan)

        if not scans_with_markers:
            if self.attached_res_edit_status_var is not None:
                self.attached_res_edit_status_var.set("Selected scans do not have attached resonator markers.")
            return

        names = [Path(str(scan.filename)).name for scan in scans_with_markers]
        ok = messagebox.askyesno(
            "Clear Resonator Markers",
            f"Clear attached resonator markers from {len(scans_with_markers)} selected scan(s)?\n\n"
            + "\n".join(names[:10])
            + ("\n..." if len(names) > 10 else ""),
            parent=self.attached_res_edit_window,
        )
        if not ok:
            return

        self._attached_resonance_editor_push_undo_snapshot()
        cleared = 0
        for scan in scans_with_markers:
            scan.candidate_resonators.pop("sheet_resonances", None)
            scan.processing_history.append(
                _make_event(
                    "clear_attached_resonator_markers",
                    {"scan": scan.filename},
                )
            )
            cleared += 1
        self.dataset.processing_history.append(
            _make_event(
                "clear_attached_resonator_markers_selected_scans",
                {
                    "selected_count": len(scans),
                    "cleared_count": cleared,
                    "filenames": [scan.filename for scan in scans_with_markers],
                },
            )
        )
        self._attached_res_edit_selected = None
        self._attached_res_edit_changed = True
        self._attached_resonance_editor_reset_working_number()
        if self.attached_res_edit_status_var is not None:
            self.attached_res_edit_status_var.set(
                f"Cleared attached resonator markers from {cleared} selected scan(s)."
            )
        self._render_attached_resonance_editor()

    def _attached_resonance_editor_on_click(self, event) -> None:
        if self.attached_res_edit_ax is None or event.inaxes != self.attached_res_edit_ax:
            return
        if event.xdata is None or event.ydata is None:
            return
        if not self._attached_resonance_editor_click_is_within_plot(float(event.xdata), float(event.ydata)):
            if self.attached_res_edit_status_var is not None:
                self.attached_res_edit_status_var.set("Click inside the visible plot window to edit resonators.")
            return
        if self.attached_res_edit_toolbar is not None:
            mode = getattr(self.attached_res_edit_toolbar, "mode", "")
            if str(mode).strip():
                if self.attached_res_edit_status_var is not None:
                    self.attached_res_edit_status_var.set(
                        f"Navigation mode active ({str(mode).strip()}). Finish pan/zoom to resume editing."
                    )
                return

        nearest = self._attached_resonance_editor_find_nearest_point(float(event.xdata), float(event.ydata))
        if self._attached_res_edit_pending_add:
            self._attached_resonance_editor_add_at_click(float(event.xdata), float(event.ydata))
            return

        if event.dblclick and self._attached_res_edit_selected is not None:
            self._attached_resonance_editor_move_selected(float(event.xdata), float(event.ydata))
            return

        if nearest is None:
            self._attached_res_edit_selected = None
            if self.attached_res_edit_status_var is not None:
                self.attached_res_edit_status_var.set("No resonator selected.")
            self._render_attached_resonance_editor()
            return

        self._attached_res_edit_selected = (nearest["scan_key"], nearest["resonator_number"])
        if self.attached_res_edit_status_var is not None:
            self.attached_res_edit_status_var.set(
                f"Selected resonator {nearest['resonator_number']} on {Path(nearest['scan'].filename).name}."
            )
        self._render_attached_resonance_editor()

    def _attached_resonance_editor_find_nearest_point(self, x_ghz: float, y_val: float) -> Optional[dict]:
        if not self._attached_res_edit_points or self.attached_res_edit_ax is None:
            return None
        x0, x1 = self.attached_res_edit_ax.get_xlim()
        y0, y1 = self.attached_res_edit_ax.get_ylim()
        x_span = max(abs(x1 - x0), 1e-9)
        y_span = max(abs(y1 - y0), 1e-9)
        best = None
        best_metric = None
        for point in self._attached_res_edit_points:
            metric = ((point["x_ghz"] - x_ghz) / x_span) ** 2 + ((point["y"] - y_val) / y_span) ** 2
            if best_metric is None or metric < best_metric:
                best_metric = metric
                best = point
        if best_metric is None or best_metric > 0.0025:
            return None
        return best

    def _attached_resonance_editor_click_is_within_plot(self, x_ghz: float, y_val: float) -> bool:
        if self.attached_res_edit_ax is None:
            return False
        try:
            x0, x1 = self.attached_res_edit_ax.get_xlim()
            y0, y1 = self.attached_res_edit_ax.get_ylim()
        except Exception:
            return False
        lo_x, hi_x = (float(x0), float(x1)) if float(x0) <= float(x1) else (float(x1), float(x0))
        lo_y, hi_y = (float(y0), float(y1)) if float(y0) <= float(y1) else (float(y1), float(y0))
        return lo_x <= float(x_ghz) <= hi_x and lo_y <= float(y_val) <= hi_y

    def _attached_resonance_editor_signal_success(self) -> None:
        widget = self.attached_res_edit_window if self.attached_res_edit_window is not None else self.root
        if winsound is not None:
            try:
                winsound.PlaySound(
                    "SystemDefault",
                    winsound.SND_ALIAS | winsound.SND_ASYNC | winsound.SND_NODEFAULT,
                )
                return
            except Exception:
                pass
        try:
            widget.bell()
        except Exception:
            pass

    def _attached_resonance_editor_row_for_add_click(
        self,
        rows: list[dict],
        offset_by_scan_key: dict[str, float],
        click_hz: float,
        y_val: float,
        *,
        window_hz: float,
        visible_range_hz: Optional[tuple[float, float]],
    ) -> Optional[tuple[dict, float]]:
        group_offsets_in_order: list[float] = []
        seen_offsets: set[float] = set()
        for row in rows:
            offset = float(offset_by_scan_key.get(str(row["scan_key"]), 0.0))
            if offset in seen_offsets:
                continue
            seen_offsets.add(offset)
            group_offsets_in_order.append(offset)

        chosen_offset = None
        for offset in group_offsets_in_order:
            if offset <= float(y_val) <= offset + 1.0:
                chosen_offset = offset
                break
        if chosen_offset is None:
            return None

        best_row = None
        best_target_hz = None
        best_metric = None
        for row in rows:
            offset = float(offset_by_scan_key.get(str(row["scan_key"]), 0.0))
            if abs(offset - chosen_offset) > 1e-12:
                continue
            freq_arr = np.asarray(row["freq"], dtype=float)
            if freq_arr.size == 0:
                continue
            if not (float(freq_arr[0]) <= float(click_hz) <= float(freq_arr[-1])):
                continue
            target_hz = self._attached_resonance_editor_minimum_near_click(
                row["freq"],
                row["amp"],
                click_hz,
                window_hz=window_hz,
                visible_range_hz=visible_range_hz,
            )
            if target_hz is None:
                continue
            metric = abs(float(target_hz) - float(click_hz))
            if best_metric is None or metric < best_metric:
                best_metric = metric
                best_row = row
                best_target_hz = float(target_hz)
        if best_row is None or best_target_hz is None:
            return None
        return best_row, best_target_hz

    def _show_attached_resonance_minimum_search_diagnostic(
        self,
        *,
        row: dict,
        click_hz: float,
        window_hz: float,
        visible_range_hz: Optional[tuple[float, float]],
        detail: str,
    ) -> None:
        if self.attached_res_edit_window is None or not self.attached_res_edit_window.winfo_exists():
            return
        freq_arr = np.asarray(row["freq"], dtype=float)
        amp_display = self._attached_resonance_editor_display_amp(row["amp"])
        if freq_arr.size == 0 or amp_display.size == 0:
            messagebox.showwarning("Add resonator failed", detail, parent=self.attached_res_edit_window)
            return

        if visible_range_hz is not None:
            lo_visible, hi_visible = visible_range_hz
        else:
            lo_visible, hi_visible = float(freq_arr[0]), float(freq_arr[-1])
        display_mask = (freq_arr >= float(lo_visible)) & (freq_arr <= float(hi_visible))
        if not np.any(display_mask):
            display_mask = np.ones(freq_arr.shape, dtype=bool)
            lo_visible, hi_visible = float(freq_arr[0]), float(freq_arr[-1])

        lo_search = max(float(lo_visible), float(click_hz) - float(window_hz))
        hi_search = min(float(hi_visible), float(click_hz) + float(window_hz))
        search_mask = (freq_arr >= lo_search) & (freq_arr <= hi_search)

        window = tk.Toplevel(self.attached_res_edit_window)
        window.title("Add Resonator Failed")
        window.geometry("920x560")
        window.transient(self.attached_res_edit_window)

        controls = tk.Frame(window, padx=8, pady=8)
        controls.pack(side="top", fill="x")
        tk.Label(
            controls,
            text=detail,
            anchor="w",
            justify="left",
            wraplength=880,
        ).pack(side="left", fill="x", expand=True)
        tk.Button(controls, text="Close", width=10, command=window.destroy).pack(side="right", padx=(8, 0))

        fig = Figure(figsize=(9.2, 4.8))
        ax = fig.add_subplot(111)
        canvas = FigureCanvasTkAgg(fig, master=window)
        canvas.get_tk_widget().pack(fill="both", expand=True)

        ax.plot(freq_arr[display_mask] / 1.0e9, amp_display[display_mask], color="tab:blue", linewidth=1.2)
        ax.axvspan(lo_search / 1.0e9, hi_search / 1.0e9, color="gold", alpha=0.25, label="searched window")
        ax.axvline(click_hz / 1.0e9, color="crimson", linestyle="--", linewidth=1.5, label="mouse click")
        if np.any(search_mask):
            ax.plot(
                freq_arr[search_mask] / 1.0e9,
                amp_display[search_mask],
                color="darkorange",
                linewidth=2.0,
            )
        ax.set_xlim(float(lo_visible) / 1.0e9, float(hi_visible) / 1.0e9)
        ax.set_xlabel("Frequency (GHz)")
        ax.set_ylabel("Displayed |S21|")
        ax.set_title(Path(row["scan"].filename).name)
        ax.grid(True, alpha=0.3)
        ax.legend(loc="best")
        fig.tight_layout()
        canvas.draw_idle()

    def _attached_resonance_editor_add_at_click(self, x_ghz: float, y_val: float) -> None:
        rows = self._attached_res_edit_rows_cache
        offset_by_scan_key = self._attached_res_edit_offset_by_scan_key
        if not rows or not offset_by_scan_key:
            rows, _warnings = self._selected_scans_for_attached_resonance_editor()
            offset_by_scan_key, _tick_info = self._attached_resonance_editor_offset_map(
                rows,
                self._attached_resonance_editor_curve_spacing(),
            )
        if not rows:
            return
        visible_range_hz = self._attached_resonance_editor_visible_range_hz()
        click_hz = x_ghz * 1.0e9
        target = self._attached_resonance_editor_row_for_add_click(
            rows,
            offset_by_scan_key,
            click_hz,
            y_val,
            window_hz=self._attached_resonance_editor_search_window_hz(),
            visible_range_hz=visible_range_hz,
        )
        if target is None:
            if self.attached_res_edit_status_var is not None:
                self.attached_res_edit_status_var.set(
                    "Click inside a group's 0-1 band at a frequency where a visible local minimum can be found."
                )
            return
        best_row, target_hz = target
        if visible_range_hz is not None:
            lo_hz, hi_hz = visible_range_hz
            if not (lo_hz <= click_hz <= hi_hz):
                if self.attached_res_edit_status_var is not None:
                    self.attached_res_edit_status_var.set("Click inside the visible plot window to add a resonator.")
                return
        resonator_number = self._attached_resonance_editor_working_number()
        scan = best_row["scan"]
        if visible_range_hz is not None:
            lo_hz, hi_hz = visible_range_hz
            if not (lo_hz <= target_hz <= hi_hz):
                detail = (
                    f"Attached-resonator add aborted because snapped target "
                    f"{target_hz / 1.0e9:.9g} GHz is outside the visible range "
                    f"{lo_hz / 1.0e9:.9g} to {hi_hz / 1.0e9:.9g} GHz."
                )
                if self.attached_res_edit_status_var is not None:
                    self.attached_res_edit_status_var.set(detail)
                messagebox.showwarning("Add resonator failed", detail, parent=self.attached_res_edit_window)
                self._log(
                    f"Attach resonator warning: click at {x_ghz:.9g} GHz on {Path(scan.filename).name} "
                    f"snapped to off-screen target {target_hz / 1.0e9:.9g} GHz."
                )
                return
        self._attached_resonance_editor_signal_success()
        payload = self._sheet_resonance_attachment(scan)
        assignments = payload["assignments"]
        self._attached_resonance_editor_push_undo_snapshot()
        assignments[resonator_number] = {
            "frequency_hz": target_hz,
            "sheet_path": "",
            "sheet_name": "",
            "row": 0,
            "column": 0,
            "identifier": self._sheet_identifier_for_scan(scan),
        }
        self.dataset.processing_history.append(
            _make_event(
                "add_attached_resonator",
                {"scan": scan.filename, "resonator_number": resonator_number, "frequency_hz": target_hz},
            )
        )
        self._attached_res_edit_changed = True
        self._attached_res_edit_selected = (self._scan_key(scan), resonator_number)
        self._render_attached_resonance_editor()

    def _attached_resonance_editor_visible_range_hz(self) -> Optional[tuple[float, float]]:
        if self.attached_res_edit_ax is None:
            return None
        try:
            x0, x1 = self.attached_res_edit_ax.get_xlim()
        except Exception:
            return None
        lo_ghz, hi_ghz = (float(x0), float(x1)) if float(x0) <= float(x1) else (float(x1), float(x0))
        return (lo_ghz * 1.0e9, hi_ghz * 1.0e9)

    def _attached_resonance_editor_minimum_near_click(
        self,
        freq_hz: Sequence[float],
        amp: Sequence[float],
        click_hz: float,
        window_hz: float = 3.0e5,
        visible_range_hz: Optional[tuple[float, float]] = None,
    ) -> Optional[float]:
        freq_arr = np.asarray(freq_hz, dtype=float)
        amp_arr = np.asarray(amp, dtype=float)
        if freq_arr.size == 0 or amp_arr.size == 0:
            return None
        window_mask = np.abs(freq_arr - float(click_hz)) <= float(window_hz)
        if visible_range_hz is not None:
            lo_hz, hi_hz = visible_range_hz
            visible_mask = (freq_arr >= float(lo_hz)) & (freq_arr <= float(hi_hz))
            window_mask = window_mask & visible_mask
        if np.any(window_mask):
            candidate_indices = np.flatnonzero(window_mask)
            best_local_idx = int(candidate_indices[int(np.argmin(amp_arr[window_mask]))])
            return float(freq_arr[best_local_idx])
        return None


    def _attached_resonance_editor_move_selected(self, x_ghz: float, y_val: float) -> None:
        if self._attached_res_edit_selected is None:
            return
        scan_key, resonator_number = self._attached_res_edit_selected
        rows = self._attached_res_edit_rows_cache
        offset_by_scan_key = self._attached_res_edit_offset_by_scan_key
        if not rows or not offset_by_scan_key:
            rows, _warnings = self._selected_scans_for_attached_resonance_editor()
            offset_by_scan_key, _tick_info = self._attached_resonance_editor_offset_map(
                rows,
                self._attached_resonance_editor_curve_spacing(),
            )
        target_row = None
        for row in rows:
            if row["scan_key"] != scan_key:
                continue
            offset = float(offset_by_scan_key[str(row["scan_key"])])
            amp_display = self._attached_resonance_editor_display_amp(row["amp"])
            y_trace = self._interpolate_y(row["freq"], amp_display, x_ghz * 1.0e9) + offset
            if abs(y_trace - y_val) <= 0.35:
                target_row = row
                break
        if target_row is None:
            if self.attached_res_edit_status_var is not None:
                self.attached_res_edit_status_var.set("Double-click on the same scan trace to move the selected resonator.")
            return
        scan = target_row["scan"]
        payload = scan.candidate_resonators.get("sheet_resonances")
        if not isinstance(payload, dict):
            return
        assignments = payload.get("assignments")
        if not isinstance(assignments, dict) or resonator_number not in assignments:
            return
        self._attached_resonance_editor_push_undo_snapshot()
        nearest_idx = int(np.argmin(np.abs(np.asarray(target_row["freq"], dtype=float) - x_ghz * 1.0e9)))
        target_hz = float(target_row["freq"][nearest_idx])
        assignments[resonator_number]["frequency_hz"] = target_hz
        self.dataset.processing_history.append(
            _make_event(
                "move_attached_resonator",
                {"scan": scan.filename, "resonator_number": resonator_number, "frequency_hz": target_hz},
            )
        )
        self._attached_res_edit_changed = True
        self._render_attached_resonance_editor()

def main() -> None:
    root = tk.Tk()
    DataAnalysisGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
