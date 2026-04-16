from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Callable, Optional

import matplotlib.pyplot as plt
import numpy as np
import tkinter as tk
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.figure import Figure
from openpyxl import Workbook, load_workbook
from tkinter import filedialog, messagebox

from analysis_gui_support.analysis_io import _dataset_dir
from analysis_gui_support.analysis_models import VNAScan, _make_event, _read_polar_series

class ResonanceSheetIOMixin:
    def open_resonance_sheet_loader(self) -> None:
        path_text = filedialog.askopenfilename(
            parent=self.root,
            title="Select resonance spreadsheet",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")],
        )
        if not path_text:
            return
        sheet_path = Path(path_text)
        if not sheet_path.exists():
            messagebox.showwarning("Missing file", "Select a valid spreadsheet file first.", parent=self.root)
            return
        try:
            loaded_count = self._load_resonances_from_sheet(sheet_path)
        except Exception as exc:
            self._log(f"Load resonators from sheet failed: {exc}")
            messagebox.showerror("Load failed", str(exc), parent=self.root)
            return
        messagebox.showinfo(
            "Resonators loaded",
            f"Loaded {loaded_count} resonator assignment(s) from:\n{sheet_path}",
            parent=self.root,
        )


    def open_resonance_sheet_saver(self) -> None:
        path_text = filedialog.asksaveasfilename(
            parent=self.root,
            title="Save resonance spreadsheet",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if not path_text:
            return
        sheet_path = Path(path_text)
        if str(sheet_path).strip() == "":
            messagebox.showwarning("Missing file", "Select an output spreadsheet path first.", parent=self.root)
            return
        try:
            saved_count = self._save_resonances_to_sheet(sheet_path)
        except Exception as exc:
            self._log(f"Save resonators to sheet failed: {exc}")
            messagebox.showerror("Save failed", str(exc), parent=self.root)
            return
        messagebox.showinfo(
            "Resonators saved",
            f"Saved {saved_count} resonator assignment(s) to:\n{sheet_path}",
            parent=self.root,
        )


    def open_attached_resonance_plotter(self) -> None:
        dialog = tk.Toplevel(self.root)
        dialog.title("Plot Resonator Markers")
        dialog.geometry("760x150")
        dialog.transient(self.root)
        dialog.grab_set()
        status_var = tk.StringVar(value="Ready to generate PDF.")

        generate_button = None

        def run() -> None:
            if generate_button is not None:
                generate_button.configure(state="disabled")
            status_var.set("Preparing pages...")
            dialog.update_idletasks()

            def on_progress(completed: int, total: int) -> None:
                status_var.set(f"Generating page {completed} of {total}...")
                dialog.update_idletasks()

            try:
                saved = self._plot_attached_resonances(progress_callback=on_progress)
            except Exception as exc:
                if generate_button is not None:
                    generate_button.configure(state="normal")
                status_var.set("PDF generation failed.")
                self._log(f"Plot resonator markers failed: {exc}")
                messagebox.showerror("Plot failed", str(exc), parent=dialog)
                return
            if not saved:
                if generate_button is not None:
                    generate_button.configure(state="normal")
                status_var.set("No PDF was generated.")
                messagebox.showwarning("No plots saved", "No plot files were generated.", parent=dialog)
                return
            out_dir = saved[0].parent
            status_var.set("PDF generation complete.")
            self._log(f"Plotted resonator markers into {out_dir}")
            messagebox.showinfo(
                "Plots saved",
                f"Saved {len(saved)} plot file(s) to:\n{out_dir}",
                parent=dialog,
            )
            dialog.destroy()

        top = tk.Frame(dialog, padx=10, pady=10)
        top.pack(fill="both", expand=True)
        tk.Label(
            top,
            text="Generate a single PDF using the same grouped resonator-marker view as the editor: one full-span page plus overlapping 100 MHz zoom pages.",
            anchor="w",
            justify="left",
            wraplength=720,
        ).pack(anchor="w", pady=(0, 8))
        tk.Label(top, textvariable=status_var, anchor="w", justify="left").pack(anchor="w")

        btns = tk.Frame(top)
        btns.pack(fill="x", pady=(12, 0))
        tk.Button(btns, text="Cancel", width=12, command=dialog.destroy).pack(side="right")
        generate_button = tk.Button(btns, text="Generate PDF", width=14, command=run)
        generate_button.pack(side="right", padx=(0, 8))


    @staticmethod
    def _coerce_frequency_to_scan_hz(value: float, scan: VNAScan) -> Optional[float]:
        freq = np.asarray(scan.freq, dtype=float)
        if freq.size == 0 or not np.isfinite(value):
            return None
        lo = float(np.min(freq))
        hi = float(np.max(freq))
        candidates = [float(value), float(value) * 1e9, float(value) * 1e6, float(value) * 1e3]
        for cand in candidates:
            if lo <= cand <= hi:
                return cand
        return None


    def _find_scan_for_sheet_identifier(self, identifier: str, target_value: float) -> tuple[VNAScan, float]:
        text = identifier.strip()
        lower = text.lower()
        scans = list(self.dataset.vna_scans)

        if lower.startswith("group "):
            try:
                group_num = int(lower.split()[1])
            except Exception as exc:
                raise ValueError(f"Invalid group name: {identifier}") from exc
            grouped = [scan for scan in scans if scan.plot_group == group_num]
            if not grouped:
                raise ValueError(f"No scans found for {identifier}.")
            for scan in grouped:
                target_hz = self._coerce_frequency_to_scan_hz(target_value, scan)
                if target_hz is not None:
                    return scan, target_hz
            raise ValueError(f"No scan in {identifier} contains frequency value {target_value}.")

        for scan in scans:
            name = Path(scan.filename).name.lower()
            stem = Path(scan.filename).stem.lower()
            if lower == name or lower == stem:
                target_hz = self._coerce_frequency_to_scan_hz(target_value, scan)
                if target_hz is None:
                    raise ValueError(f"Frequency value {target_value} is outside {Path(scan.filename).name}.")
                return scan, target_hz
        raise ValueError(f"No scan found matching filename {identifier}.")


    @staticmethod
    def _resonance_plot_window_hz(scan: VNAScan, target_hz: float) -> tuple[float, float]:
        freq = np.sort(np.asarray(scan.freq, dtype=float))
        if freq.size == 0:
            return target_hz, target_hz
        span = float(freq[-1] - freq[0])
        diffs = np.diff(freq)
        diffs = diffs[np.isfinite(diffs) & (diffs > 0)]
        step = float(np.median(diffs)) if diffs.size else max(span / max(freq.size - 1, 1), 1.0)
        half_width = max(250.0 * step, 0.002 * span)
        half_width = min(half_width, 0.05 * span if span > 0 else half_width)
        lo = max(float(freq[0]), target_hz - half_width)
        hi = min(float(freq[-1]), target_hz + half_width)
        if hi <= lo:
            half_width = max(1000.0 * step, 1.0)
            lo = max(float(freq[0]), target_hz - half_width)
            hi = min(float(freq[-1]), target_hz + half_width)
        return lo, hi


    @staticmethod
    def _interpolate_y(freq_hz: np.ndarray, y: np.ndarray, target_hz: float) -> float:
        x = np.asarray(freq_hz, dtype=float)
        vals = np.asarray(y, dtype=float)
        order = np.argsort(x)
        x = x[order]
        vals = vals[order]
        return float(np.interp(target_hz, x, vals))


    @staticmethod
    def _sheet_resonator_label(value: object) -> str:
        if value is None:
            return ""
        try:
            numeric = float(value)
            if np.isfinite(numeric) and numeric.is_integer():
                return str(int(numeric))
        except Exception:
            pass
        return str(value).strip()


    @staticmethod
    def _sheet_resonance_attachment(scan: VNAScan) -> dict[str, object]:
        payload = scan.candidate_resonators.get("sheet_resonances")
        if not isinstance(payload, dict):
            payload = {"assignments": {}}
            scan.candidate_resonators["sheet_resonances"] = payload
        assignments = payload.get("assignments")
        if not isinstance(assignments, dict):
            payload["assignments"] = {}
        return payload


    @staticmethod
    def _resonator_sort_key(label: str) -> tuple[int, object]:
        text = str(label).strip()
        try:
            value = int(text)
        except Exception:
            return (1, text)
        return (0, value)


    @staticmethod
    def _sheet_identifier_for_scan(scan: VNAScan, record: Optional[dict] = None) -> str:
        if scan.plot_group is not None:
            return f"Group {int(scan.plot_group)}"
        if isinstance(record, dict):
            identifier = str(record.get("identifier") or "").strip()
            if identifier:
                return identifier
        return Path(scan.filename).name


    def _normalize_sheet_resonance_identifiers(self) -> int:
        updated_count = 0
        for scan in self.dataset.vna_scans:
            payload = scan.candidate_resonators.get("sheet_resonances")
            if not isinstance(payload, dict):
                continue
            assignments = payload.get("assignments")
            if not isinstance(assignments, dict):
                continue
            for record in assignments.values():
                if not isinstance(record, dict):
                    continue
                expected = self._sheet_identifier_for_scan(scan, record)
                current = str(record.get("identifier") or "").strip()
                if current != expected:
                    record["identifier"] = expected
                    updated_count += 1
        return updated_count


    def _save_resonances_to_sheet(self, sheet_path: Path) -> int:
        normalized_count = self._normalize_sheet_resonance_identifiers()
        if normalized_count > 0:
            self.dataset.processing_history.append(
                _make_event(
                    "normalize_sheet_resonance_identifiers",
                    {"updated_count": int(normalized_count)},
                )
            )
        row_map: dict[str, dict[str, float]] = {}
        row_order: list[str] = []
        resonator_numbers: set[str] = set()
        assignment_count = 0

        for scan in self.dataset.vna_scans:
            payload = scan.candidate_resonators.get("sheet_resonances")
            if not isinstance(payload, dict):
                continue
            assignments = payload.get("assignments")
            if not isinstance(assignments, dict):
                continue
            for resonator_number, record in assignments.items():
                if not isinstance(record, dict):
                    continue
                try:
                    target_hz = float(record.get("frequency_hz"))
                except Exception:
                    continue
                identifier = self._sheet_identifier_for_scan(scan, record)
                if identifier not in row_map:
                    row_map[identifier] = {}
                    row_order.append(identifier)
                resonator_label = str(resonator_number).strip()
                row_map[identifier][resonator_label] = target_hz
                resonator_numbers.add(resonator_label)
                assignment_count += 1

        if not row_map or not resonator_numbers:
            raise ValueError("No marked resonator assignments were found on the loaded VNA scans.")

        wb = Workbook()
        ws = wb.active
        ws.title = "Resonators"
        ws.cell(row=1, column=1, value="")
        ws.cell(row=2, column=1, value="Identifier")

        ordered_resonators = sorted(resonator_numbers, key=self._resonator_sort_key)
        for col_idx, resonator_label in enumerate(ordered_resonators, start=2):
            ws.cell(row=1, column=col_idx, value=resonator_label)
            ws.cell(row=2, column=col_idx, value="Frequency (Hz)")

        for row_idx, identifier in enumerate(row_order, start=3):
            ws.cell(row=row_idx, column=1, value=identifier)
            for col_idx, resonator_label in enumerate(ordered_resonators, start=2):
                value = row_map[identifier].get(resonator_label)
                if value is not None:
                    ws.cell(row=row_idx, column=col_idx, value=float(value))

        sheet_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(sheet_path)
        wb.close()

        self.dataset.processing_history.append(
            _make_event(
                "save_resonances_to_sheet",
                {
                    "sheet_path": str(sheet_path),
                    "sheet_name": str(ws.title),
                    "assignment_count": int(assignment_count),
                    "normalized_count": int(normalized_count),
                },
            )
        )
        self._mark_dirty()
        self._autosave_dataset()
        if normalized_count > 0:
            self._log(f"Normalized {normalized_count} marked resonator identifier(s) before sheet export.")
        self._log(f"Saved {assignment_count} resonator assignment(s) to {sheet_path}")
        return assignment_count


    def _load_resonances_from_sheet(self, sheet_path: Path) -> int:
        wb = load_workbook(sheet_path, data_only=True, read_only=True)
        try:
            ws = wb.active

            column_headers: dict[int, str] = {}
            for col_idx in range(2, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_idx).value
                header = self._sheet_resonator_label(cell)
                column_headers[col_idx] = header

            row_records: list[dict] = []
            warnings: list[str] = []
            assignment_count = 0
            for row_idx, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                if not row:
                    continue
                identifier_raw = row[0]
                if identifier_raw is None or str(identifier_raw).strip() == "":
                    continue
                identifier = str(identifier_raw).strip()
                freq_entries: dict[int, dict] = {}
                for col_offset, cell in enumerate(row[1:], start=2):
                    resonator_number = column_headers.get(col_offset, "")
                    if cell is None or str(cell).strip() == "":
                        freq_entries[col_offset] = None
                        continue
                    try:
                        target_value = float(cell)
                    except Exception:
                        warnings.append(f"Row {row_idx}, column {col_offset}: skipped non-numeric frequency cell {cell!r}.")
                        freq_entries[col_offset] = None
                        continue
                    try:
                        scan, target_hz = self._find_scan_for_sheet_identifier(identifier, target_value)
                    except Exception as exc:
                        warnings.append(f"Row {row_idx}, column {col_offset}: {exc}")
                        freq_entries[col_offset] = None
                        continue
                    if resonator_number:
                        payload = self._sheet_resonance_attachment(scan)
                        assignments = payload["assignments"]
                        assignments[str(resonator_number)] = {
                            "frequency_hz": float(target_hz),
                            "sheet_path": str(sheet_path),
                            "sheet_name": str(ws.title),
                            "row": int(row_idx),
                            "column": int(col_offset),
                            "identifier": identifier,
                        }
                        assignment_count += 1
                    else:
                        warnings.append(
                            f"Row 1, column {col_offset}: blank resonator number header; frequency was plotted but not attached."
                        )
                    freq_entries[col_offset] = {
                        "scan": scan,
                        "target_hz": target_hz,
                        "target_value": target_value,
                        "resonator_number": resonator_number,
                    }
                row_records.append(
                    {
                        "row_idx": row_idx,
                        "identifier": identifier,
                        "entries": freq_entries,
                    }
                )

            data_columns = [col for col in range(2, ws.max_column + 1)]
        finally:
            wb.close()

        if not row_records or not data_columns:
            raise ValueError("No loadable resonance entries were found in the selected spreadsheet.")

        self.dataset.processing_history.append(
            _make_event(
                "load_resonances_from_sheet",
                {
                    "sheet_path": str(sheet_path),
                    "sheet_name": str(ws.title),
                    "assignment_count": int(assignment_count),
                },
            )
        )
        self._mark_dirty()
        self._autosave_dataset()
        for warning in warnings[:20]:
            self._log(f"Sheet resonance load warning: {warning}")
        self._log(f"Loaded {assignment_count} resonator assignment(s) from {sheet_path}")
        return assignment_count


    def _collect_attached_resonance_rows(self) -> tuple[list[dict], dict[int, str], list[int], list[str], str]:
        self._normalize_sheet_resonance_identifiers()
        grouped_rows: dict[int, dict] = {}
        column_headers: dict[int, str] = {}
        warnings: list[str] = []
        sheet_labels: set[str] = set()

        for scan in self.dataset.vna_scans:
            payload = scan.candidate_resonators.get("sheet_resonances")
            if not isinstance(payload, dict):
                continue
            assignments = payload.get("assignments")
            if not isinstance(assignments, dict):
                continue
            for resonator_number, record in assignments.items():
                if not isinstance(record, dict):
                    continue
                try:
                    row_idx = int(record.get("row"))
                    col_idx = int(record.get("column"))
                    target_hz = float(record.get("frequency_hz"))
                except Exception:
                    warnings.append(f"{Path(scan.filename).name}: skipped malformed resonator marker record.")
                    continue
                identifier = str(record.get("identifier") or Path(scan.filename).name)
                sheet_path = str(record.get("sheet_path") or "")
                sheet_name = str(record.get("sheet_name") or "")
                if sheet_path or sheet_name:
                    sheet_labels.add(f"{Path(sheet_path).name} | {sheet_name}".strip(" |"))
                column_headers[col_idx] = str(resonator_number).strip()
                row_record = grouped_rows.setdefault(
                    row_idx,
                    {
                        "row_idx": row_idx,
                        "identifier": identifier,
                        "entries": {},
                    },
                )
                row_record["entries"][col_idx] = {
                    "scan": scan,
                    "target_hz": target_hz,
                    "target_value": target_hz,
                    "resonator_number": str(resonator_number).strip(),
                }

        if not grouped_rows or not column_headers:
            raise ValueError("No marked resonator assignments were found on the loaded VNA scans.")

        row_records = [grouped_rows[idx] for idx in sorted(grouped_rows)]
        data_columns = sorted(column_headers)
        source_label = "; ".join(sorted(label for label in sheet_labels if label))
        return row_records, column_headers, data_columns, warnings, source_label


    def _plot_resonance_rows(
        self,
        *,
        row_records: list[dict],
        column_headers: dict[int, str],
        data_columns: list[int],
        data_mode: str,
        warnings: list[str],
        source_label: str,
    ) -> list[Path]:
        if not row_records or not data_columns:
            raise ValueError("No plottable resonance entries are available.")

        out_dir = _dataset_dir(self.dataset) / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_sheet_resonance_plots"
        out_dir.mkdir(parents=True, exist_ok=True)
        saved_paths: list[Path] = []

        column_chunks = [data_columns[i : i + 3] for i in range(0, len(data_columns), 3)]
        half_window_hz = 0.5e6

        for page_idx, page_columns in enumerate(column_chunks, start=1):
            nrows = len(row_records)
            ncols = len(page_columns)
            fig, axes = plt.subplots(
                nrows,
                ncols,
                figsize=(5.2 * ncols, max(2.2 * nrows, 4.0)),
                squeeze=False,
                sharex="col",
            )

            column_ranges: dict[int, tuple[float, float]] = {}
            for col in page_columns:
                targets = []
                for row_record in row_records:
                    entry = row_record["entries"].get(col)
                    if isinstance(entry, dict):
                        targets.append(float(entry["target_hz"]))
                if targets:
                    column_ranges[col] = (min(targets) - half_window_hz, max(targets) + half_window_hz)
                else:
                    column_ranges[col] = (0.0, 1.0)

            for row_pos, row_record in enumerate(row_records):
                for col_pos, col in enumerate(page_columns):
                    ax = axes[row_pos][col_pos]
                    entry = row_record["entries"].get(col)
                    if entry is None:
                        ax.axis("off")
                        continue

                    scan = entry["scan"]
                    freq = np.asarray(scan.freq, dtype=float)
                    if data_mode == "normalized":
                        norm = scan.baseline_filter.get("normalized", {})
                        amp, _phase = _read_polar_series(
                            norm,
                            amplitude_key="norm_amp",
                            phase_key="norm_phase_deg_unwrapped",
                        )
                        if amp.shape == scan.freq.shape:
                            y = np.asarray(amp, dtype=float)
                            y_label = "Normalized |S21|"
                        else:
                            y = np.asarray(scan.amplitude(), dtype=float)
                            y_label = "|S21|"
                            warnings.append(
                                f"{Path(scan.filename).name}: normalized data missing; used raw amplitude instead."
                            )
                    else:
                        y = np.asarray(scan.amplitude(), dtype=float)
                        y_label = "|S21|"

                    order = np.argsort(freq)
                    freq_sorted = freq[order]
                    y_sorted = y[order]
                    col_lo, col_hi = column_ranges[col]
                    full_mask = (freq_sorted >= col_lo) & (freq_sorted <= col_hi)
                    if not np.any(full_mask):
                        ax.text(0.5, 0.5, "No data in column range", ha="center", va="center")
                        ax.axis("off")
                        continue

                    ax.plot(
                        freq_sorted[full_mask] / 1.0e9,
                        y_sorted[full_mask],
                        color="0.65",
                        linewidth=0.8,
                        zorder=1,
                    )

                    target_hz = float(entry["target_hz"])
                    local_lo = target_hz - half_window_hz
                    local_hi = target_hz + half_window_hz
                    local_mask = (freq_sorted >= local_lo) & (freq_sorted <= local_hi)
                    if np.any(local_mask):
                        ax.plot(
                            freq_sorted[local_mask] / 1.0e9,
                            y_sorted[local_mask],
                            color="tab:blue",
                            linewidth=2.2,
                            zorder=2,
                        )

                    show_labels = row_pos == 0 and col_pos == 0
                    cand = scan.candidate_resonators
                    gaussian_freqs = np.asarray(
                        cand.get("gaussian_convolution", {}).get("candidate_freq", np.array([])),
                        dtype=float,
                    )
                    dsdf_freqs = np.asarray(
                        cand.get("dsdf_gaussian_convolution", {}).get("candidate_freq", np.array([])),
                        dtype=float,
                    )
                    for freqs_hz, style, label in (
                        (
                            gaussian_freqs,
                            dict(
                                linestyle="none",
                                marker="o",
                                markersize=7,
                                markerfacecolor="none",
                                markeredgewidth=1.5,
                                color="green",
                            ),
                            "Gaussian candidates",
                        ),
                        (
                            dsdf_freqs,
                            dict(linestyle="none", marker="D", markersize=5, color="red"),
                            "dS21/df peaks",
                        ),
                    ):
                        in_win = freqs_hz[(freqs_hz >= col_lo) & (freqs_hz <= col_hi)]
                        if in_win.size == 0:
                            continue
                        y_pts = [self._interpolate_y(freq_sorted, y_sorted, float(f)) for f in in_win]
                        ax.plot(in_win / 1.0e9, y_pts, label=(label if show_labels else None), **style)

                    target_y = self._interpolate_y(freq_sorted, y_sorted, target_hz)
                    ax.plot(
                        [target_hz / 1.0e9],
                        [target_y],
                        linestyle="none",
                        marker="s",
                        markersize=7,
                        color="black",
                        label=("Spreadsheet frequency" if show_labels else None),
                    )
                    ax.set_xlim(col_lo / 1.0e9, col_hi / 1.0e9)
                    ax.grid(True, alpha=0.3)
                    ax.set_title(
                        f"{row_record['identifier']} | {Path(scan.filename).name}\n{target_hz / 1.0e9:.9g} GHz",
                        fontsize=8,
                    )
                    if col_pos == 0:
                        ax.set_ylabel(y_label)
                    if row_pos == nrows - 1:
                        ax.set_xlabel("Frequency (GHz)")
                    if show_labels:
                        ax.legend(loc="upper right", fontsize=8)

            for col_pos, col in enumerate(page_columns):
                resonator_number = column_headers.get(col, "")
                resonator_label = (
                    f"Resonator {resonator_number}"
                    if resonator_number
                    else f"Sheet column {col}"
                )
                axes[0][col_pos].set_title(
                    f"{resonator_label} | {column_ranges[col][0] / 1.0e9:.9g} to {column_ranges[col][1] / 1.0e9:.9g} GHz",
                    fontsize=9,
                )

            fig.suptitle(
                f"Resonator Marker Plots | mode={'baseline normalized' if data_mode == 'normalized' else 'raw'} | page {page_idx}"
                + (f" | {source_label}" if source_label else ""),
                fontsize=12,
            )
            fig.tight_layout()
            out_path = out_dir / f"resonance_sheet_page_{page_idx:02d}.pdf"
            fig.savefig(out_path)
            plt.close(fig)
            saved_paths.append(out_path)

        self.dataset.processing_history.append(
            _make_event(
                "plot_attached_resonances",
                {
                    "data_mode": data_mode,
                    "plot_count": int(
                        sum(sum(1 for entry in r["entries"].values() if isinstance(entry, dict)) for r in row_records)
                    ),
                    "page_count": len(saved_paths),
                    "output_dir": str(out_dir),
                    "source_label": source_label,
                },
            )
        )
        self._mark_dirty()
        self._autosave_dataset()
        for warning in warnings[:20]:
            self._log(f"Attached resonance plot warning: {warning}")
        return saved_paths


    def _plot_attached_resonances_draw_page(
        self,
        ax,
        rows: list[dict],
        *,
        spacing: float = 1.5,
        truncate_threshold: float = 1.5,
        xlim_ghz: Optional[tuple[float, float]] = None,
        title: str = "",
    ) -> int:
        offset_by_scan_key, tick_info = self._attached_resonance_editor_offset_map(rows, spacing)
        trace_colors = self._attached_resonance_editor_trace_colors()
        resonator_tracks: dict[str, list[tuple[float, float]]] = {}
        resonator_markers: list[dict] = []
        marker_count = 0
        x_range_ghz = xlim_ghz

        for row in rows:
            offset = float(offset_by_scan_key[str(row["scan_key"])])
            freq_ghz = np.asarray(row["freq"], dtype=float) / 1.0e9
            amp_display = np.minimum(np.asarray(row["amp"], dtype=float), truncate_threshold)
            y = amp_display + offset
            trace_color = trace_colors[int(row.get("scan_index", 0)) % len(trace_colors)]
            ax.plot(freq_ghz, y, linewidth=1.0, color=trace_color, alpha=0.9, zorder=1)

            for resonator in row["resonators"]:
                target_hz = float(resonator["target_hz"])
                target_ghz = target_hz / 1.0e9
                y_pt = self._interpolate_y(row["freq"], amp_display, target_hz) + offset
                resonator_number = str(resonator["resonator_number"])
                resonator_markers.append(
                    {
                        "resonator_number": resonator_number,
                        "x_ghz": target_ghz,
                        "y": y_pt,
                    }
                )
                resonator_tracks.setdefault(resonator_number, []).append((target_ghz, y_pt))
                marker_count += 1

        for resonator_number, points in resonator_tracks.items():
            if x_range_ghz is not None:
                points = [pt for pt in points if x_range_ghz[0] <= pt[0] <= x_range_ghz[1]]
            if len(points) < 2:
                continue
            points = sorted(points, key=lambda item: item[1], reverse=True)
            ax.plot(
                [pt[0] for pt in points],
                [pt[1] for pt in points],
                linestyle=":",
                linewidth=1.0,
                color="tab:red",
                alpha=0.9,
                zorder=2,
            )

        for point in resonator_markers:
            if x_range_ghz is not None and not (x_range_ghz[0] <= point["x_ghz"] <= x_range_ghz[1]):
                continue
            ax.plot(
                [point["x_ghz"]],
                [point["y"]],
                linestyle="none",
                marker="o",
                markersize=6,
                markerfacecolor="none",
                markeredgecolor="tab:red",
                markeredgewidth=1.5,
                zorder=4,
            )
            ax.text(
                point["x_ghz"],
                point["y"] - 0.18,
                point["resonator_number"],
                ha="center",
                va="top",
                fontsize=8,
                color="tab:red",
                zorder=5,
            )

        y_low = min(
            float(np.min(np.minimum(np.asarray(row["amp"], dtype=float), truncate_threshold)))
            + float(offset_by_scan_key[str(row["scan_key"])])
            for row in rows
        )
        y_high = max(
            float(np.max(np.minimum(np.asarray(row["amp"], dtype=float), truncate_threshold)))
            + float(offset_by_scan_key[str(row["scan_key"])])
            for row in rows
        )
        freq_min = min(float(row["freq"][0]) for row in rows) / 1.0e9
        freq_max = max(float(row["freq"][-1]) for row in rows) / 1.0e9
        x_pad = max((freq_max - freq_min) * 0.01, 1e-6)

        ax.set_xlabel("Frequency (GHz)")
        ax.set_ylabel("Normalized |S21| + vertical offset")
        ax.set_yticks([item[0] for item in tick_info])
        ax.set_yticklabels([item[1] for item in tick_info], fontsize=8)
        ax.set_ylim(y_low - 0.2, y_high + 0.2)
        if xlim_ghz is None:
            ax.set_xlim(freq_min - 0.5 * x_pad, freq_max + 2.0 * x_pad)
        else:
            ax.set_xlim(xlim_ghz)
        ax.grid(True, alpha=0.3)
        if title:
            ax.set_title(title, fontsize=11)
        return marker_count


    def _plot_attached_resonances(
        self,
        *,
        progress_callback: Optional[Callable[[int, int], None]] = None,
    ) -> list[Path]:
        rows, warnings = self._selected_scans_for_attached_resonance_editor()
        if not rows:
            raise ValueError("No selected scans with normalized data are available for resonator-marker plotting.")

        out_dir = _dataset_dir(self.dataset) / f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_resonator_marker_plots"
        out_dir.mkdir(parents=True, exist_ok=True)
        pdf_path = out_dir / "resonator_markers.pdf"

        freq_min_hz = min(float(row["freq"][0]) for row in rows)
        freq_max_hz = max(float(row["freq"][-1]) for row in rows)
        marker_freqs_hz = sorted(
            float(resonator["target_hz"])
            for row in rows
            for resonator in row["resonators"]
        )
        zoom_span_hz = 100.0e6
        zoom_step_hz = 80.0e6
        zoom_windows: list[tuple[float, float]] = []
        if marker_freqs_hz and freq_max_hz - freq_min_hz > zoom_span_hz:
            start_hz = freq_min_hz
            seen_starts: set[float] = set()
            while True:
                rounded_start = round(start_hz, 3)
                if rounded_start in seen_starts:
                    break
                seen_starts.add(rounded_start)
                end_hz = min(start_hz + zoom_span_hz, freq_max_hz)
                if any(start_hz <= marker_hz <= end_hz for marker_hz in marker_freqs_hz):
                    zoom_windows.append((start_hz, end_hz))
                if end_hz >= freq_max_hz:
                    break
                next_start = start_hz + zoom_step_hz
                if next_start + zoom_span_hz > freq_max_hz:
                    next_start = max(freq_min_hz, freq_max_hz - zoom_span_hz)
                if next_start <= start_hz:
                    break
                start_hz = next_start

        total_pages = 1 + len(zoom_windows)
        page_count = 0
        marker_count = 0
        with PdfPages(pdf_path) as pdf:
            fig = Figure(figsize=(12, 7))
            ax = fig.add_subplot(111)
            marker_count += self._plot_attached_resonances_draw_page(
                ax,
                rows,
                title="Resonator Markers | Full Span",
            )
            fig.tight_layout()
            pdf.savefig(fig)
            plt.close(fig)
            page_count += 1
            if progress_callback is not None:
                progress_callback(page_count, total_pages)

            for page_idx, (lo_hz, hi_hz) in enumerate(zoom_windows, start=1):
                fig = Figure(figsize=(12, 7))
                ax = fig.add_subplot(111)
                marker_count += self._plot_attached_resonances_draw_page(
                    ax,
                    rows,
                    xlim_ghz=(lo_hz / 1.0e9, hi_hz / 1.0e9),
                    title=(
                        f"Resonator Markers | Zoom {page_idx} | "
                        f"{lo_hz / 1.0e9:.6f} to {hi_hz / 1.0e9:.6f} GHz"
                    ),
                )
                fig.tight_layout()
                pdf.savefig(fig)
                plt.close(fig)
                page_count += 1
                if progress_callback is not None:
                    progress_callback(page_count, total_pages)

        self.dataset.processing_history.append(
            _make_event(
                "plot_attached_resonances",
                {
                    "page_count": int(page_count),
                    "plot_count": int(marker_count),
                    "output_pdf": str(pdf_path),
                    "zoom_span_mhz": 100.0,
                    "zoom_overlap_mhz": 20.0,
                },
            )
        )
        self._mark_dirty()
        self._autosave_dataset()
        for warning in warnings[:20]:
            self._log(f"Resonator marker plot warning: {warning}")
        self._log(f"Saved resonator marker PDF: {pdf_path}")
        return [pdf_path]
